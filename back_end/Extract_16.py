"""
增加获取文本信息功能
将文本信息构造成树
更新为解析xml文档的形式
v11版本整理了代码注释，删除了无用代码
v14版本将遍历节点单独设置为一个模块，增加了目录处理模块
v15版本完善了标题类型的确定机制
v16版本完善了表格信息处理
"""
import os
import shutil
import tkinter as tk
import zipfile
from tkinter import filedialog
import pandas as pd
import win32com.client as win32
import xml.dom.minidom
from win32com.client import constants  # visio信息提取
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from xml.dom.minidom import parse
import networkx as nx
import matplotlib.pyplot as plt
import cn2an  # 中文字符转阿拉伯数字
from pylab import *
import re

"""Node类，存储包括标题、文本、表格、图片4种信息"""


class Node:
    def __init__(self, id, name='', text='', father_id=[], father_text=[], son_id=[], son_text=[], level=0):
        # 下面为Node对象增加n个实例变量
        self.id = id  # 自己的ID
        self.name = name  # 节点的名称
        self.text = text  # 存储的文本信息
        self.father_id = father_id  # 父亲们的ID
        self.father_text = father_text  # 父亲们的名字
        self.son_id = son_id  # 儿子们的ID
        self.son_text = son_text  # 儿子们的名字
        self.level = level  # 所在层

    def node_info_print(self):  # 打印各节点信息
        print('')
        print("\033[1;40m\033[1;31m" + '节点id：' + str(self.id) + "\033[0m")
        print('节点title：' + "\033[1;" + str(30 + self.level) + "m" + str(self.name) + "\033[0m")
        print('节点文本信息：' + "\033[1;" + str(30 + self.level) + "m" + str(self.text) + "\033[0m")
        print('所在层：' + "\033[1;" + str(30 + self.level) + "m" + str(self.level) + "\033[0m")
        print('父节点id：' + str(self.father_id))
        print('父节点文本信息：' + str(self.father_text))
        print('子节点id：' + str(self.son_id))
        print('子节点文本信息：' + str(self.son_text))


'''打开选择文件夹对话框，读取docx文件，删除已经存在的zip以及解压文件夹'''


def Get_filepath():
    root = tk.Tk()
    root.withdraw()
    Filepath = filedialog.askopenfilename()  # 获得选择好的文件
    if Filepath == '':
        exit()
    else:
        Filepath = Filepath.replace("\\", "/")  # 路径字符串斜杠替换
    (Folderpath, Filename) = os.path.split(Filepath)  # 分离文件夹路径和文件名.后缀

    # 为了后续copy文件改为zip并解压，删除已经存在的同名文件
    (souce_name, souce_suffix) = os.path.splitext(Filename)  # 分离文件名和后缀
    if souce_suffix != '.docx':
        exit('请选择docx文档')
    extract_path = Folderpath + '/' + souce_name + '_copy'
    zip_path = Folderpath + '/' + souce_name + '_copy.zip'
    Delete_Copy(extract_path, zip_path)  # 删除该路径下的文件
    print('docx文件路径:', Filepath)

    return Filepath


"""将dataframe数据存入excel，居中，自动调整列宽，每列相同数据合并单元格"""


def to_excel(df, Filepath, stack_dfname):
    wb = Workbook()  # 创建一个workbook
    ws = wb.active  # 获取当前workbook的第一个worksheet，默认的索引值是0，它是可以改变的

    # 将每行数据写入ws中
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # 遍历ws.cell，设置居中
    for cell_i in range(df.shape[0] + 1):
        for cell_j in range(df.shape[1]):
            ws.cell(row=cell_i + 1, column=cell_j + 1).alignment = Alignment(horizontal='center', vertical='center')

    # 自动调整列宽
    for col in df.columns:
        index = list(df.columns).index(col)  # 列序号
        letter = get_column_letter(index + 1)  # 列字母
        collen = df[col].apply(lambda x: len(str(x).encode())).max()  # 获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
        if collen > 150:
            collen = 150  # 限宽
        ws.column_dimensions[letter].width = collen

    # 合并单元格
    for dfname in stack_dfname:
        if dfname == 'Normal':  # 不合并Normal列
            break
        dfname_index = stack_dfname.index(dfname)
        df_key = list(set(df[dfname].values))
        #   遍历第一列去重后id
        for i in df_key:
            # 获取id等于指定值的几行数据
            df_id = df[df.iloc[:, dfname_index] == i].index.tolist()  # 索引值从0开始
            ws.merge_cells(start_row=df_id[0] + 2, end_row=df_id[-1] + 2, start_column=dfname_index + 1,
                           end_column=dfname_index + 1)  # 序号从1开始，所以行序号需要加2

    # 存入excel
    (Folderpath, Filename) = os.path.split(Filepath)  # 分离文件夹路径和文件名.后缀
    (souce_name, souce_suffix) = os.path.splitext(Filename)  # 分离文件名和后缀
    excel_path = Folderpath + '/' + souce_name + '_info_extraction.xlsx'
    wb.save(excel_path)
    # 保存writer中的数据至excel
    # 如果省略该语句，则数据不会写入到上边创建的excel文件中
    print('导入成功！')


"""删除文件夹和zip压缩包"""


def Delete_Copy(extract_path, zip_path):
    # 删除解压文件夹
    if os.path.exists(extract_path):  # 如果文件存在
        # 删除文件夹，可使用以下两种方法。
        shutil.rmtree(extract_path)  # 执行删除文件夹
        print('\n已删除文件夹:' + extract_path)
    else:
        print('no such file:' + extract_path)  # 则返回文件不存在

    # 删除zip压缩包
    if os.path.exists(zip_path):  # 如果文件存在
        # 删除文件夹，可使用以下两种方法。
        os.remove(zip_path)  # 删除zip文件
        print('已删除压缩包:' + zip_path)
    else:
        print('no such file:' + zip_path)  # 则返回文件不存在


"""从docx文件中提取xml文档，输出解压前zip文件以及解压后文件夹路径"""


def Get_xmlmkdir(Filepath):
    # 分离文件夹路径、文件名、后缀
    (Folderpath, Filename) = os.path.split(Filepath)  # 分离文件夹路径和文件名.后缀
    (souce_name, souce_suffix) = os.path.splitext(Filename)  # 分离文件名和后缀

    # 克隆一份docx文件，用于改后缀并解压
    os.chdir(Folderpath)
    app = win32.Dispatch('Word.Application')
    source_doc = app.Documents.Open(Filepath)  # 源docx文件路径
    source_doc.Content.Copy()
    copy_name = 'copy'
    copy_path = Folderpath + '/' + souce_name + '_' + copy_name + '.docx'  # copy的docx文件路径
    source_doc.SaveAs(copy_path)  # 另存
    source_doc.Close()

    # 该后缀改成zip文件
    zip_path = Folderpath + '/' + souce_name + '_' + copy_name + '.zip'  # zip路径
    os.rename(copy_path, zip_path)  # 重命名为zip文件

    # 创建新文件夹并把zip内容解压进来
    extract_path = Folderpath + '/' + souce_name + '_' + copy_name
    os.mkdir(extract_path)  # 创建目录
    f = zipfile.ZipFile(zip_path, 'r')  # 进行解压
    for file in f.namelist():
        f.extract(file, extract_path)

    # 输出解压信息
    print('\n源docx文件路径：' + Filepath)
    print('克隆docx文件路径：' + copy_path)
    print('zip文件路径：' + zip_path)
    print('解压文件夹：' + extract_path)
    print('')

    return extract_path, zip_path, Folderpath, souce_name


"""输入图片存放路径获取图片信息"""


def Get_vsdx_from_path(picture_path):
    (Folderpath, Filename) = os.path.split(picture_path)  # 分离文件夹路径和文件名.后缀
    (souce_name, souce_suffix) = os.path.splitext(Filename)  # 分离文件名和后缀
    appVisio = win32.gencache.EnsureDispatch("Visio.Application")
    if souce_suffix == '.vsdx':
        vdoc = appVisio.Documents.Open(picture_path)  # 打开vsdx文档
        connects = vdoc.Pages(1).Connects
        text = []  # 初始化text列表，存储visio图片中的每一条边信息
        for conn in connects:  # 遍历每一条边
            from_shp = conn.FromSheet
            info = from_shp.Name
            if conn.FromPart == constants.visBegin:
                info += ', ' + '起点'
            elif conn.FromPart == constants.visEnd:
                info += ', ' + '终点'  # visBegin ==9 ,visEnd == 12
            to_shp = conn.ToSheet
            info += ', ' + to_shp.Text  # 连接的矩形名称
            text.append(info)
        vdoc.Close()  # 关闭vsdx文档
    else:
        text = '非vsdx文档无法读取'
    return text


"""在输入的字符串S中，找到字符串str_begin与字符串str_end之间包含的所有文本的并集"""
def Get_string(S, str_begin, str_end):
    pattern = [str_begin, '(.+?)', str_end]
    pattern = ''.join(pattern)
    ret = re.findall(pattern, S)  # 非贪婪
    text_info = ''.join(ret)
    L_str = len(text_info)
    return text_info

"""在输入的字符串S中，找到字符串str_begin与字符串str_end之间包含的所有文本的并集"""
def Get_string_2node(S, str_begin, str_end, str_begin2, str_end2):
    str_len = len(str_begin) + len(str_end)  # 俩字符串的总长度
    text_info = ''  # 初始化文本信息为空
    index_begin = []  # 记录每个文本信息开始位置的索引（文本信息可能分布在字符串的各个位置）
    index_end = []  # 记录每个文本信息结束位置的索引（文本信息可能分布在字符串的各个位置）
    if len(S) >= str_len:  # 只有总长度大于节点字符长度之和才能有可能包含信息
        str_index_begin = 0
        while str_index_begin < (len(S) - str_len + 1):  # 对字符串S的遍历范围
            S_tmp = S[str_index_begin:str_index_begin + len(str_begin)]
            S_tmp2 = S[str_index_begin:str_index_begin + len(str_begin2)]
            if S_tmp == str_begin:  # 与第一个字符串匹配
                index_begin.append(str_index_begin)  # 文本信息开始位置的索引
                str_index_end = str_index_begin + len(str_begin)
                while str_index_end < len(S):
                    S_tmp = S[str_index_end:str_index_end + len(str_end)]
                    text_info = text_info + S[str_index_end]
                    if S_tmp == str_end:
                        index_end.append(str_index_end)  # 文本信息结束位置的索引
                        text_info = text_info[:-1]  # 删除文本信息的最后一个字符
                        str_index_begin = str_index_end  # 开启下一个文本信息
                        break
                    str_index_end = str_index_end + 1

            elif S_tmp2 == str_begin2:  # 与第二个字符串匹配
                index_begin.append(str_index_begin)  # 文本信息开始位置的索引
                str_index_end = str_index_begin + len(str_begin2)
                while str_index_end < len(S):
                    S_tmp2 = S[str_index_end:str_index_end + len(str_end2)]
                    text_info = text_info + S[str_index_end]
                    if S_tmp2 == str_end:
                        index_end.append(str_index_end)  # 文本信息结束位置的索引
                        text_info = text_info[:-1]  # 删除文本信息的最后一个字符
                        str_index_begin = str_index_end  # 开启下一个文本信息
                        break
                    str_index_end = str_index_end + 1

            str_index_begin = str_index_begin + 1
    return text_info


"""从字符串S里面获取第一个x字符串等号后面的信息（以字符"为结尾）"""
def Get_info_from_string(S, x):
    len_x = len(x)
    len_S = len(S)
    target = ''  # 初始化被提取的信息
    tip = True
    for S_index in range(len_S - len_x + 1):  # 按照固定模式的匹配，比较死板，暂时没有更好的方法
        tmp = S[S_index:S_index + len_x]
        if tmp == x and S[S_index + len_x] == '=' and S[S_index + len_x + 1] == '"':
            S_index = S_index + len_x + 2
            begin_index = S_index
            while 1:
                if S[S_index + 1] == '"':
                    end_index = S_index
                    break
                S_index = S_index + 1
            target = S[begin_index:end_index + 1]
            tip = False
            break
    if tip:  # 如果该字符串中没有能被匹配到的信息
        print(x + '未能匹配上')
    return target


"""在指定路径（extract_path）下找到对应的visio图片中包含rid的节点下的文本"""


def Get_picture_info(rid, extract_path):
    rels_path = extract_path + '/word/_rels/document.xml.rels'
    text = ''
    if os.path.exists(rels_path):  # 如果文件存在
        DOMTree = xml.dom.minidom.parse(rels_path)
        collection = DOMTree.documentElement
        relations = collection.getElementsByTagName("Relationship")
        for relation in relations:
            if relation.hasAttribute("Id"):
                if relation.getAttribute("Id") == rid:
                    Target_path = relation.getAttribute("Target")
                    Picture_path = extract_path + '/word/' + Target_path
                    Picture_path = Picture_path.replace("\\", "/")  # 路径字符串斜杠替换
                    print('提取visio信息，vsdx文档路径：' + Picture_path)
                    text = Get_vsdx_from_path(Picture_path)  # 根据图片路径，提取图片中的边信息
                    # print(text)
                    break
    else:
        print(rels_path + '的文件不存在！！！')
    return text


"""获取前缀标识符"""
def Get_prefix_level(S):
    # 将中文部分转为阿拉伯数字
    ZH = '零一二三四五六七八九十'
    separator = '.、\/-'
    i = 0
    tmp_run = ''
    all_run = []
    prefix_level = 0  # 前缀的层次
    while i < len(S):
        if S[i] in ZH:
            while S[i] in ZH:
                tmp_run = tmp_run + S[i]
                i = i + 1
                if i >= len(S):
                    break
            all_run.append(tmp_run)
            tmp_run = ''
        else:
            while S[i] not in ZH:
                tmp_run = tmp_run + S[i]
                i = i + 1
                if i >= len(S):
                    break
            all_run.append(tmp_run)
            tmp_run = ''
    #print(all_run)

    for i in range(len(all_run)):
        if all_run[i][0] in ZH:  # 如果是数字，转换为阿拉伯字符
            all_run[i] = str(cn2an.cn2an(all_run[i], "normal"))
        else:
            if len(all_run[i]) > 1:
                break
    #print(all_run)
    Arabic_S = ''.join(all_run)
    #print(Arabic_S)
    prefix = ''
    first_in = True
    for i in range(len(Arabic_S)):
        if not Arabic_S[i].isdigit() and Arabic_S[i] not in separator:
            break
        if Arabic_S[i].isdigit() and first_in:
            prefix_level = prefix_level + 1
            first_in = False
        if Arabic_S[i] in separator:
            first_in = True
        prefix = prefix + Arabic_S[i]
    if len(Arabic_S) - len(prefix) > 30 and prefix_level > 0:  # 太长了，就不认为是标题了
        print('文本过长，不认为是标题')
        prefix_level = 0
    prefix_level = str(prefix_level)
    return prefix_level


"""获取节点标识符和前缀标识符，用于是否是标题的判定"""
def Get_node_prefix_identifier(Node):
    identifier = []
    node_pStyle = Get_string(Node, '<w:pStyle w:val="', '"/>')  # pStyle信息
    node_outlineLvl = Get_string(Node, '<w:outlineLvl w:val="', '"/>')  # outlineLvl信息
    node_ilvl = Get_string(Node, '<w:ilvl w:val="', '"/>')  # outlineLvl信息
    node_numId = Get_string(Node, '<w:numId w:val="', '"/>')  # outlineLvl信息
    node_identifier = node_pStyle #+ node_outlineLvl
    if node_identifier != '':  # 如果节点存在pStyle信息和outlineLvl信息，才被认为是标题，才会附加ilvl信息和numId信息
        node_identifier = node_identifier #+ node_ilvl + node_numId
    if node_identifier != '':  # 存在节点标识符的情况下，才提取前缀信息
        text = Get_string_2node(Node, '<w:t>', '</w:t>', '<w:t xml:space="preserve">', '</w:t>')
        if text != '':  # 空字符串不能成为标题
            prefix_identifier = Get_prefix_level(text)
            identifier.append(node_identifier)
            identifier.append(prefix_identifier)
    return identifier


"""确定xml中标题的层级，通过关键字符串pStyle来判断结点是否是标题以及结点等级"""
def Get_heading_level(S_cut):
    dom_needed = ['<w:p', '<w:tbl>', '<w:sdt>']
    all_identifier = []  # 记录所有的pStyle，用于判断标题的级别
    first_identifier = True
    level = 1
    for i in range(len(S_cut)):
        Node = S_cut[i][0]
        Node_type = S_cut[i][1]
        # 处理扫描到的节点类型，记录其中的内容
        if Node_type == dom_needed[0]:  # 用于处理文本和图形信息
            identifier = Get_node_prefix_identifier(Node)
            if identifier != []:
                #print(identifier)
                if first_identifier:
                    all_identifier.append([identifier, 'Heading ' + str(level)])
                    first_identifier = False
                else:
                    nofind = True
                    for j in range(len(all_identifier)):
                        tmp_node_identifier = all_identifier[j][0][0]
                        tmp_prefix_identifier = all_identifier[j][0][1]
                        if identifier[0] == tmp_node_identifier or (identifier[1] == tmp_prefix_identifier and identifier[1] != '0'):
                            # 只要有一个标识符相等，则认为是同级标题，同时，如果前缀标识符为'0'说明根本没有前缀，因此也没有用前缀来判定的必要
                            all_identifier.append([identifier, all_identifier[j][1]])
                            nofind = False
                            break
                    if nofind:
                        level = level + 1
                        all_identifier.append([identifier, 'Heading ' + str(level)])

    for each_identifier in all_identifier:
        print(each_identifier)

    return all_identifier, level


"""遍历xml字符串，对标题、文本、图像、表格分类，存入TreeList列表中"""
def Traverse_dom(S_cut, extract_path):
    dom_needed = ['<w:p', '<w:tbl>', '<w:sdt>']
    text = ''
    title = ''

    # 初始化树根节点
    tip = False  # 判断是否找到可记录的文本信息
    ID = 0
    Tree = Node(id=ID, name='Root', text='根节点', level=0)
    ID = ID + 1
    TreeList = []
    TreeList.append(Tree)

    all_identifier, level = Get_heading_level(S_cut)  # 记录所有的pStyle，用于判断标题的级别

    # 标题的层级是不确定的，但是文本（Normal）、表格（Table）、图片（Picture）是确定的
    stack_dfname = []
    for i in range(level):
        stack_dfname.append('Heading ' + str(i + 1))
    stack_dfname.append('Normal')
    stack_dfname.append('Table')
    stack_dfname.append('Picture')

    # 遍历xml字符串，对各结点进行分类
    for i in range(len(S_cut)):
        S_each = S_cut[i][0]
        S_type = S_cut[i][1]
        title = 'nofind'

        # 处理扫描到的节点类型，记录其中的内容
        if S_type == dom_needed[0]:  # 用于处理文本和图形信息
            if '</v:shape>' in S_each:  # 该结点为图形节点
                text = Get_string(S_each, '<o:OLEObject', '/>')
                rid = Get_string(text, 'r:id="', '"')  # 提取图片rid
                title = 'Picture'
                text = Get_picture_info(rid, extract_path)
            else:  # 按文本处理
                text = Get_string_2node(S_each, '<w:t>', '</w:t>', '<w:t xml:space="preserve">', '</w:t>')
                identifier = Get_node_prefix_identifier(S_each)
                if identifier != []:  # 没有找到标题特征则定义为Normal
                    for each_identifier in all_identifier:
                        if each_identifier[0] == identifier:
                            title = each_identifier[1]
                            break
                else:
                    title = 'Normal'
            tip = True
        elif S_type == dom_needed[1]:  # 用于处理表格信息
            text = Get_string_2node(S_each, '<w:t>', '</w:t>', '<w:t xml:space="preserve">', '</w:t>')

            tr_cut = Cut_xml_node(S_each, [['<w:tr', '</w:tr>']])

            title = 'Table'
            tip = True

        # 在上面提取到了新的text和title，则可存入TreeList中
        if tip:
            if text != '':  # 仅对非空字符构造
                if title not in stack_dfname:
                    print('存在未定义title')
                if 'Heading' in title:
                    current_level = stack_dfname.index(title) + 1
                elif title == 'Table':
                    current_level = len(stack_dfname) - 2 + 1
                else:
                    current_level = len(stack_dfname) - 3 + 1  # 将Normal，Table，Picture归为一层
                Leaf = Node(id=ID, name=title, text=text, father_id=[], father_text=[], son_id=[], son_text=[],
                            level=current_level)
                ID = ID + 1
                for j in range(len(TreeList)):
                    father_index = len(TreeList) - j - 1
                    if TreeList[father_index].level <= current_level - 1:  # 找到父节点在列表中所在位置，上一层应该是标题
                        Leaf.father_id = TreeList[father_index].id
                        Leaf.father_text = TreeList[father_index].text
                        TreeList[father_index].son_id.append(Leaf.id)
                        TreeList[father_index].son_text.append(Leaf.text)
                        break
                TreeList.append(Leaf)  # 将节点压入列表中
            tip = False

    return TreeList


def Cut_xml_node(S, dom_needed):  # 分割整个xml字符串
    S_cut = []
    S_index = 0  # 用于遍历字符串S
    String_tmp = ''  # 临时记录字符串
    start_record = False
    current_dom = ''
    #dom_needed = ['<w:p', '<w:tbl>', '<w:sdt>']

    while S_index < len(S):
        if S[S_index] == '<' and start_record == False:
            start_record = True  # 遇见该符号开始记录
        elif S[S_index] == ' ' or S[S_index] == '\n':
            start_record = False  # 遇见该符号结束记录
            current_dom = String_tmp
            String_tmp = ''
        elif S[S_index] == '<' and start_record == True:
            current_dom = String_tmp
            String_tmp = ''
        if start_record:
            String_tmp = String_tmp + S[S_index]

        # 处理扫描到的节点类型，记录其中的内容
        for dom in dom_needed:
            if current_dom == dom[0]:  # 用于处理文本和图形信息
                end_dom = dom[1]
                end_index = S_index
                while end_index < len(S) - len(end_dom):  # 找到该结点终止的索引
                    tmp_dom = S[end_index:end_index + len(end_dom)]
                    if tmp_dom == end_dom:
                        end_index = end_index + len(end_dom)
                        break
                    end_index = end_index + 1
                S_input = S[S_index - len(dom[0]):end_index]  # 起始索引和终止索引中间包含的字串
                S_cut.append([S_input, current_dom])
                S_index = end_index - 1

        current_dom = ''
        S_index = S_index + 1
    return S_cut


"""与Get_string函数的区别在于他不合并找到的文本，将找到的多个文本存在列表中"""
def Get_string_separated(S, str_begin, str_end):
    str_len = len(str_begin) + len(str_end)  # 俩字符串的总长度
    text_info = ''  # 初始化文本信息为空
    index_begin = []  # 记录每个文本信息开始位置的索引（文本信息可能分布在字符串的各个位置）
    index_end = []  # 记录每个文本信息结束位置的索引（文本信息可能分布在字符串的各个位置）
    all_text = []
    if len(S) >= str_len:  # 只有总长度大于节点字符长度之和才能有可能包含信息
        str_index_begin = 0
        while str_index_begin < (len(S) - str_len + 1):  # 对字符串S的遍历范围
            S_tmp = S[str_index_begin:str_index_begin + len(str_begin)]
            if S_tmp == str_begin:  # 与第一个字符串匹配
                index_begin.append(str_index_begin)  # 文本信息开始位置的索引
                str_index_end = str_index_begin + len(str_begin)
                while str_index_end < len(S):
                    S_tmp = S[str_index_end:str_index_end + len(str_end)]
                    text_info = text_info + S[str_index_end]
                    if S_tmp == str_end:
                        index_end.append(str_index_end)  # 文本信息结束位置的索引
                        text_info = text_info[:-1]  # 删除文本信息的最后一个字符
                        all_text.append(text_info)
                        text_info = ''  # 重新置为空字符串
                        str_index_begin = str_index_end  # 开启下一个文本信息
                        break
                    str_index_end = str_index_end + 1
            str_index_begin = str_index_begin + 1
    return all_text


"""分离目录节点"""
def Separating_directory_node(S_cut):
    i = 0
    while i < len(S_cut):
        node = S_cut[i][0]
        node_type = S_cut[i][1]

        # 针对<w:p结点，如果结点中文本仅有目录两字，则将该结点删除
        if node_type == '<w:p':
            text = Get_string_2node(node, '<w:t>', '</w:t>', '<w:t xml:space="preserve">', '</w:t>')
            hyperlink = Get_string(node, '<w:hyperlink', '</w:hyperlink>')
            if text == '目录':
                S_cut.pop(i)
                i = i - 1
            elif hyperlink != '':  # 针对<w:p结点，存在hyperlink结点，并且如果结点中的文本有多段，且最后一段是数字，那么认为这也是目录
                text_info = Get_string_separated(hyperlink, '<w:t>', '</w:t>')
                if len(text_info) >= 2 and text_info[-1].isdigit():  # 有多段文本，且最后一段是数字
                    S_cut.pop(i)
                    i = i - 1

        # 针对<w:sdt>结点，如果存在该结点并且里面也存在sdtContent结点，则删除
        if node_type == '<w:sdt>':
            sdtContent = Get_string(node, '<w:sdtContent>', '</w:sdtContent>')
            if sdtContent != '':
                S_cut.pop(i)
                i = i - 1

        i = i + 1
    return S_cut


"""进入解压后文件夹，读取xml文档，将文档信息转化为字符串"""
def Parsing_xml(extract_path):
    Parsing_path = extract_path + '/word/document.xml'
    DOMTree = parse(Parsing_path)
    type(DOMTree)
    domlist = DOMTree.documentElement
    node_name = '<w:body>'
    string = domlist.toxml()
    S = Get_string(string, node_name, '</w:body>')  # 仅针对body中的信息
    dom_needed = [['<w:p', '</w:p>'], ['<w:tbl>', '</w:tbl>'], ['<w:sdt>', '</w:sdt>']]
    S_cut = Cut_xml_node(S, dom_needed)  # 分割整个xml字符串
    S_cut = Separating_directory_node(S_cut)
    TreeList = Traverse_dom(S_cut, extract_path)  # 读取xml字符串，将各类结点存入列表中
    return TreeList


"""打印列表各节点"""
def Print_TreeList(TreeList):
    for each_node in TreeList:  # 遍历节点
        each_node.node_info_print()  # 打印各节点信息


"""将结点信息存入dataframe中"""
def TreeList_to_df(TreeList):
    dfname = ['节点id', 'Title', '文本信息', '层级', '父节点id', '父节点文本信息', '子节点id', '子节点文本信息']
    df = pd.DataFrame(columns=dfname)  # 初始化dataframe
    index = 0
    for node in TreeList:
        if node.name == 'Root':  # 不存根节点
            continue
        tmp_list = []
        tmp_list.append(str(node.id))
        tmp_list.append(str(node.name))
        tmp_list.append(str(node.text))
        tmp_list.append(str(node.level))
        tmp_list.append(str(node.father_id))
        tmp_list.append(str(node.father_text))
        if node.son_id == []:
            tmp_list.append('')
        else:
            tmp_list.append(str(node.son_id))
        if node.son_text == []:
            tmp_list.append('')
        else:
            tmp_list.append(str(node.son_text))
        df.loc[index] = tmp_list
        index = index + 1
    return df


"""将dataframe数据存入excel，居中，自动调整列宽，无合并单元格"""
def to_excel_nomerger(df, Filepath):
    wb = Workbook()  # 创建一个workbook
    ws = wb.active  # 获取当前workbook的第一个worksheet，默认的索引值是0，它是可以改变的

    # 将每行数据写入ws中
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    # 遍历ws.cell，设置居中
    for cell_i in range(df.shape[0] + 1):
        for cell_j in range(df.shape[1]):
            ws.cell(row=cell_i + 1, column=cell_j + 1).alignment = Alignment(horizontal='center', vertical='center')

    # 自动调整列宽
    for col in df.columns:
        index = list(df.columns).index(col)  # 列序号
        letter = get_column_letter(index + 1)  # 列字母
        collen = df[col].apply(lambda x: len(str(x).encode())).max()  # 获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
        if collen > 150:
            collen = 150  # 限宽
        elif collen < 20:
            collen = 20
        ws.column_dimensions[letter].width = collen

    # 存入excel
    (Folderpath, Filename) = os.path.split(Filepath)  # 分离文件夹路径和文件名.后缀
    (souce_name, souce_suffix) = os.path.splitext(Filename)  # 分离文件名和后缀
    excel_path = Folderpath + '/' + souce_name + '_info_extraction.xlsx'
    if os.path.exists(excel_path):  # 如果文件存在
        # 删除文件夹，可使用以下两种方法。
        os.remove(excel_path)  # 删除zip文件
        print('\n已删除旧excel')
    wb.save(excel_path)
    # 保存writer中的数据至excel
    print('已将节点信息导入excel  ' + excel_path)


"""传入树列表生成一个图"""
def Structural_network(TreeList):
    fig, ax = plt.subplots()
    G = nx.DiGraph()  # 无多重边有向图
    max_level = 0
    mycolor = ['#ff0000', 'dodgerblue', 'gold', 'limegreen', 'c', 'lightpink', 'orange', 'violet']
    all_color = []
    node_info = {}
    all_pos = {}
    for node in TreeList:
        if node.level > max_level:
            max_level = node.level
    for level in range(max_level + 1):
        tmpG = nx.DiGraph()
        if level < len(mycolor):
            current_color = mycolor[level]
        else:
            current_color = 'grey'
        for node in TreeList:
            if node.level == level:
                G.add_node(node.id, name=node.name, id=node.id)
                tmpG.add_node(node.id, name=node.name, id=node.id)
                all_color.append(current_color)
                node_info[node.id] = [node.name, node.text, node.level, node.father_id, node.son_id]
        R = level + pow(1.5, level + 1)  # 设置同心圆半径
        if R > 15:
            R = 15
        pos = nx.circular_layout(tmpG, scale=R)
        if len(pos) == 1:
            for key in pos:
                pos[key] = [R, 0]
        if level == 0:
            all_pos = pos
        else:
            all_pos.update(pos)
    for node in TreeList:
        begin_id = node.id
        all_end_id = node.son_id
        if all_end_id != []:
            for i, end_id in enumerate(all_end_id):
                G.add_edges_from([(begin_id, end_id)])
    nx.draw(G, all_pos, node_size=150, node_color=all_color, alpha=0.5)
    node_labels = nx.get_node_attributes(G, 'id')
    nx.draw_networkx_labels(G, all_pos, labels=node_labels, font_size=6)

    po_annotation = []
    for key in all_pos:
        # 标注点的坐标
        point_x = all_pos[key][0]
        point_y = all_pos[key][1]
        point, = plt.plot(point_x, point_y, '.', c='red', markersize=0.1, alpha=0.1)
        # 标注plt.annotate
        if node_info[key][3] == []:
            father_id = '无'
        else:
            father_id = str(node_info[key][3])
        if node_info[key][4] == []:
            son_id = '无'
        else:
            son_id = str(node_info[key][4])
        display_info = 'title:' + node_info[key][0] + '\ntext:' + str(node_info[key][1]) + '\nlevel:' + str(
            node_info[key][2]) + '\nfather_id:' + father_id + '\nson_id:' + son_id
        annotation = plt.annotate(display_info, xy=(point_x, point_y), size=7,
                                  bbox=dict(boxstyle='round,pad=0.5', fc='yellow', ec='k', lw=1, alpha=0.3))
        # 默认鼠标未指向时不显示标注信息
        annotation.set_visible(False)
        po_annotation.append([point, annotation])

    def on_move(event):
        visibility_changed = False
        for point, annotation in po_annotation:
            should_be_visible = (point.contains(event)[0] == True)
            if should_be_visible != annotation.get_visible():
                visibility_changed = True
                annotation.set_visible(should_be_visible)
        if visibility_changed:
            plt.draw()

    on_move_id = fig.canvas.mpl_connect('motion_notify_event', on_move)
    mpl.rcParams['font.sans-serif'] = ['SimHei']  # ''' 添加中文字体'''
    plt.rcParams['axes.unicode_minus'] = False  # ''' 防止负号无法正常显示  ''''
    plt.show()


def Output_to_txt(TreeList, Folderpath, souce_name):
    print(Folderpath)
    txt_path = Folderpath + '/' + souce_name + '.txt'

    # 删除存在的txt文档
    if os.path.exists(txt_path):  # 如果文件存在
        # 删除文件夹，可使用以下两种方法。
        os.remove(txt_path)  # 删除zip文件
        print('已删除文档:' + txt_path)
    else:
        print('no such file:' + txt_path)  # 则返回文件不存在

    f = open(txt_path, 'w+', encoding='utf-8')
    for node in TreeList:
        if node.level > 0:
            suojin = ''
            i = 1
            while i < node.level:
                suojin = suojin + '    ' + '    '
                i = i + 1
            f.write(suojin + str(node.text) + '\n')
    f.close()


if __name__ == '__main__':
    # Filepath = Get_filepath()  # 获取docx文件路径

    Filepath = "D:/master/master1/CEvaluation/uploads/cu/docs/测试说明标准大纲.docx"
    (Folderpath, Filename) = os.path.split(Filepath)  # 分离文件夹路径和文件名.后缀
    # 为了后续copy文件改为zip并解压，删除已经存在的同名文件
    (souce_name, souce_suffix) = os.path.splitext(Filename)  # 分离文件名和后缀
    if souce_suffix != '.docx':
        exit('请选择docx文档')
    extract_path = Folderpath + '/' + souce_name + '_copy'
    zip_path = Folderpath + '/' + souce_name + '_copy.zip'
    Delete_Copy(extract_path, zip_path)  # 删除该路径下的文件
    print('docx文件路径:', Filepath)

    extract_path, zip_path, Folderpath, souce_name = Get_xmlmkdir(Filepath)  # 将docx的克隆文件解压成含xml文件夹
    exit(0)
    TreeList = Parsing_xml(extract_path)  # 将xml字符串中的各结点信息存入列表中
    Print_TreeList(TreeList)  # 打印列表中各结点
    df = TreeList_to_df(TreeList)  # 将结点信息存入dataframe中
    to_excel_nomerger(df, Filepath)
    Output_to_txt(TreeList, Folderpath, souce_name)  # 生成txt文件
    Structural_network(TreeList)  # 生成图

    # Delete_Copy(extract_path, zip_path)  # 删除解压文件夹和zip压缩包

    #D:\master\master1\cgui\标准大纲\测试报告标准大纲.docx
