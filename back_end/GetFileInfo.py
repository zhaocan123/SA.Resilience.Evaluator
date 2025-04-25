import sys
import hashlib
from openpyxl import Workbook
import copy


def getfile(oldproject, newproject):
    '''
    比较文件级信息
    :param oldproject: 旧项目信息
    :param newproject: 新项目信息
    :return:
    '''
    olddata = dict()  # 文件树信息
    newdata = dict()
    oldhash = dict()
    newhash = dict()
    oldfileinfo = oldproject['codeFileInfo']
    oldfilelist = oldproject['projectInfo']['fileTypes']['c']
    oldfilelist.extend(oldproject['projectInfo']['fileTypes']['header'])
    oldfilelist.extend(oldproject['projectInfo']['fileTypes']['cpp'])
    newfileinfo = newproject['codeFileInfo']
    newfilelist = newproject['projectInfo']['fileTypes']['c']
    newfilelist.extend(newproject['projectInfo']['fileTypes']['header'])
    newfilelist.extend(newproject['projectInfo']['fileTypes']['cpp'])
    get_file_tree(oldfileinfo, olddata, oldhash, oldfilelist)
    get_file_tree(newfileinfo, newdata, newhash, newfilelist)
    numlist = check_md5(oldhash, newhash, olddata, newdata)
    # print("olddata",olddata)
    # print("oldmd5",oldhash)
    # print("newdata",newdata)
    # print("newmd5",newhash)
    filelevel = {}
    filelevel['updateFileNum'] = numlist[0]
    filelevel['deleteFileNum'] = numlist[1]
    filelevel['insertFileNum'] = numlist[2]
    filelevel['oldexceldata'] = olddata
    filelevel['newexceldata'] = newdata
    oldtree = tree_json(olddata)
    get_tree_id(oldtree)
    newtree = tree_json(newdata)
    get_tree_id(newtree)
    filelevel['selectedFileTree'] = oldtree
    filelevel['latestFileTree'] = newtree

    return filelevel


def getfileCPP(oldproject, newproject):
    '''
    比较文件级信息
    :param oldproject: 旧项目信息
    :param newproject: 新项目信息
    :return:
    '''
    olddata = dict()  # 文件树信息
    newdata = dict()
    oldhash = dict()
    newhash = dict()
    oldfileinfo = oldproject['codeFileInfo']
    oldfilelist = oldproject['projectInfo']['fileTypes']['cpp']
    newfileinfo = newproject['codeFileInfo']
    newfilelist = newproject['projectInfo']['fileTypes']['cpp']
    get_file_treeCPP(oldfileinfo, olddata, oldhash, oldfilelist)
    get_file_treeCPP(newfileinfo, newdata, newhash, newfilelist)
    numlist = check_md5(oldhash, newhash, olddata, newdata)
    # print("olddata",olddata)
    # print("oldmd5",oldhash)
    # print("newdata",newdata)
    # print("newmd5",newhash)
    filelevel = {}
    filelevel['updateFileNum'] = numlist[0]
    filelevel['deleteFileNum'] = numlist[1]
    filelevel['insertFileNum'] = numlist[2]
    filelevel['oldexceldata'] = olddata
    filelevel['newexceldata'] = newdata
    oldtree = tree_jsonCPP(olddata)
    get_tree_id(oldtree)
    newtree = tree_jsonCPP(newdata)
    get_tree_id(newtree)
    filelevel['selectedFileTree'] = oldtree
    filelevel['latestFileTree'] = newtree

    return filelevel


def generate_tree(tree, nodelist):
    if nodelist[0] not in tree.keys():
        tree[nodelist[0]] = dict()
    if len(nodelist) > 1:
        return generate_tree(tree[nodelist[0]], nodelist[1:])
    else:
        return tree[nodelist[0]]


def get_file_tree(fileinfo, tree, hashdict, filelist):
    '''
    生成文件数
    :param fileinfo: 文件信息
    :param tree: 空字典
    :param hashdict: md5
    :return:
    '''
    for filepath in fileinfo.keys():
        if filepath not in filelist:
            continue

        tmp = filepath.split('/')
        idx = tmp.index("code")
        lens = len(tmp)
        # md5 = getmd5(filepath)
        # hashdict[filepath] = md5
        tmppath = ''
        for k in range(idx+2, lens):
            tmppath = tmppath + tmp[k] + '/'
        tmppath = tmppath[:-1]
        md5 = getmd5(filepath)
        hashdict[tmppath] = md5
        j = generate_tree(tree, tmp[idx+1:])
        # i = idx + 1
        # j = tree
        # while i < lens:
        #     if i not in j.keys():
        #         j[tmp[i]] = dict()
        #
        #     j = j[tmp[i]]
        #     i += 1
        j['path'] = filepath
        j['functionNumber'] = fileinfo[filepath]['functionNumber']
        j['globalVariable'] = fileinfo[filepath]['globalVariable']
        j['size'] = fileinfo[filepath]['size']
        # j[''] 修改时间？
        # j[''] 圈复杂度？
        j['fanIn'] = len(fileinfo[filepath]['fanIn'])
        j['fanOut'] = len(fileinfo[filepath]['fanOut'])
        j['commentLine'] = fileinfo[filepath]['codeInfo']['commentLine']
        j['status'] = 0


def tree_json(tree):
    res = []

    for key in tree.keys():
        if 'functionNumber' in tree[key].keys():
            temp = {}
            temp['label'] = key
            temp["type"] = 'file'
            temp['status'] = tree[key]['status']
            temp['fileInfo'] = {}
            temp['fileInfo']['funcNum'] = tree[key]['functionNumber']
            temp['fileInfo']['globalVariableNum'] = tree[key]['globalVariable']
            temp['fileInfo']['fileSize'] = tree[key]['size']
            temp['fileInfo']['outDegree'] = tree[key]['fanOut']
            temp['fileInfo']['inDegree'] = tree[key]['fanIn']
            temp['fileInfo']['annotationLine'] = tree[key]['commentLine']
            res.append(copy.deepcopy(temp))
        else:
            temp = {}
            temp['label'] = key
            temp["type"] = 'dir'
            temp['children'] = tree_json(tree[key])
            res.append(copy.deepcopy(temp))
    return res


def get_tree_id(tree_json, id=0):
    for i in tree_json:
        i['id'] = id
        id += 1
        if 'children' in i.keys():
            get_tree_id(i['children'], id)


def get_file_treeCPP(fileinfo, tree, hashdict, filelist):
    '''
    生成文件数
    :param fileinfo: 文件信息
    :param tree: 空字典
    :param hashdict: md5
    :return:
    '''
    for filepath in fileinfo.keys():
        if filepath not in filelist:
            continue

        tmp = filepath.split('/')
        idx = tmp.index("code")
        lens = len(tmp)
        # md5 = getmd5(filepath)
        # hashdict[filepath] = md5
        tmppath = ''
        for k in range(idx+2, lens):
            tmppath = tmppath + tmp[k] + '/'
        tmppath = tmppath[:-1]
        md5 = getmd5(filepath)
        hashdict[tmppath] = md5
        j = generate_tree(tree, tmp[idx+1:])
        # i = idx + 1
        # j = tree
        # while i < lens:
        #     if i not in j.keys():
        #         j[tmp[i]] = dict()
        #
        #     j = j[tmp[i]]
        #     i += 1
        j['path'] = filepath
        j['classNumber'] = len(fileinfo[filepath]['classList'])
        j['globalVariable'] = fileinfo[filepath]['globalVariable']
        j['size'] = fileinfo[filepath]['size']
        # j[''] 修改时间？
        # j[''] 圈复杂度？
        j['fanIn'] = len(fileinfo[filepath]['fanIn'])
        j['fanOut'] = len(fileinfo[filepath]['fanOut'])
        j['commentLine'] = fileinfo[filepath]['codeInfo']['commentLine']
        j['status'] = 0


def tree_jsonCPP(tree):
    res = []

    for key in tree.keys():
        if 'classNumber' in tree[key].keys():
            temp = {}
            temp['label'] = key
            temp["type"] = 'file'
            temp['status'] = tree[key]['status']
            temp['fileInfo'] = {}
            temp['fileInfo']['classNum'] = tree[key]['classNumber']
            temp['fileInfo']['globalVariableNum'] = tree[key]['globalVariable']
            temp['fileInfo']['fileSize'] = tree[key]['size']
            temp['fileInfo']['outDegree'] = tree[key]['fanOut']
            temp['fileInfo']['inDegree'] = tree[key]['fanIn']
            temp['fileInfo']['annotationLine'] = tree[key]['commentLine']
            res.append(copy.deepcopy(temp))
        else:
            temp = {}
            temp['label'] = key
            temp["type"] = 'dir'
            temp['children'] = tree_jsonCPP(tree[key])
            res.append(copy.deepcopy(temp))
    return res


def getmd5(filename):
    # filename ="main.c"
    with open(filename, "rb") as f:
        bytes = f.read()  # read file as bytes
        readable_hash = hashlib.md5(bytes).hexdigest()
        # print(readable_hash)
    return readable_hash


def check_md5(oldhash, newhash, olddata, newdata):
    '''
    比较文件级变化
    :param oldhash:
    :param newhash:
    :param olddata:
    :param newdata:
    :return:
    '''
    oldkey = list(olddata.keys())[0]
    newkey = list(newdata.keys())[0]
    # define status
    deletetype = 1
    addtype = 2
    changetype = 3
    standtype = 0
    deletenum = 0
    addnum = 0
    changenum = 0
    for oldfile in oldhash.keys():
        if oldfile not in newhash.keys():
            tmp = oldfile.split('/')
            node = olddata[oldkey]
            for i in tmp:
                node = node[i]
            node['status'] = deletetype
            deletenum += 1
        else:
            if oldhash[oldfile] == newhash[oldfile]:
                continue
            else:
                tmp = oldfile.split('/')
                node = olddata[oldkey]
                for i in tmp:
                    node = node[i]
                node['status'] = changetype
                changenum += 1

                tmp = oldfile.split('/')
                node = newdata[newkey]
                for i in tmp:
                    node = node[i]
                node['status'] = changetype

    for newfile in newhash.keys():
        if newfile not in oldhash.keys():
            tmp = newfile.split('/')
            node = newdata[newkey]
            for i in tmp:
                node = node[i]
            node['status'] = addtype
            addnum += 1
    return changenum, deletenum, addnum


def writeexcel(data, sheet):
    if 'path' in data.keys() and 'globalVariable' in data.keys():
        if data['status'] == 1:
            msg = '删除'
        elif data['status'] == 2:
            msg = '新增'
        elif data['status'] == 3:
            msg = '修改'
        else:
            msg = '无变化'
        row_data = [data['path'].split('/')[-1], data['path'], data['functionNumber'], data['globalVariable'], data['size'], data['fanIn'], data['fanOut'], data['commentLine'], msg]
        sheet.append(row_data)
    else:
        for value in data.values():
            writeexcel(value, sheet)


def export_info(filelevel, outputpath):
    olddata = filelevel['oldexceldata']
    newdata = filelevel['newexceldata']
    workbook = Workbook()
    sheet1 = workbook.create_sheet(title='olddata')
    # 写入表头
    headers = ['文件名', '路径', '函数数量', '全局变量数量', '文件大小', '出度', '入度', '注释行数', '变更信息']
    sheet1.append(headers)
    writeexcel(olddata, sheet1)
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)
    sheet2 = workbook.create_sheet(title='newdata')
    # 写入表头
    headers = ['文件名', '路径', '函数数量', '全局变量数量', '文件大小', '出度', '入度', '注释行数', '变更信息']
    sheet2.append(headers)
    writeexcel(newdata, sheet2)

    # 保存文件
    workbook.save(outputpath)


def export_infoCPP(filelevel, outputpath):
    olddata = filelevel['selectedfileTree']
    newdata = filelevel['latestfileTree']
    workbook = Workbook()
    sheet1 = workbook.create_sheet(title='olddata')
    # 写入表头
    headers = ['文件名', '路径', '类数量', '全局变量数量', '文件大小', '出度', '入度', '注释行数', '变更信息']
    sheet1.append(headers)
    writeexcel(olddata, sheet1)
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)
    sheet2 = workbook.create_sheet(title='newdata')
    # 写入表头
    headers = ['文件名', '路径', '类数量', '全局变量数量', '文件大小', '出度', '入度', '注释行数', '变更信息']
    sheet2.append(headers)
    writeexcel(newdata, sheet2)

    # 保存文件
    workbook.save(outputpath)


if __name__ == '__main__':
    # print(getmd5("main.c") == getmd5("main1.c"))
    import json
    f = open('E:\CPP_master\dev0807\CPP_support\\uploads\saolei24\code\saolei/projectInfo.json', 'r')
    content = f.read()
    a = json.loads(content)
    olddata = {"projectInfo": a}
    f.close()
    f = open('E:\CPP_master\dev0807\CPP_support\\uploads\saolei24\code\saolei/codeFileInfo.json', 'r')
    content = f.read()
    a = json.loads(content)
    olddata['codeFileInfo'] = a
    f.close()
    f = open('E:\CPP_master\dev0807\CPP_support\\uploads\saolei24\code\saolei/projectInfo.json', 'r')
    content = f.read()
    a = json.loads(content)
    newdata = {"projectInfo": a}
    f.close()
    f = open('E:\CPP_master\dev0807\CPP_support\\uploads\saolei24\code\saolei/codeFileInfo.json', 'r')
    content = f.read()
    a = json.loads(content)
    newdata['codeFileInfo'] = a
    f.close()
    filelevel = getfile(olddata, newdata)
    print(filelevel)

    export_info(filelevel, 'output.xlsx')
    data = json.dumps(filelevel, indent=4, ensure_ascii=False)
    with open('E:\CPP_master\dev0807\CPP_support\\uploads\saolei24\code\saolei/fileinfo.json', 'w', newline='\n') as f:
        f.write(data)

    filelevel = getfileCPP(olddata, newdata)
    print(filelevel)
    data = json.dumps(filelevel, indent=4, ensure_ascii=False)
    with open('E:\CPP_master\dev0807\CPP_support\\uploads\saolei24\code\saolei/fileinfo1.json', 'w', newline='\n') as f:
        f.write(data)
