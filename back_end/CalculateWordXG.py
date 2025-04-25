import json
from Extract2_14 import *
import re
import datetime
import time

import openpyxl

import Extract2_14 as ex
import pandas as pd

pd.set_option('display.max_columns', None)  # 显示所有列
pd.set_option('display.max_rows', None)  # 显示所有行
pd.set_option('display.width', None)  # 自动调整列宽
pd.set_option('display.max_colwidth', None)  # 显示所有单元格的内容


def get_testcase(path):
    fp = pd.read_excel(
        io=path,
        sheet_name=None,  # 传入列表，指定Sheet索引
    )
    usecaseTable = pd.DataFrame(fp['测试用例'], columns=['用例编号', '测试步骤'])
    resultTable = pd.DataFrame(fp['执行结果'], columns=['用例编号', '测试类型', '执行结果', '实际结果', '预期结果'])
    generalTable = pd.merge(usecaseTable, resultTable, how='inner', on='用例编号')
    # print(generalTable.loc[:,'用例编号'])
    # print(generalTable.loc[:, '测试步骤'])
    # print(generalTable.loc[:, '执行结果'])
    # print(generalTable.loc[:, '实际结果'])
    # print(generalTable)
    # for i in generalTable.itertuples():
    #     print(getattr(i,'实际结果'),getattr(i,'执行结果'))
    generalTable["用例编号"] = generalTable["用例编号"].astype(str)
    return generalTable


def get_metric(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet = wb["需求跟踪矩阵"]
    r = sheet.max_row
    c = sheet.max_column
    df = pd.DataFrame(
        {"序号": pd.Series(), "需求名称": pd.Series(), "r_id": pd.Series(), "软件需求编号": pd.Series(),
         "tc_id": pd.Series(), "status": pd.Series()})
    position = [3, 0]
    for i in range(position[0], r):
        row = []
        row.append(sheet[i][position[1]].value)
        row.append(sheet[i][position[1] + 1].value)
        row.append(str(sheet[i][position[1] + 2].value))
        row.append(sheet[i][position[1] + 3].value)
        row.append(str(sheet[i][position[1] + 8].value))
        row.append(sheet[i][position[1] + 9].value)
        df.loc[row[0]] = row
    return df


def get_interface(treeList):
    df = pd.DataFrame({
        "r_id": pd.Series(),
        "des": pd.Series(),
        "support": pd.Series(),
        "receive": pd.Series(),
        "back": pd.Series(),
        "standard": pd.Series(),

    })
    for node in treeList:
        if node.text == "软件系统外部接口需求":
            for son in node.son_id:
                table = treeList[treeList[son].son_id[0]].table_info
                df.loc[table[0][1]] = pd.Series({
                    "r_id": table[0][1],
                    "des": table[2][1],
                    "support": table[4][1],
                    "receive": table[5][1],
                    "back": table[6][1],
                    "standard": table[7][1],

                })
    return df


def get_requirements(treeList):
    df = pd.DataFrame({
        "r_id": pd.Series(),
        "des": pd.Series(),
        "standard": pd.Series(),
        "tag": pd.Series(),
        "special": pd.Series(),
        "定制": pd.Series()
    })
    df.set_index("r_id")
    for node in treeList:
        if node.text == "功能需求":
            for i in node.son_id:
                if treeList[i].name == "Normal Table":
                    table = treeList[i].table_info
                    if df[df["r_id"] == table[0][1]].size > 0:
                        df.loc[table[0][1], "tag"] = "功能"
                    else:
                        df.loc[table[0][1]] = pd.Series(
                            {"r_id": table[0][1], "des": table[2][1], "standard": table[4][1], "定制": table[5][1],
                             "tag": "功能"})

        if node.text == "用户特定需求描述":
            for i in node.son_id:
                if treeList[i].name == "Normal Table":
                    table = treeList[i].table_info
                    if df[df["r_id"] == table[0][1]].size > 0:
                        df.loc[table[0][1], "special"] = True
                    else:
                        df.loc[table[0][1]] = pd.Series(
                            {"r_id": table[0][1], "des": table[2][1], "standard": table[4][1], "定制": table[5][1],
                             "special": True})
        # if node.text.startswith("适应性"):
        if node.text == "适用性：软件是否提供了相应的功能;":
            for i in node.son_id:
                if treeList[i].name == "Normal Table":
                    table = treeList[i].table_info
                    if df[df["r_id"] == table[0][1]].size > 0:
                        df.loc[table[0][1], "tag"] = "适用"
                    else:
                        df.loc[table[0][1]] = pd.Series(
                            {"r_id": table[0][1], "des": table[2][1], "standard": table[4][1], "定制": None,
                             "tag": "适用"})

        # if node.text.startswith("可靠性"):
        if node.text == "可靠性:产品在规定的条件下，在规定的时间内完成规定功能的能力;":
            for i in node.son_id:
                if treeList[i].name == "Normal Table":
                    table = treeList[i].table_info
                    if df[df["r_id"] == table[0][1]].size > 0:
                        df.loc[table[0][1], "tag"] = "可靠"
                    else:
                        df.loc[table[0][1]] = pd.Series(
                            {"r_id": table[0][1], "des": table[2][1], "standard": table[4][1], "定制": None,
                             "tag": "可靠"})
        # if node.text.startswith("易用性"):
        if node.text == "易用性：在指定使用条件下，产品被理解、学习、使用和吸引用户的能力;":
            for i in node.son_id:
                if treeList[i].name == "Normal Table":
                    table = treeList[i].table_info
                    if df[df["r_id"] == table[0][1]].size > 0:
                        df.loc[table[0][1], "tag"] = "易用"
                    else:
                        df.loc[table[0][1]] = pd.Series(
                            {"r_id": table[0][1], "des": table[2][1], "standard": table[4][1], "定制": None,
                             "tag": "易用"})
        if node.text == "效率性：在规定的条件下，相对于所用资源的数量，软件产品可提供适当性能的能力;":
            # if node.text.startswith("效率性"):
            for i in node.son_id:
                if treeList[i].name == "Normal Table":
                    table = treeList[i].table_info
                    if df[df["r_id"] == table[0][1]].size > 0:
                        df.loc[table[0][1], "tag"] = "效率"
                    else:
                        df.loc[table[0][1]] = pd.Series(
                            {"r_id": table[0][1], "des": table[2][1], "standard": table[4][1], "定制": None,
                             "tag": "效率"})

        # if node.text.startswith("维护性"):
        if node.text == "维护性：“四规”，在规定条件下，规定的时间内，使用规定的工具或方法修复规定功能的能力;":
            for i in node.son_id:
                if treeList[i].name == "Normal Table":
                    table = treeList[i].table_info
                    if df[df["r_id"] == table[0][1]].size > 0:
                        df.loc[table[0][1], "tag"] = "维护"
                    else:
                        df.loc[table[0][1]] = pd.Series(
                            {"r_id": table[0][1], "des": table[2][1], "standard": table[4][1], "定制": None,
                             "tag": "维护"})
        if node.text.startswith("可移植性"):
            # if node.text == "可移植性：从一种环境迁移到另一种环境的能力;":
            for i in node.son_id:
                if treeList[i].name == "Normal Table":
                    table = treeList[i].table_info
                    if df[df["r_id"] == table[0][1]].size > 0:
                        df.loc[table[0][1], "tag"] = "可移植"
                    else:
                        df.loc[table[0][1]] = pd.Series(
                            {"r_id": table[0][1], "des": table[2][1], "standard": table[4][1], "定制": None,
                             "tag": "可移植"})
        if node.text.startswith("兼容性"):
            # if node.text == "兼容性：产品在不同平台，不同设备，不同仪器上运行情况，是否稳定;":
            for i in node.son_id:
                if treeList[i].name == "Normal Table":
                    table = treeList[i].table_info
                    if df[df["r_id"] == table[0][1]].size > 0:
                        df.loc[table[0][1], "tag"] = "兼容"
                    else:
                        df.loc[table[0][1]] = pd.Series(
                            {"r_id": table[0][1], "des": table[2][1], "standard": table[4][1], "定制": None,
                             "tag": "兼容"})
        if node.text.startswith("信息安全性"):
            # if node.text == "信息安全性：产品在数据安全，网络安全方面是否有相应的防护措施":
            for i in node.son_id:
                if treeList[i].name == "Normal Table":
                    table = treeList[i].table_info
                    if df[df["r_id"] == table[0][1]].size > 0:
                        df.loc[table[0][1], "tag"] = "安全"
                    else:
                        df.loc[table[0][1]] = pd.Series(
                            {"r_id": table[0][1], "des": table[2][1], "standard": table[4][1], "定制": None,
                             "tag": "安全"})
    return df


def getindex(requirements, name, desc):
    reslist = []
    for i in requirements:
        if i.text == '非功能需求':
            for id in i.son_id:
                node = requirements[id]
                if name in node.text:
                    for cid in node.son_id:
                        child = requirements[cid]
                        table = child.table_info
                        if table:
                            for val in table.values():
                                if '描述' in val[0] and desc in val[1]:
                                    reslist.append(table[0][1])
                                    break
    print(reslist)
    return reslist


def getlistlens(description, query):
    i = '、'
    j = ';'
    idx = description.index(query)
    desc = description[idx:]
    lens_i = len(desc.split(i))
    lens_j = len(desc.split(j))
    return max(lens_i, lens_j)


def gettime(description):
    timelist = []
    while 1:
        ret_d = re.findall(r"\d+.\d+天|\d+天|\d+.\d+d|\d+d", description)
        if ret_d:
            for each_time in ret_d:
                time = re.match(r'\d+.\d+|\d+', each_time)
                time = float(time.group()) * 24 * 3600 * 1000  # 统一成毫秒
                # time = float(time.group())
                timelist.append(time)
            break
        ret_s = re.findall(r"\d+.\d+s|\d+s|\d+.\d+秒|\d+秒", description)
        if ret_s:
            for each_time in ret_s:
                time = re.match(r'\d+.\d+|\d+', each_time)
                time = float(time.group()) * 1000  # 统一成毫秒
                timelist.append(time)
            break
        ret_m_s = re.findall(r"(\d+?)分(\d+?)秒", description)
        if ret_m_s:
            for each_time in ret_m_s:
                time = float(each_time[0]) * 60 + float(each_time[1])
                time = time * 1000  # 统一成毫秒
                timelist.append(time)
            break
        ret_hms = re.findall(r"(\d+?)时(\d+?)分(\d+?)秒", description)
        if ret_hms:
            for each_time in ret_hms:
                time = float(each_time[0]) * 3600 + float(each_time[1]) * 60 + float(
                    each_time[2])
                time = time * 1000  # 统一成毫秒
                timelist.append(time)
            break
        ret_ms = re.findall(r"\d+.\d+ms|\d+ms|\d+.\d+毫秒|\d+毫秒", description)
        if ret_ms:
            for each_time in ret_ms:
                time = re.match(r'\d+.\d+|\d+', each_time)
                time = float(time.group())
                timelist.append(time)
            break

        break

    if timelist:
        return sum(timelist)/len(timelist)
    else:
        return 0


def getuid(df):
    id = ':'
    for i in df.itertuples():
        id = id + getattr(i, '用例编号') + ';'
    return id


def fcp1g(req, trace):
    try:
        gn = req[req["tag"] == "功能"]
        B = int(gn.count(axis=0)["r_id"])
        joined = gn.merge(trace[["r_id", "status"]], on="r_id", how="inner")
        A = int(joined.r_id.nunique())
        if joined[joined["status"] == "已关闭"].size != 0:
            A -= int(joined[joined["status"] == "已关闭"].groupby("r_id").r_id.nunique().iloc[0])
        A_dict = {
            "id": "A",
            "type": "需求跟踪矩阵",
            "source": tracefile,
            "des": "缺少的功能数量",
            "val": A
        }
        B_dict = {
            "id": "B",
            "type": "需求文档",
            "source": reqfile+"/功能需求",
            "des": "指定的功能数量",
            "val": B
        }
        res_dict = {
            "val": 1 - A / B,
            "sub": [A_dict, B_dict]
        }
        return res_dict
    except:
        print("fcp1g:err")


def fcr1g(req, trace, testcases):
    try:
        gn = req[req["tag"] == "功能"]
        B = int(gn.count(axis=0)["r_id"])
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "执行结果", "测试类型"]], left_on="tc_id", right_on="用例编号")
        A = int(all[(all["测试类型"] == "功能测试-反向") | (all["测试类型"] == "功能测试-正向")].r_id.nunique())
        if all[-(all["执行结果"] == "通过")].size != 0:
            A -= int(all[-(all["执行结果"] == "通过")].groupby("r_id").r_id.nunique().iloc[0])
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(all),
                "des": "功能不正确的数量",
                "val": int(all[-(all["执行结果"] == "通过")].groupby("r_id").r_id.nunique().iloc[0])
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/功能需求",
                "des": "考虑的功能数量",
                "val": B
            }
            res_dict = {
                "val": A / B,
                "sub": [A_dict, B_dict]
            }
            return res_dict
        res_dict = {
            "val": A / B,
            "sub": []
        }
        return res_dict
    except:
        print("fcr1g:err")


def fap1g(req, trace, testcases):
    try:
        gn = req[(req["special"] == True) & (req["tag"] == "功能")]
        open = trace[(trace["status"] == "已关闭")]
        joined = gn.merge(open[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "执行结果", "测试类型"]], left_on="tc_id", right_on="用例编号")
        filtered = all[(all["测试类型"] == "功能测试-反向") | (all["测试类型"] == "功能测试-正向")]
        res = []
        sublist = []
        j = 0
        for i in filtered.groupby("r_id"):
            A_dict = {
                "id": "A" + str(j),
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(i[1]),
                "des": "为实现特定使用目标所需的功能中缺少或不正确功能的数量",
                "val": len(i[1][i[1]["执行结果"] == "不通过"])
            }
            B_dict = {
                "id": "B" + str(j),
                "type": "需求文档",
                "source": reqfile + "/功能需求",
                "des": "为实现特定使用目标所需的功能数量",
                "val": len(i[1])
            }
            sublist.append(A_dict)
            sublist.append(B_dict)
            j += 1
            res.append(len(i[1][i[1]["执行结果"] == "通过"]) / len(i[1]))
        res_dict = {
            "val": res[0],
            "sub": sublist,
            "list": res
        }
        return res_dict
    except:
        print("fap1g:err")


def fap2g(list):
    try:
        A_dict = {
            "id": "A",
            "type": "测试文档",
            "source": "FAp-1-G",
            "des": "特定使用目标FAp-1-G的测量值的和",
            "val": sum(list)
        }
        B_dict = {
            "id": "B",
            "type": "测试文档",
            "source": "FAp-1-G",
            "des": "使用目标的数量",
            "val": len(list)
        }
        res_dict = {
            "val": sum(list) / len(list),
            "sub": [A_dict, B_dict]
        }
        return res_dict
    except:
        print("fap2g:err")


def filter_standard(row):
    standards = row["standard"].split(";")
    for standard in standards:
        m = re.match(".*?(" + standard + ").*", row["测试步骤"])
        if m is not None:
            return True
    return False


def fcl1g(req, trace, testcases):
    try:
        gn = req[req["tag"] == "功能"]
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '测试步骤', '测试类型', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")
        filtered = all[(all["测试类型"] == "功能测试-反向") | (all["测试类型"] == "功能测试-正向")]
        filter = filtered.apply(filter_standard, axis=1)
        standard_case = filtered[filter]
        if standard_case.size == 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "证实实现的功能依从性相关项数",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "测试文档",
                "source": "",
                "des": "与功能依从性相关项数",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(standard_case[standard_case["执行结果"] == "通过"]),
                "des": "证实实现的功能依从性相关项数",
                "val": len(standard_case[standard_case["执行结果"] == "通过"])
            }
            B_dict = {
                "id": "B",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(standard_case),
                "des": "与功能依从性相关项数",
                "val": len(standard_case)
            }
            res_dict = {
                "val": len(standard_case[standard_case["执行结果"] == "通过"]) / len(standard_case),
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("fcl1g:err")


def ptb1g(req, trace, testcases):
    try:
        xl = req[req["tag"] == "效率"]
        joined = xl.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果']], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*响应时间.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        sumA = 0
        sublist = []
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)(ms|s|秒|毫秒).*", row["实际结果"])
            A_dict = {
                "id": "A" + str(n),
                "type": "测试文档",
                "source": testfile + "/用例编号:" + row["tc_id"],
                "des": "第" + str(n + 1) + "次测量时系统响应一个特定用户任务或系统任务花费的时间",
                "val": None
            }

            if m is not None:
                if m.group(2) == "s" or m.group(2) == "秒":
                    A_dict["val"] = 1000 * float(m.group(1))
                    sumA += 1000 * float(m.group(1))
                else:
                    A_dict["val"] = 1000 * float(m.group(1))
                    sumA += float(m.group(1))
                n += 1
                sublist.append(A_dict)
        if n > 0:
            N_dict = {
                "id": "N",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "测得的响应次数",
                "val": n
            }
            sublist.append(N_dict)
            res_dict = {
                "val": sumA / n,
                "sub": sublist
            }
            return res_dict
        else:
            return -1
    except:
        print("ptb1g:err")


def ptb2g(req, ptb1g):
    try:
        xl = req[req["tag"] == "效率"]

        def filter_row(row):
            m = re.match(".*响应时间.*", row["des"])
            if m is not None:
                return True
            return False

        filter = xl.apply(filter_row, axis=1)
        target = xl[filter]
        B = 0
        for index, row in target.iterrows():
            m = re.match(".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)(ms|s|秒|毫秒).*", row["des"])
            if m is not None:
                if m.group(2) == "s" or m.group(2) == "秒":
                    B = 1000 * float(m.group(1))
                else:
                    B = float(m.group(1))
            else:
                return 0
        A_dict = {
            "id": "A",
            "type": "测试文档",
            "source": "PTb-1-G",
            "des": "PTb-1-G测度中所测量的平均响应时间",
            "val": ptb1g
        }
        B_dict = {
            "id": "B",
            "type": "需求文档",
            "source": reqfile + "/效率性",
            "des": "规定的任务响应时间",
            "val": B
        }
        res_dict = {
            "val": ptb1g / B,
            "sub": [A_dict, B_dict]
        }
        return res_dict
    except:
        print("ptb2g:err")


def ptb3g(req, trace, testcases):
    try:
        xl = req[req["tag"] == "效率"]
        joined = xl.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果']], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*周转时间.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        sumA = 0
        sublist = []
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*开始时刻为(\d*年\d*月\d*日\d*时\d*分\d*秒).*完成时刻为(\d*年\d*月\d*日\d*时\d*分\d*秒).*",
                         row["实际结果"])
            if m is not None:
                timestamp1 = time.mktime(time.strptime(m.group(1), "%Y年%m月%d日%H时%M分%S秒"))
                timestamp2 = time.mktime(time.strptime(m.group(2), "%Y年%m月%d日%H时%M分%S秒"))
                t = timestamp2 - timestamp1
                A_dict = {
                    "id": "B" + str(n) + "-A" + str(n),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + row["tc_id"],
                    "des": "周转时间",
                    "val": t
                }
                sublist.append(A_dict)
                sumA += t
                n += 1
            else:
                pass
        if n > 0:
            N_dict = {
                "id": "N",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "测量的次数",
                "val": n
            }
            sublist.append(N_dict)
            res_dict = {
                "val": sumA / n,
                "sub": sublist
            }
            return res_dict
        else:
            return 0
    except:
        print("ptb3g:err")


def ptb4g(req, ptb3g):
    try:
        xl = req[req["tag"] == "效率"]

        def filter_row(row):
            m = re.match(".*周转时间.*", row["des"])
            if m is not None:
                return True
            return False

        filter = xl.apply(filter_row, axis=1)
        target = xl[filter]
        B = 0
        for index, row in target.iterrows():
            m = re.match(".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)(s|秒).*", row["des"])
            if m is not None:
                B = float(m.group(1))
            else:
                return 0
        A_dict = {
            "id": "A",
            "type": "测试文档",
            "source": "PTb-3-G",
            "des": "PTb-3-G测度中所测量的平均周转时间",
            "val": ptb3g
        }
        B_dict = {
            "id": "B",
            "type": "需求文档",
            "source": reqfile + "/效率性",
            "des": "规定的作业或异步进程的周转时间",
            "val": B
        }
        res_dict = {
            "val": ptb3g / B,
            "sub": [A_dict, B_dict]
        }
        return res_dict
    except:
        print("ptb4g:err")


def ptb5g(req, trace, testcases):
    try:
        xl = req[req["tag"] == "效率"]
        joined = xl.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果']], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*吞吐量.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        # print(target)
        sublist = []
        sumA = 0
        n = 0
        for index, row in target.iterrows():
            m = re.match(
                ".*作业数量为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)(KB|MB|GB).*时间周期为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)(s|秒).*",
                row["实际结果"])
            if m is not None:
                A_dict = {
                    "id": "A" + str(n),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + row["tc_id"],
                    "des": "第" + str(n + 1) + "次观察时间内完成的作业数量",
                    "val": None
                }
                if m.group(2) == "MB":
                    A_dict["val"] = float(m.group(1)) * 1024
                    sumA += float(m.group(1)) * 1024 / float(m.group(3))
                if m.group(2) == "GB":
                    A_dict["val"] = float(m.group(1)) * 1024 * 1024
                    sumA += float(m.group(1)) * 1024 * 1024 / float(m.group(3))
                if m.group(2) == "KB":
                    A_dict["val"] = float(m.group(1))
                    sumA += float(m.group(1)) / float(m.group(3))

                B_dict = {
                    "id": "B" + str(n),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + row["tc_id"],
                    "des": "第" + str(n + 1) + "次观察时间的周期",
                    "val": float(m.group(3))
                }
                sublist.append(A_dict)
                sublist.append(B_dict)
                n += 1
            else:
                pass
        if n > 0:
            N_dict = {
                "id": "N",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "观察的次数",
                "val": n
            }
            sublist.append(N_dict)
            res_dict = {
                "val": sumA / n,
                "sub": sublist
            }

            return res_dict
        else:
            res_dict = {
                "val": sumA / n,
                "sub": sublist
            }
            return res_dict
    except:
        print("ptb5g:err")


def pru1g(req, trace, testcases):
    try:
        xl = req[req["tag"] == "效率"]
        joined = xl.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果']], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*处理器占用率.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        sumA = 0
        sublist = []
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*处理器占用率为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)%.*", row["实际结果"])
            m1 = re.match(
                ".*任务运行时间为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)s.*处理器使用时间为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)s.*",
                row["实际结果"])
            if m is not None:
                X_dict = {
                    "id": "X" + str(n),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + row["tc_id"],
                    "des": "第" + str(n + 1) + "次观察中,执行一组给定任务所占用处理器平均占用率",
                    "val": float(m.group(1)) / 100
                }
                sublist.append(X_dict)
                sumA += float(m.group(1)) / 100
                n += 1
            else:
                pass
            if m1 is not None:
                A_dict = {
                    "id": "A" + str(n),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + row["tc_id"],
                    "des": "第" + str(n + 1) + "次观察中,处理器执行一组给定任务所用的时间",
                    "val": float(m1.group(2))
                }
                B_dict = {
                    "id": "B" + str(n),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + row["tc_id"],
                    "des": "第" + str(n + 1) + "次观察中,执行任务的运行时间",
                    "val": float(m1.group(1))
                }
                sublist.append(A_dict)
                sublist.append(B_dict)
                sumA += float(m1.group(2)) / float(m1.group(1))
                n += 1
            else:
                pass

        if n > 0:
            N_dict = {
                "id": "N",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "处理的样本数",
                "val": n
            }
            sublist.append(N_dict)
            res_dict = {
                "val": sumA / n,
                "sub": sublist
            }
            return res_dict
        else:
            res_dict = {
                "val": None,
                "sub": sublist
            }
            return res_dict
    except:
        print("pru1g:err")


def pru2g(req, trace, testcases):
    try:
        xl = req[req["tag"] == "效率"]
        joined = xl.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果']], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*内存占用率.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        sumA = 0
        sublist = []
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*内存占用率为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)%.*", row["实际结果"])
            m1 = re.match(
                ".*观测期间系统可用内存值为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)M.*系统使用的实际内存值为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)M",
                row["实际结果"])
            if m is not None:
                X_dict = {
                    "id": "X" + str(n),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + row["tc_id"],
                    "des": "第" + str(n + 1) + "次观察中,执行一组给定任务所占用I/O设备的占用率",
                    "val": float(m.group(1)) / 100
                }
                sublist.append(X_dict)
                sumA += float(m.group(1)) / 100
                n += 1
            else:
                pass
            if m1 is not None:
                A_dict = {
                    "id": "A" + str(n),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + row["tc_id"],
                    "des": "第" + str(n + 1) + "次观察中,执行一组给定任务所占用的实际内存大小",
                    "val": float(m1.group(2))
                }
                B_dict = {
                    "id": "B" + str(n),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + row["tc_id"],
                    "des": "第" + str(n + 1) + "次观察中,可用于执行任务的内存大小",
                    "val": float(m1.group(1))
                }
                sublist.append(A_dict)
                sublist.append(B_dict)
                sumA += float(m1.group(2)) / float(m1.group(1))
                n += 1
            else:
                pass

        if n > 0:
            N_dict = {
                "id": "N",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "处理的样本数",
                "val": n
            }
            sublist.append(N_dict)
            res_dict = {
                "val": sumA / n,
                "sub": sublist
            }
            return res_dict
        else:
            res_dict = {
                "val": None,
                "sub": sublist
            }
            return res_dict
    except:
        print("pru2g:err")


def pru3g(req, trace, testcases):
    try:
        xl = req[req["tag"] == "效率"]
        joined = xl.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果']], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*I/O设备占用率.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        sumA = 0
        sublist = []
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*I/O设备占用率为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)%.*", row["实际结果"])
            m1 = re.match(
                ".*I/O设备所需时间为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)s.*实际I/O设备占用时间为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)s.*",
                row["实际结果"])
            if m is not None:
                X_dict = {
                    "id": "X" + str(n),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + row["tc_id"],
                    "des": "第" + str(n + 1) + "次观察中,执行一组给定任务所占用I/O设备的占用率",
                    "val": float(m.group(1)) / 100
                }
                sublist.append(X_dict)
                sumA += float(m.group(1)) / 100
                n += 1
            else:
                pass
            if m1 is not None:
                A_dict = {
                    "id": "A" + str(n),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + row["tc_id"],
                    "des": "第" + str(n + 1) + "次观察中,执行一组给定任务所占用I/O设备的持续时间",
                    "val": float(m1.group(2))
                }
                B_dict = {
                    "id": "B" + str(n),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + row["tc_id"],
                    "des": "第" + str(n + 1) + "次观察中,执行任务所需I/O运行的持续时间",
                    "val": float(m1.group(1))
                }
                sublist.append(A_dict)
                sublist.append(B_dict)
                sumA += float(m1.group(2)) / float(m1.group(1))
                n += 1
            else:
                pass

        if n > 0:
            N_dict = {
                "id": "N",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "观察次数",
                "val": n
            }
            sublist.append(N_dict)
            res_dict = {
                "val": sumA / n,
                "sub": sublist
            }
            return res_dict
        else:
            res_dict = {
                "val": None,
                "sub": sublist
            }
            return res_dict
    except:
        print("pru3g:err")


def pru4s(req, trace, testcases):
    try:
        xl = req[req["tag"] == "效率"]
        joined = xl.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果']], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*带宽占用率.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        sublist = []
        sumA = 0
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*带宽占用率为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)%.*", row["实际结果"])
            m1 = re.match(
                ".*测试期间系统可用带宽为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)Mbps，实际传输带宽为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)Mbps.*",
                row["实际结果"])
            if m is not None:
                X_dict = {
                    "id": "X" + str(n),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + row["tc_id"],
                    "des": "第" + str(n + 1) + "次观察中,执行一组给定任务时测得的实际带宽占用率",
                    "val": float(m.group(1)) / 100
                }
                sublist.append(X_dict)
                sumA += float(m.group(1)) / 100
                n += 1
            else:
                pass
            if m1 is not None:
                A_dict = {
                    "id": "A" + str(n),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + row["tc_id"],
                    "des": "第" + str(n + 1) + "次观察中,执行一组给定任务时测得的实际传输带宽",
                    "val": float(m1.group(2))
                }
                B_dict = {
                    "id": "B" + str(n),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + row["tc_id"],
                    "des": "第" + str(n + 1) + "次观察中,执行一组任务时可用带宽容量",
                    "val": float(m1.group(1))
                }
                sublist.append(A_dict)
                sublist.append(B_dict)
                sumA += float(m1.group(2)) / float(m1.group(1))
                n += 1
            else:
                pass
        if n > 0:
            res_dict = {
                "val": sumA / n,
                "sub": sublist
            }
            return res_dict
        else:
            res_dict = {
                "val": None,
                "sub": sublist
            }
            return res_dict
    except:
        print("pru4s:err")


def pca1g(req, trace, testcases):
    try:
        xl = req[req["tag"] == "效率"]
        joined = xl.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果']], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*事务处理容量.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        sumA = 0
        A = []
        B = []
        C = []
        n = 0
        for index, row in target.iterrows():
            m = re.match(
                ".*观察时长为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)s，完成事务数量为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*",
                row["实际结果"])
            if m is not None:
                sumA += float(m.group(2)) / float(m.group(1))
                A.append(float(m.group(2)))
                B.append(float(m.group(1)))
                C.append(row["tc_id"])
                n += 1
            else:
                pass
        if n > 0:
            sublist = []
            for i in range(n):
                A_dict = {
                    "id": "A" + str(i),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + C[i],
                    "des": "第" + str(i + 1) + "次观察时间内完成事务的数量",
                    "val": A[i]
                }
                B_dict = {
                    "id": "B" + str(i),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + C[i],
                    "des": "第" + str(i + 1) + "次观察时间",
                    "val": B[i]
                }
                sublist.append(A_dict)
                sublist.append(B_dict)
            res_dict = {
                "val": sumA / n,
                "sub": sublist
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "观察时间内完成事务的数量",
                "val": None
            }
            N_dict = {
                "id": "N",
                "type": "测试文档",
                "source": "",
                "des": "观察时间",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, N_dict]
            }
            return res_dict
    except:
        print("pca1g:err")


def pca2g(req, trace, testcases):
    try:
        xl = req[req["tag"] == "效率"]
        joined = xl.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果']], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*用户访问量.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        sumA = 0
        A = []
        C = []
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*同时访问用户数量为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*", row["实际结果"])
            if m is not None:
                sumA += float(m.group(1))
                A.append(float(m.group(1)))
                C.append(row["tc_id"])
                n += 1
            else:
                pass
        if n > 0:
            sublist = []
            N_dict = {
                "id": "N",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "观察次数",
                "val": n
            }
            sublist.append(N_dict)
            for i in range(n):
                A_dict = {
                    "id": "A" + str(i),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + C[i],
                    "des": "第" + str(i + 1) + "次观察中,同时访问系统的最大用户数量",
                    "val": A[i]
                }
                sublist.append(A_dict)
            res_dict = {
                "val": sumA / n,
                "sub": sublist
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "第i次观察中,同时访问系统的最大用户数量",
                "val": None
            }
            N_dict = {
                "id": "N",
                "type": "测试文档",
                "source": "",
                "des": "观察次数",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, N_dict]
            }

            return res_dict
    except:
        print("pca2g:err")


def pca3s(req, trace, testcases):
    try:
        xl = req[req["tag"] == "效率"]
        joined = xl.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果']], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*用户访问增长.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        A = []
        B = []
        C = []
        sumA = 0
        n = 0
        for index, row in target.iterrows():
            m = re.match(
                ".*观察时长为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*)s.*新增用户数量为([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*",
                row["实际结果"])
            if m is not None:
                sumA += float(m.group(2)) / float(m.group(1))
                A.append(float(m.group(2)))
                B.append(float(m.group(1)))
                C.append(row["tc_id"])
                n += 1
            else:
                pass
        if n > 0:
            sublist = []
            for i in range(n):
                A_dict = {
                    "id": "A" + str(i),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + C[i],
                    "des": "第" + str(i + 1) + "次观察时间内成功增加的用户数量",
                    "val": A[i]
                }
                B_dict = {
                    "id": "B" + str(i),
                    "type": "测试文档",
                    "source": testfile + "/用例编号:" + C[i],
                    "des": "第" + str(i + 1) + "次观察时间",
                    "val": B[i]
                }
                sublist.append(A_dict)
                sublist.append(B_dict)
            res_dict = {
                "val": sumA / n,
                "sub": sublist
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "观察时间内成功增加的用户数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "观察时间",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("pca2g:err")


def pcl1g(req, trace, testcases):
    try:
        xl = req[req["tag"] == "效率"]
        joined = xl.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '测试步骤', '测试类型', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")
        filter = all.apply(filter_standard, axis=1)
        standard_case = all[filter]
        if standard_case.size == 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "证实实现的效率依从性相关项数",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "测试文档",
                "source": "",
                "des": "与效率依从性相关项数",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(standard_case[standard_case["执行结果"] == "通过"]),
                "des": "证实实现的效率依从性相关项数",
                "val": len(standard_case[standard_case["执行结果"] == "通过"])
            }
            B_dict = {
                "id": "B",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(standard_case),
                "des": "与效率依从性相关项数",
                "val": len(standard_case)
            }
            res_dict = {
                "val": len(standard_case[standard_case["执行结果"] == "通过"]) / len(standard_case),
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("pcl1g:err")


def cco1g(req, trace, testcases):
    try:

        jr = req[req["tag"] == "兼容"]
        joined = jr.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果']], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*与以下软件共享运行环境.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        A = 0
        B = 0
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*测试期间系统可共存的软件有：(.*)。", row["实际结果"])
            m1 = re.match(".*与以下软件共享运行环境：(.*)，这样可以便于使用。", row["des"])
            if m is not None and m1 is not None:
                A = len(m.group(1).split("、"))
                B = len(m1.group(1).split("、"))
                n += 1
            else:
                pass
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "与该产品可共存的其他规定的软件产品数量",
                "val": A
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/兼容性",
                "des": "在运行环境中,该产品需要与其他软件产品共存的数量",
                "val": B
            }
            res_dict = {
                "val": A / B,
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "与该产品可共存的其他规定的软件产品数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "在运行环境中,该产品需要与其他软件产品共存的数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("cco1g:err")


def cin1g(interface, trace, testcases):
    try:
        joined = interface.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "测试步骤", '实际结果', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")
        B = 0
        for index, row in interface.iterrows():
            B += len(row["receive"].split(";"))
            B += len(row["back"].split(";"))

        def filter_row(row):
            m = re.match(".*数据格式.*", row["实际结果"])
            if m is not None:
                return True
            return False
        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        A = len(target[target["执行结果"] == "通过"])
        if B > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "与其他软件或系统可交换数据格式的数量",
                "val": A
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/软件系统外部接口需求",
                "des": "需要交换的数据格式数量",
                "val": B
            }
            res_dict = {
                "val": A / B,
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "与该产品可共存的其他规定的软件产品数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "在运行环境中,该产品需要与其他软件产品共存的数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("cin1g:err")


def cin2g(interface, trace, testcases):
    try:
        joined = interface.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "测试步骤", '实际结果', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")
        B = 0
        for index, row in interface.iterrows():
            B += len(row["support"].split(";"))

        def filter_row(row):
            m = re.match(".*协议类型.*", row["实际结果"])
            if m is not None:
                return True
            return False
        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        A = len(target[target["执行结果"] == "通过"])
        if B > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "实际支持数据交换协议的数量",
                "val": A
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/软件系统外部接口需求",
                "des": "规定支持的数据交换协议数量",
                "val": B
            }
            res_dict = {
                "val": A / B,
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "实际支持数据交换协议的数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "规定支持的数据交换协议数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("cin2g:err")


def cin3s(interface, trace, testcases):
    try:
        joined = interface.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "测试步骤", '实际结果', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")
        B = len(interface)
        group = all.groupby("r_id")
        A = 0
        for i in group:
            if len(i[1][i[1]["执行结果"] == "通过"]) == len(i[1]):
                A += 1
        if B > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(all),
                "des": "有效的外部接口数量",
                "val": A
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/软件系统外部接口需求",
                "des": "规定的外部接口数量",
                "val": B
            }
            res_dict = {
                "val": A / B,
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "有效的外部接口数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "规定的外部接口数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("cin3s:err")


def ccl1g(interface, req, trace, testcases):
    try:
        jr = req[req["tag"] == "兼容"]
        joined1 = jr.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all1 = joined1.merge(testcases[["用例编号", '测试步骤', '测试类型', "执行结果"]], left_on="tc_id",
                             right_on="用例编号")
        filter1 = all1.apply(filter_standard, axis=1)
        standard_case1 = all1[filter1]

        joined2 = interface.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all2 = joined2.merge(testcases[["用例编号", "测试步骤", '实际结果', "执行结果"]], left_on="tc_id",
                             right_on="用例编号")
        filter2 = all2.apply(filter_standard, axis=1)
        standard_case2 = all2[filter2]

        if standard_case1.size == 0 and standard_case2.size == 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "证实实现的兼容性的依从性相关项数",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "测试文档",
                "source": "",
                "des": "与兼容性的依从性相关项数",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(standard_case1[standard_case1["执行结果"] == "通过"])+getuid(standard_case2[standard_case2["执行结果"] == "通过"])[1:],
                "des": "证实实现的兼容性的依从性相关项数",
                "val": len(standard_case1[standard_case1["执行结果"] == "通过"]) + len(standard_case2[standard_case2["执行结果"] == "通过"])
            }
            B_dict = {
                "id": "B",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(standard_case1)+getuid(standard_case2)[1:],
                "des": "与兼容性的依从性相关项数",
                "val": len(standard_case1) + len(standard_case2)
            }
            res_dict = {
                "val": (len(standard_case1[standard_case1["执行结果"] == "通过"]) + len(standard_case2[standard_case2["执行结果"] == "通过"])) / (len(standard_case1) + len(standard_case2)),
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("ccl1g:err")


def uap1g(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果']], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*场景中系统使用方法.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        # print(target)
        A = []
        B = []
        sumA = 0
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*场景中系统使用方法：(.*)，.*", row["des"])
            m1 = re.match(".*使用场景有：(.*)。", row["实际结果"])
            if m is not None and m1 is not None:
                sumA += len(m1.group(1).split("、"))/len(m.group(1).split("、"))
                A.append(len(m1.group(1).split("、")))
                B.append(len(m.group(1).split("、")))
                n += 1
            else:
                pass
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "在产品描述或用户文档中所描述的使用场景数量",
                "val": A[0]
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "要求实现的功能总数量",
                "val": B[0]
            }
            res_dict = {
                "val": A[0]/B[0],
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "在产品描述或用户文档中所描述的使用场景数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "要求实现的功能总数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("uap1g:err")


def uap2g(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果']], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*演示功能.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        A = []
        B = []
        sumA = 0
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*演示功能：(.*)，.*", row["des"])
            m1 = re.match(".*演示功能的任务有：(.*)。", row["实际结果"])
            if m is not None and m1 is not None:
                sumA += len(m1.group(1).split("、"))/len(m.group(1).split("、"))
                A.append(len(m1.group(1).split("、")))
                B.append(len(m.group(1).split("、")))
                n += 1
            else:
                pass
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "具有演示功能的任务的数量",
                "val": A[0]
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/易用性",
                "des": "期望能从演示功能中获益的任务数量",
                "val": B[0]
            }
            res_dict = {
                "val": A[0]/B[0],
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "具有演示功能的任务的数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "期望能从演示功能中获益的任务数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("uap2g:err")


def uap3g(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果']], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*引导页.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        A = []
        B = []
        sumA = 0
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*引导页数量大于(\d+).*", row["des"])
            m1 = re.match(".*引导页数量为(\d+).*", row["实际结果"])
            if m is not None and m1 is not None:
                sumA += float(m1.group(1))/float(m.group(1))
                A.append(float(m1.group(1)))
                B.append(float(m.group(1)))
                n += 1
            else:
                pass
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "能说明网站目的的引导页数量",
                "val": A[0]
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/易用性",
                "des": "网站中需要引导页的数量",
                "val": B[0]
            }
            res_dict = {
                "val": A[0]/B[0],
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "能说明网站目的的引导页数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "网站中需要引导页的数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("uap3g:err")


def ule1g(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果']], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*用户文档和帮助机制中能描述以下功能.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        A = []
        B = []
        sumA = 0
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*用户文档和帮助机制中能描述以下功能：(.*)，.*", row["des"])
            m1 = re.match(".*测试中系统帮助机制描述的功能有：(.*)。", row["实际结果"])
            if m is not None:
                sumA += len(m1.group(1).split("、"))/len(m.group(1).split("、"))
                A.append(len(m1.group(1).split("、")))
                B.append(len(m.group(1).split("、")))
                n += 1
            else:
                pass
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "在用户文档和/或帮助机制中按要求描述的功能数量",
                "val": A[0]
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/易用性",
                "des": "要求实现的功能总数量",
                "val": B[0]
            }
            res_dict = {
                "val": A[0]/B[0],
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "在用户文档和/或帮助机制中按要求描述的功能数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "要求实现的功能总数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("ule1g:err")


def ule3s(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果']], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*提供出错原因和可能解决方法.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        A = []
        B = []
        sumA = 0
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*差错信息时提供出错原因和可能解决方法：(.*)，.*", row["des"])
            m1 = re.match(".*系统给出差错发生原因及可能解决方法的差错信息有：(.*)。", row["实际结果"])
            if m is not None and m1 is not None:
                sumA += len(m1.group(1).split("、"))/len(m.group(1).split("、"))
                A.append(len(m1.group(1).split("、")))
                B.append(len(m.group(1).split("、")))
                n += 1
            else:
                pass
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "在用户文档和/或帮助机制中按要求描述的功能数量",
                "val": A[0]
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/易用性",
                "des": "要求实现的功能总数量",
                "val": B[0]
            }
            res_dict = {
                "val": A[0]/B[0],
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "在用户文档和/或帮助机制中按要求描述的功能数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "要求实现的功能总数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("ule3s:err")


def ule4s(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果']], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*完成常规任务所需信息元素和步骤的数量.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        A = []
        B = []
        sumA = 0
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*新用户完成常规任务所需信息元素和步骤的数量大于(\d+).*", row["des"])
            m1 = re.match(".*用户可理解的信息元素和步骤的数量为(\d+).*", row["实际结果"])
            if m is not None and m1 is not None:
                sumA += float(m1.group(1)) / float(m.group(1))
                A.append(float(m1.group(1)))
                B.append(float(m.group(1)))
                n += 1
            else:
                pass
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "以用户可以理解的方式所呈现信息元素和步骤的数量",
                "val": A[0]
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/易用性",
                "des": "对于新用户来说完成常规任务所需信息元素和步骤的数量",
                "val": B[0]
            }
            res_dict = {
                "val": A[0] / B[0],
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "以用户可以理解的方式所呈现信息元素和步骤的数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "对于新用户来说完成常规任务所需信息元素和步骤的数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("ule4g:err")


def uop1g(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果']], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*操作结果与外观具有一致性.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        A = []
        B = []
        sumA = 0
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*系统在以下交互式任务中操作结果与外观具有一致性：(.*)，.*", row["des"])
            m1 = re.match(".*不一致的交互任务有：(.*)。", row["实际结果"])
            if m is not None and m1 is not None:
                sumA += len(m1.group(1).split("、"))/len(m.group(1).split("、"))
                A.append(len(m1.group(1).split("、")))
                B.append(len(m.group(1).split("、")))
                n += 1
            else:
                pass
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "不一致的特定交互式任务数量",
                "val": A[0]
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/易用性",
                "des": "需要一致的交互任务的数量",
                "val": B[0]
            }
            res_dict = {
                "val": 1 - (A[0]/B[0]),
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "不一致的特定交互式任务数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "需要一致的交互任务的数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("uop1g:err")


def uop2g(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*传达正确结果或指令.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        # print(target)
        n = len(target)
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "传达给用户正确结果或指令的消息数量",
                "val": len(target[target["执行结果"] == "通过"])
            }
            B_dict = {
                "id": "B",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "实现的消息数量",
                "val": len(target)
            }
            res_dict = {
                "val": len(target[target["执行结果"] == "通过"])/len(target),
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "传达给用户正确结果或指令的消息数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "测试文档",
                "source": "",
                "des": "实现的消息数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("uop2g:err")


def uop3s(req, trace, testcases):
    try:
        gn = req[req["tag"] == "功能"]
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*定制.*", row["实际结果"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        # print(target)
        n = len(target)
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "为用户使用方便而提供的可被定制的功能和操作规程的数量",
                "val": len(target[target["执行结果"] == "通过"])
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/易用性",
                "des": "用户能够受益于定制的功能和操作规程的数量",
                "val": len(target)
            }
            res_dict = {
                "val": len(target[target["执行结果"] == "通过"])/len(target),
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "为用户使用方便而提供的可被定制的功能和操作规程的数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "用户能够受益于定制的功能和操作规程的数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("uop3s:err")


def uop5s(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*监视以下功能状态.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        # print(target)
        A = []
        B = []
        sumA = 0
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*能够监视以下功能状态：(.*)，.*", row["des"])
            m1 = re.match(".*状态监视的功能有：(.*)。", row["实际结果"])
            if m is not None and m1 is not None:
                sumA += len(m1.group(1).split("、")) / len(m.group(1).split("、"))
                A.append(len(m1.group(1).split("、")))
                B.append(len(m.group(1).split("、")))
                n += 1
            else:
                pass
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "具有状态监视能力的功能数量",
                "val": A[0]
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/易用性",
                "des": "期望受益于监视能力的功能数量",
                "val": B[0]
            }
            res_dict = {
                "val": A[0] / B[0],
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "具有状态监视能力的功能数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "期望受益于监视能力的功能数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("uop3s:err")


def uop6s(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*支持撤销操作或重新确认.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        A = []
        B = []
        sumA = 0
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*支持撤销操作或重新确认：(.*)，.*", row["des"])
            m1 = re.match(".*系统提供撤销操作或重新确认的任务有：(.*)。", row["实际结果"])
            if m is not None and m1 is not None:
                sumA += len(m1.group(1).split("、")) / len(m.group(1).split("、"))
                A.append(len(m1.group(1).split("、")))
                B.append(len(m.group(1).split("、")))
                n += 1
            else:
                pass
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "提供撤销操作或重新确认的任务数量",
                "val": A[0]
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/易用性",
                "des": "用户能够从重新确认或撒销操作中获益的任务数量",
                "val": B[0]
            }
            res_dict = {
                "val": A[0] / B[0],
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "提供撤销操作或重新确认的任务数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "用户能够从重新确认或撒销操作中获益的任务数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("uop6s:err")


def uop7s(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*使用以下信息结构.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        A = []
        B = []
        sumA = 0
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*以下信息结构：(.*)，.*", row["des"])
            m1 = re.match(".*信息结构数量有：(.*)。", row["实际结果"])
            if m is not None and m1 is not None:
                sumA += len(m1.group(1).split("、")) / len(m.group(1).split("、"))
                # if row["执行结果"] == "通过":
                #     A.append(len(m1.group(1).split("、")))
                # else:
                #     A.append(0)
                A.append(len(m1.group(1).split("、")))
                B.append(len(m.group(1).split("、")))
                n += 1
            else:
                pass
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "对于预期用户来说,熟悉和方便的信息结构数量",
                "val": A[0]
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/易用性",
                "des": "使用的信息结构数量",
                "val": B[0]
            }
            res_dict = {
                "val": A[0] / B[0],
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "对于预期用户来说,熟悉和方便的信息结构数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "使用的信息结构数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("uop7s:err")


def uop8s(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*具有相似项且外观相似的比例.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        # print(target)
        A = []
        B = []
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*具有相似项的用户界面数量为(\d*)，具有相似项但外观不同的用户界面数量为(\d*).*", row["实际结果"])
            if m is not None:
                # if row["执行结果"] == "通过":
                #     A.append(len(m1.group(1).split("、")))
                # else:
                #     A.append(0)
                A.append(float(m.group(2)))
                B.append(float(m.group(1)))
                n += 1
            else:
                pass
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "具有相似项但外观不同的用户界面的数量",
                "val": A[0]
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/易用性",
                "des": "只有相似项的用户界面的数量",
                "val": B[0]
            }
            res_dict = {
                "val": 1 - (A[0] / B[0]),
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "具有相似项但外观不同的用户界面的数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "只有相似项的用户界面的数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("uop8s:err")


def uop9s(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*启动以下任务.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        # print(target)
        A = []
        B = []
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*启动以下任务：(.*)，.*", row["des"])
            m1 = re.match(".*启动的任务有：(.*)。", row["实际结果"])
            if m is not None and m1 is not None:
                # if row["执行结果"] == "通过":
                #     A.append(len(m1.group(1).split("、")))
                # else:
                #     A.append(0)
                A.append(len(m1.group(1).split("、")))
                B.append(len(m.group(1).split("、")))
                n += 1
            else:
                pass
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "可由所有适当的输入方法启动任务的数量",
                "val": A[0]
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/易用性",
                "des": "系统支持的任务数量",
                "val": B[0]
            }
            res_dict = {
                "val": A[0] / B[0],
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "可由所有适当的输入方法启动任务的数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "系统支持的任务数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("uop9s:err")


def uep1g(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*抵御误操作.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        # print(target)
        A = []
        B = []
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*抵御误操作：(.*)，.*", row["des"])
            m1 = re.match(".*系统能够抵御的用户误操作有：(.*)。", row["实际结果"])
            if m is not None and m1 is not None:
                # if row["执行结果"] == "通过":
                #     A.append(len(m1.group(1).split("、")))
                # else:
                #     A.append(0)
                A.append(len(m1.group(1).split("、")))
                B.append(len(m.group(1).split("、")))
                n += 1
            else:
                pass
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "可以防止导致系统故障的用户操作和输人的数量",
                "val": A[0]
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/易用性",
                "des": "实际操作中可以防止导致系统故障的用户操作和输入的数量",
                "val": B[0]
            }
            res_dict = {
                "val": A[0] / B[0],
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "可以防止导致系统故障的用户操作和输人的数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "实际操作中可以防止导致系统故障的用户操作和输入的数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("uep1g:err")


def uep2s(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*检测到以下输入差错时提示纠正.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        # print(target)
        A = []
        B = []
        n = 0
        for index, row in target.iterrows():
            m = re.match(".*输入差错时提示纠正：(.*)，.*", row["des"])
            m1 = re.match(".*纠正的输入差错有：(.*)。", row["实际结果"])
            if m is not None and m1 is not None:
                # if row["执行结果"] == "通过":
                #     A.append(len(m1.group(1).split("、")))
                # else:
                #     A.append(0)
                A.append(len(m1.group(1).split("、")))
                B.append(len(m.group(1).split("、")))
                n += 1
            else:
                pass
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "系统提供建议的修改值的输入差错数量",
                "val": A[0]
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/易用性",
                "des": "检测到的输入差错数量",
                "val": B[0]
            }
            res_dict = {
                "val": A[0] / B[0],
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "系统提供建议的修改值的输入差错数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "检测到的输入差错数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("uep2s:err")


def uep3s(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*用户操作的差错可以进行纠正.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        n = len(target)
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "由系统恢复的用户差错数量,这些用户差错是经设计并测试的",
                "val": len(target[target["执行结果"] == "通过"])
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/易用性",
                "des": "操作过程中可能发生的用户差错数量",
                "val": len(target)
            }
            res_dict = {
                "val": len(target[target["执行结果"] == "通过"]) / len(target),
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "由系统恢复的用户差错数量,这些用户差错是经设计并测试的",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "操作过程中可能发生的用户差错数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("uep3s:err")


def uac1g(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*特殊群体.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        n = len(target)
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "特殊群体用户成功使用的功能数量",
                "val": len(target[target["执行结果"] == "通过"])
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/易用性",
                "des": "实现的功能数量",
                "val": len(target)
            }
            res_dict = {
                "val": len(target[target["执行结果"] == "通过"]) / len(target),
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "特殊群体用户成功使用的功能数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "实现的功能数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("uac1g:err")


def uac2s(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '实际结果', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")

        def filter_row(row):
            m = re.match(".*语种.*", row["des"])
            if m is not None:
                return True
            return False

        filter = all.apply(filter_row, axis=1)
        target = all[filter]
        # print(target)
        n = len(target)
        A = []
        B = []
        for index, row in target.iterrows():
            m = re.match(".*语种：(.*)，.*", row["des"])
            m1 = re.match(".*语种有：(.*)。", row["实际结果"])
            if m is not None and m1 is not None:
                # if row["执行结果"] == "通过":
                #     A.append(len(m1.group(1).split("、")))
                # else:
                #     A.append(0)
                A.append(len(m1.group(1).split("、")))
                B.append(len(m.group(1).split("、")))
                n += 1
            else:
                pass
        if n > 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(target),
                "des": "实际支持的语种数量",
                "val": A[0]
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": reqfile + "/易用性",
                "des": "需要支持的语种数量",
                "val": B[0]
            }
            res_dict = {
                "val": A[0] / B[0],
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "实际支持的语种数量",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "需求文档",
                "source": "",
                "des": "需要支持的语种数量",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("uac2s:err")


def ucl1g(req, trace, testcases):
    try:
        yy = req[req["tag"] == "易用"]
        joined = yy.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '测试步骤', '测试类型', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")
        filter = all.apply(filter_standard, axis=1)
        standard_case = all[filter]
        # print(standard_case)
        if standard_case.size == 0:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": "",
                "des": "证实实现的易用性的依从性相关项数",
                "val": None
            }
            B_dict = {
                "id": "B",
                "type": "测试文档",
                "source": "",
                "des": "与易用性的依从性相关项数",
                "val": None
            }
            res_dict = {
                "val": None,
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            A_dict = {
                "id": "A",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(standard_case[standard_case["执行结果"] == "通过"]),
                "des": "证实实现的易用性的依从性相关项数",
                "val": len(standard_case[standard_case["执行结果"] == "通过"])
            }
            B_dict = {
                "id": "B",
                "type": "测试文档",
                "source": testfile + "/用例编号" + getuid(standard_case),
                "des": "与易用性的依从性相关项数",
                "val": len(standard_case)
            }
            res_dict = {
                "val": len(standard_case[standard_case["执行结果"] == "通过"]) / len(standard_case),
                "sub": [A_dict, B_dict]
            }
            return res_dict
    except:
        print("pcl1g:err")


def rma1g(rrTreeList):
    # root = rrTreeList[0]
    A_dict = {
        "id": "A",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/故障修复率/修复的可靠性相关故障",
        "des": "设计/编码/测试阶段修复的与可靠性相关故障数",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/故障修复率/修复的可靠性相关故障",
        "des": "设计/编码/测试阶段检测到的与可靠性相关的故障数",
        "val": None
    }
    try:
        for i in rrTreeList:
            if i.text == '修复的可靠性相关故障':
                for sonid in i.son_id:
                    child = rrTreeList[sonid]
                    table = child.table_info
                    if table:
                        A = len(table) - 1
                        A_dict = {
                            "id": "A",
                            "type": "可靠性测试方案",
                            "source": reliablityfile + "/修复的可靠性相关故障",
                            "des": "设计/编码/测试阶段修复的与可靠性相关故障数",
                            "val": A
                        }
                        break

            if i.text == '检测到的可靠性相关故障':
                for sonid in i.son_id:
                    child = rrTreeList[sonid]
                    table = child.table_info
                    if table:
                        B = len(table) - 1
                        B_dict = {
                            "id": "B",
                            "type": "可靠性测试方案",
                            "source": reliablityfile + "/修复的可靠性相关故障",
                            "des": "设计/编码/测试阶段检测到的与可靠性相关的故障数",
                            "val": B
                        }
                        break

        res = A_dict['val'] / B_dict['val']
        res_dict = {
            "val": res,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def rma2g(rrTreeList):
    A_dict = {
        "id": "A",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/平均失效间隔时间/运行时间",
        "des": "运行时间",
        "val": None
    }
    A_list = []
    A_data = {}
    B_dict = {
        "id": "B",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/平均失效间隔时间/系统/软件失效次数",
        "des": "实际发生的系统/软件失效次数",
        "val": None
    }
    B_list = []
    B_data = {}
    try:
        for i in rrTreeList:
            if i.text == '平均失效间隔时间':
                if '运行时间' in i.son_text:
                    idx = i.son_text.index('运行时间')
                    id = i.son_id[idx]
                    child = rrTreeList[id]
                    for sonid in child.son_id:
                        cchild = rrTreeList[sonid]
                        table = cchild.table_info
                        if table:
                            for row in range(1, len(table)):
                                content = table[row][3]
                                pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                ave_time = pattern.findall(content)[0]
                                A_data[table[row][2]] = float(ave_time)
                            # print(ave_time)
                if '系统/软件失效次数' in i.son_text:
                    idx = i.son_text.index('系统/软件失效次数')
                    id = i.son_id[idx]
                    child = rrTreeList[id]
                    for sonid in child.son_id:
                        cchild = rrTreeList[sonid]
                        table = cchild.table_info
                        if table:
                            for row in range(1, len(table)):
                                content = table[row][3]
                                pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                ave_time = pattern.findall(content)[0]
                                B_data[table[row][2]] = float(ave_time)

        Xlist = []
        sub = []
        num = 0
        for B in B_data:
            B_dict = {
                "id": "B" + str(num),
                "type": "可靠性测试方案",
                "source": reliablityfile + "/平均失效间隔时间/系统/软件失效次数",
                "des": "实际发生的系统/软件失效次数",
                "val": B_data[B]
            }
            sub.append(B_dict)
            if B in A_data:
                A_dict = {
                    "id": "A" + str(num),
                    "type": "可靠性测试方案",
                    "source": reliablityfile + "/平均失效间隔时间/运行时间",
                    "des": "运行时间",
                    "val": A_data[B]
                }
                sub.append(A_dict)
                if B_data[B] == 0:
                    Xlist.append(A_data[B])
                else:
                    Xlist.append(A_data[B]/B_data[B])

        # res = A_dict['val'] / B_dict['val']
        res_dict = {
            "val": round(sum(Xlist)/len(Xlist), 3),
            "sub": sub
        }
    except:
        A_dict = {
            "id": "A",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/平均失效间隔时间/运行时间",
            "des": "运行时间",
            "val": None
        }

        B_dict = {
            "id": "B",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/平均失效间隔时间/系统/软件失效次数",
            "des": "实际发生的系统/软件失效次数",
            "val": None
        }
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def rma3g(rrTreeList):
    A_dict = {
        "id": "A",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/周期失效率/观察持续周期数",
        "des": "在观察时间内检测到的失效数量",
        "val": None
    }
    A_data = {}
    B_dict = {
        "id": "B",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/周期失效率/观察周期内检测到的失效数",
        "des": "观察持续周期数",
        "val": None
    }
    B_data = {}
    try:
        for i in rrTreeList:
            if i.text == '周期失效率':
                if '观察周期内检测到的失效数' in i.son_text:
                    idx = i.son_text.index('观察周期内检测到的失效数')
                    id = i.son_id[idx]
                    child = rrTreeList[id]
                    for sonid in child.son_id:
                        cchild = rrTreeList[sonid]
                        table = cchild.table_info
                        if table:
                            for row in range(1, len(table)):
                                content = table[row][3]
                                pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                ave_time = pattern.findall(content)[0]
                                A_data[table[row][2]] = float(ave_time)
                            # print(ave_time)
                if '观察持续周期数' in i.son_text:
                    idx = i.son_text.index('观察持续周期数')
                    id = i.son_id[idx]
                    child = rrTreeList[id]
                    for sonid in child.son_id:
                        cchild = rrTreeList[sonid]
                        table = cchild.table_info
                        if table:
                            for row in range(1, len(table)):
                                content = table[row][4]
                                pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                ave_time = pattern.findall(content)[0]
                                B_data[table[row][2]] = float(ave_time)

        Xlist = []
        sub = []
        num = 0
        for B in B_data:
            num += 1
            B_dict = {
                "id": "B" + str(num),
                "type": "可靠性测试方案",
                "source": reliablityfile + "/周期失效率/观察周期内检测到的失效数",
                "des": "观察持续周期数",
                "val": B_data[B]
            }
            sub.append(B_dict)
            if B in A_data:
                A_dict = {
                    "id": "A" + str(num),
                    "type": "可靠性测试方案",
                    "source": reliablityfile + "/周期失效率/观察持续周期数",
                    "des": "在观察时间内检测到的失效数量",
                    "val": A_data[B]
                }
                sub.append(A_dict)
                if B_data[B] == 0:
                    Xlist.append(A_data[B])
                else:
                    Xlist.append(A_data[B]/B_data[B])

        # res = A_dict['val'] / B_dict['val']
        res_dict = {
            "val": round(sum(Xlist)/len(Xlist), 3),
            "sub": sub
        }
    except:
        A_dict = {
            "id": "A",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/周期失效率/观察持续周期数",
            "des": "在观察时间内检测到的失效数量",
            "val": None
        }

        B_dict = {
            "id": "B",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/周期失效率/观察周期内检测到的失效数",
            "des": "观察持续周期数",
            "val": None
        }
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def rma4s(rrTreeList):
    A_dict = {
        "id": "A",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/测试覆盖率/实际执行的功能",
        "des": "实际所执行的系统或软件能力、运行场景或功能的数量",
        "val": None
    }
    A_data = {}
    A_set = set()
    B_dict = {
        "id": "B",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/测试覆盖率/预期包含的功能",
        "des": "实际所执行的系统或软件能力、运行场景或功能的数量",
        "val": None
    }
    B_data = {}
    B_list = []
    try:
        for i in rrTreeList:
            if i.text == '测试覆盖率':
                if '实际执行的功能' in i.son_text:
                    idx = i.son_text.index('实际执行的功能')
                    id = i.son_id[idx]
                    child = rrTreeList[id]
                    for sonid in child.son_id:
                        cchild = rrTreeList[sonid]
                        table = cchild.table_info
                        if table:
                            for row in range(1, len(table)):
                                content = table[row][5]
                                if ',' in content:
                                    funclist = content.split(',')
                                elif '，' in content:
                                    funclist = content.split(',')
                                else:
                                    funclist = [content,]

                                for func in funclist:
                                    A_set.add(func)
                                # pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                # ave_time = pattern.findall(content)[0]
                                # A_data[table[row][2]] = float(ave_time)
                            # print(ave_time)
                if '预期包含的功能' in i.son_text:
                    idx = i.son_text.index('预期包含的功能')
                    id = i.son_id[idx]
                    child = rrTreeList[id]
                    for sonid in child.son_id:
                        cchild = rrTreeList[sonid]
                        table = cchild.table_info
                        if table:
                            for row in range(1, len(table)):
                                content = table[row][0]
                                B_list.append(content)
                                # pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                # ave_time = pattern.findall(content)[0]
                                # B_data[table[row][2]] = float(ave_time)

        # Xlist = []
        # sub = []
        # num = 0
        # for B in B_data:
        #     num+=1
        #     B_dict = {
        #         "id": "B"+ str(num),
        #         "type": "可靠性测试方案",
        #         "source": reliablityfile + "/周期失效率",
        #         "des": "观察持续周期数",
        #         "val": B_data[B]
        #     }
        #     sub.append(B_dict)
        #     if B in A_data:
        #         A_dict = {
        #             "id": "A"+ str(num),
        #             "type": "可靠性测试方案",
        #             "source": reliablityfile + "/周期失效率",
        #             "des": "在观察时间内检测到的失效数量",
        #             "val": A_data[B]
        #         }
        #         sub.append(A_dict)
        #         if B_data[B] == 0:
        #             Xlist.append(A_data[B])
        #         else:
        #             Xlist.append(A_data[B]/B_data[B])

        # res = A_dict['val'] / B_dict['val']
        num = 0
        for i in A_set:
            if i in B_list:
                num += 1
        A_dict = {
            "id": "A",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/测试覆盖率/实际执行的功能",
            "des": "实际所执行的系统或软件能力、运行场景或功能的数量",
            "val": num
        }
        # print(A_set)
        B_dict = {
            "id": "B",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/测试覆盖率/预期包含的功能",
            "des": "实际所执行的系统或软件能力、运行场景或功能的数量",
            "val": len(B_list)
        }
        res_dict = {
            "val": round(A_dict["val"]/B_dict["val"], 3),
            "sub": [A_dict, B_dict]
        }
    except:

        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def rav1g(rrTreeList):
    A_dict = {
        "id": "A",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/系统可用性/实际系统运行时间",
        "des": "实际提供的系统运行时间",
        "val": None
    }
    A_data = {}
    B_dict = {
        "id": "B",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/系统可用性/计划系统运行时间",
        "des": "操作计划中规定的系统运行时间",
        "val": None
    }
    B_data = {}
    try:
        for i in rrTreeList:
            if i.text == '系统可用性':
                if '实际系统运行时间' in i.son_text:
                    idx = i.son_text.index('实际系统运行时间')
                    id = i.son_id[idx]
                    child = rrTreeList[id]
                    for sonid in child.son_id:
                        cchild = rrTreeList[sonid]
                        table = cchild.table_info
                        if table:
                            for row in range(1, len(table)):
                                content = table[row][4]
                                pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                ave_time = pattern.findall(content)[0]
                                A_data[table[row][3]] = float(ave_time)
                            # print(ave_time)
                if '计划系统运行时间' in i.son_text:
                    idx = i.son_text.index('计划系统运行时间')
                    id = i.son_id[idx]
                    child = rrTreeList[id]
                    for sonid in child.son_id:
                        cchild = rrTreeList[sonid]
                        table = cchild.table_info
                        if table:
                            for row in range(1, len(table)):
                                content = table[row][4]
                                pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                ave_time = pattern.findall(content)[0]
                                B_data[table[row][3]] = float(ave_time)

        Xlist = []
        sub = []
        num = 0
        for B in B_data:
            num += 1
            B_dict = {
                "id": "B" + str(num),
                "type": "可靠性测试方案",
                "source": reliablityfile + "/系统可用性/计划系统运行时间",
                "des": "操作计划中规定的系统运行时间",
                "val": B_data[B]
            }
            sub.append(B_dict)
            if B in A_data:
                A_dict = {
                    "id": "A" + str(num),
                    "type": "可靠性测试方案",
                    "source": reliablityfile + "/系统可用性/实际系统运行时间",
                    "des": "实际提供的系统运行时间",
                    "val": A_data[B]
                }
                sub.append(A_dict)
                if B_data[B] == 0:
                    Xlist.append(A_data[B])
                else:
                    Xlist.append(A_data[B]/B_data[B])

        # res = A_dict['val'] / B_dict['val']
        res_dict = {
            "val": round(sum(Xlist)/len(Xlist), 3),
            "sub": sub
        }
    except:
        A_dict = {
            "id": "A",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/系统可用性/实际系统运行时间",
            "des": "实际提供的系统运行时间",
            "val": None
        }

        B_dict = {
            "id": "B",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/系统可用性/计划系统运行时间",
            "des": "操作计划中规定的系统运行时间",
            "val": None
        }
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def rav2g(rrTreeList):
    A_dict = {
        "id": "A",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/平均宕机时间/检测到的宕机",
        "des": "总的宕机时间",
        "val": None
    }
    A_data = {}
    A_list = []
    B_dict = {
        "id": "B",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/平均宕机时间/检测到的宕机",
        "des": "观察到的宕机数",
        "val": None
    }
    B_data = {}
    B_list = []
    try:
        for i in rrTreeList:
            if i.text == '平均宕机时间':
                if '检测到的宕机' in i.son_text:
                    idx = i.son_text.index('检测到的宕机')
                    id = i.son_id[idx]
                    child = rrTreeList[id]
                    for sonid in child.son_id:
                        cchild = rrTreeList[sonid]
                        table = cchild.table_info
                        if table:
                            for row in range(1, len(table)):
                                contentA = table[row][5]
                                pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                ave_timeA = pattern.findall(contentA)[0]
                                A_list.append(float(ave_timeA))

                                contentB = table[row][4]
                                pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                ave_timeB = pattern.findall(contentB)[0]
                                B_list.append(float(ave_timeB))
                            # print(ave_time)

        Xlist = []
        sub = []
        num = 0
        for B in range(len(B_list)):
            num += 1
            B_dict = {
                "id": "B" + str(num),
                "type": "可靠性测试方案",
                "source": reliablityfile + "/平均宕机时间/检测到的宕机",
                "des": "观察到的宕机数",
                "val": B_list[B]
            }
            sub.append(B_dict)

            A_dict = {
                "id": "A" + str(num),
                "type": "可靠性测试方案",
                "source": reliablityfile + "/平均宕机时间/检测到的宕机",
                "des": "总的宕机时间",
                "val": A_list[B]
            }
            sub.append(A_dict)
            if B_list[B] == 0:
                Xlist.append(A_list[B])
            else:
                Xlist.append(A_list[B]/B_list[B])

        # res = A_dict['val'] / B_dict['val']
        res_dict = {
            "val": round(sum(Xlist)/len(Xlist), 3),
            "sub": sub
        }
    except:
        A_dict = {
            "id": "A",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/平均宕机时间/检测到的宕机",
            "des": "总的宕机时间",
            "val": None
        }

        B_dict = {
            "id": "B",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/平均宕机时间/检测到的宕机",
            "des": "观察到的宕机数",
            "val": None
        }
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def rft1g(rrTreeList):
    A_dict = {
        "id": "A",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/避免失效率/避免发生严重失效的用例",
        "des": "避免发生关键和严重失效的次数(以测试用例为单位计算的数量)",
        "val": None
    }
    A_data = {}
    B_dict = {
        "id": "B",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/避免失效率/执行的故障测试用例数",
        "des": "测试中执行的故障模式(几乎导致失效)的测试用例数量",
        "val": None
    }
    B_data = {}
    try:
        for i in rrTreeList:
            if i.text == '避免失效率':
                if '避免发生严重失效的用例' in i.son_text:
                    idx = i.son_text.index('避免发生严重失效的用例')
                    id = i.son_id[idx]
                    child = rrTreeList[id]
                    for sonid in child.son_id:
                        cchild = rrTreeList[sonid]
                        table = cchild.table_info
                        if table:
                            for row in range(1, len(table)):
                                content = table[row][5]
                                # pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                # ave_time = pattern.findall(content)[0]
                                if '是' in content:
                                    A_data[table[row][2]] = 1
                            # print(ave_time)
                if '执行的故障测试用例数' in i.son_text:
                    idx = i.son_text.index('执行的故障测试用例数')
                    id = i.son_id[idx]
                    child = rrTreeList[id]
                    for sonid in child.son_id:
                        cchild = rrTreeList[sonid]
                        table = cchild.table_info
                        if table:
                            for row in range(1, len(table)):
                                # content = table[row][4]
                                # pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                # ave_time = pattern.findall(content)[0]
                                B_data[table[row][2]] = 1

        # Xlist = []
        # sub = []
        # num = 0
        # for B in B_data:
        #     num+=1
        #     B_dict = {
        #         "id": "B"+ str(num),
        #         "type": "可靠性测试方案",
        #         "source": reliablityfile + "/系统可用性/计划系统运行时间",
        #         "des": "操作计划中规定的系统运行时间",
        #         "val": B_data[B]
        #     }
        #     sub.append(B_dict)
        #     if B in A_data:
        #         A_dict = {
        #             "id": "A"+ str(num),
        #             "type": "可靠性测试方案",
        #             "source": reliablityfile + "/系统可用性/实际系统运行时间",
        #             "des": "实际提供的系统运行时间",
        #             "val": A_data[B]
        #         }
        #         sub.append(A_dict)
        #         if B_data[B] == 0:
        #             Xlist.append(A_data[B])
        #         else:
        #             Xlist.append(A_data[B]/B_data[B])

        A_dict = {
            "id": "A",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/避免失效率/避免发生严重失效的用例",
            "des": "避免发生关键和严重失效的次数(以测试用例为单位计算的数量)",
            "val": len(A_data)
        }
        B_dict = {
            "id": "B",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/避免失效率/执行的故障测试用例数",
            "des": "测试中执行的故障模式(几乎导致失效)的测试用例数量",
            "val": len(B_data)
        }

        res = round(A_dict['val'] / B_dict['val'], 3)
        res_dict = {
            "val": res,
            "sub": [A_dict, B_dict]
        }
    except:
        A_dict = {
            "id": "A",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/避免失效率/避免发生严重失效的用例",
            "des": "避免发生关键和严重失效的次数(以测试用例为单位计算的数量)",
            "val": None
        }

        B_dict = {
            "id": "B",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/避免失效率/执行的故障测试用例数",
            "des": "测试中执行的故障模式(几乎导致失效)的测试用例数量",
            "val": None
        }
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def rft2s(rTreeList):
    A_dict = {
        "id": "A",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/组件的冗余度/系统组件",
        "des": "冗余安装系统组件的数量",
        "val": None
    }
    A_data = {}
    A_list = []
    B_dict = {
        "id": "B",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/组件的冗余度/系统组件",
        "des": "系统组件数量",
        "val": None
    }
    B_data = {}
    B_list = []
    try:
        for i in rTreeList:
            if i.text == '组件的冗余度':
                if '系统组件' in i.son_text:
                    idx = i.son_text.index('系统组件')
                    id = i.son_id[idx]
                    child = rTreeList[id]
                    for sonid in child.son_id:
                        cchild = rTreeList[sonid]
                        table = cchild.table_info
                        if table:
                            for row in range(1, len(table)):
                                contentA = table[row][5]
                                # pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                # ave_timeA = pattern.findall(contentA)[0]

                                # A_list.append(float(ave_timeA))

                                contentB = table[row][2]
                                # pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                # ave_timeB = pattern.findall(contentB)[0]
                                B_list.append(contentB)
                                if '是' in contentA:
                                    A_list.append(contentB)
                            # print(ave_time)

        Xlist = []
        sub = []
        num = 0

        A_dict = {
            "id": "A",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/组件的冗余度/系统组件",
            "des": "冗余安装系统组件的数量",
            "val": len(A_list)
        }

        B_dict = {
            "id": "B",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/组件的冗余度/系统组件",
            "des": "系统组件数量",
            "val": len(B_list)
        }
        res = A_dict['val'] * B_dict['val']
        res_dict = {
            "val": round(res, 3),
            "sub": [A_dict, B_dict]
        }
    except:
        A_dict = {
            "id": "A",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/组件的冗余度/系统组件",
            "des": "冗余安装系统组件的数量",
            "val": None
        }

        B_dict = {
            "id": "B",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/组件的冗余度/系统组件",
            "des": "系统组件数量",
            "val": None
        }
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def rft3s(rTreeList):
    A_dict = {
        "id": "A",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/平均故障通告时间/系统报告故障时刻",
        "des": "系统报告故障i的时刻",
        "val": None
    }
    A_data = {}
    B_dict = {
        "id": "B",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/平均故障通告时间/检测故障时刻",
        "des": "故障i被检测到的时刻",
        "val": None
    }
    B_data = {}
    try:
        for i in rTreeList:
            if i.text == '平均故障通告时间':
                if '系统报告故障时刻' in i.son_text:
                    idx = i.son_text.index('系统报告故障时刻')
                    id = i.son_id[idx]
                    child = rTreeList[id]
                    for sonid in child.son_id:
                        cchild = rTreeList[sonid]
                        table = cchild.table_info
                        if table:
                            for row in range(1, len(table)):
                                content = table[row][4]
                                # pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                # ave_time = pattern.findall(content)[0]
                                contentslpit = content.split("-")
                                content1, content2 = contentslpit[0], contentslpit[1]
                                content1s = content1.split('/')
                                content2s = content2.split(':')
                                Atime = datetime.datetime(int(content1s[0]), int(content1s[1]), int(content1s[2]),
                                                          int(content2s[0]), int(content2s[1]))

                                A_data[table[row][2]] = Atime
                            # print(ave_time)
                if '检测故障时刻' in i.son_text:
                    idx = i.son_text.index('检测故障时刻')
                    id = i.son_id[idx]
                    child = rTreeList[id]
                    for sonid in child.son_id:
                        cchild = rTreeList[sonid]
                        table = cchild.table_info
                        if table:
                            for row in range(1, len(table)):
                                # content = table[row][4]
                                # pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                # ave_time = pattern.findall(content)[0]
                                content = table[row][4]
                                contentslpit = content.split("-")
                                content1, content2 = contentslpit[0], contentslpit[1]
                                content1s = content1.split('/')
                                content2s = content2.split(':')
                                Btime = datetime.datetime(int(content1s[0]), int(content1s[1]), int(content1s[2]),
                                                          int(content2s[0]), int(content2s[1]))

                                B_data[table[row][2]] = Btime
                                # B_data[table[row][2]] = 1

        Xlist = []
        sub = []
        num = 0
        for B in B_data:
            num += 1
            B_dict = {
                "id": "B" + str(num),
                "type": "可靠性测试方案",
                "source": reliablityfile + "/平均故障通告时间/检测故障时刻",
                "des": "故障i被检测到的时刻",
                "val": str(B_data[B])
            }
            sub.append(B_dict)
            if B in A_data:
                A_dict = {
                    "id": "A" + str(num),
                    "type": "可靠性测试方案",
                    "source": reliablityfile + "/平均故障通告时间/系统报告故障时刻",
                    "des": "系统报告故障i的时刻",
                    "val": str(A_data[B])
                }
                sub.append(A_dict)
                Xlist.append((A_data[B]-B_data[B]).seconds/3600 + (A_data[B]-B_data[B]).days*24)
                # print(Xlist[-1])

                # if B_data[B] == 0:
                #     Xlist.append(A_data[B])
                # else:
                #     Xlist.append(A_data[B]/B_data[B])

        # A_dict = {
        #     "id": "A",
        #     "type": "可靠性测试方案",
        #     "source": reliablityfile + "/避免失效率/避免发生严重失效的用例",
        #     "des": "避免发生关键和严重失效的次数(以测试用例为单位计算的数量)",
        #     "val": len(A_data)
        # }
        # B_dict = {
        #     "id": "B",
        #     "type": "可靠性测试方案",
        #     "source": reliablityfile + "/避免失效率/执行的故障测试用例数",
        #     "des": "测试中执行的故障模式(几乎导致失效)的测试用例数量",
        #     "val": len(B_data)
        # }

        res_dict = {
            "val": round(sum(Xlist)/len(Xlist), 3),
            "sub": sub
        }
    except:
        A_dict = {
            "id": "A",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/平均故障通告时间/系统报告故障时刻",
            "des": "系统报告故障i的时刻",
            "val": None
        }

        B_dict = {
            "id": "B",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/平均故障通告时间/检测故障时刻",
            "des": "故障i被检测到的时刻",
            "val": None
        }
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def rre1g(rTreeList):
    A_dict = {
        "id": "A",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/平均恢复时间/系统失效恢复时间",
        "des": "由于第i次失效而重新启动,并恢复宕机的软件/系统所花费的总时间",
        "val": None
    }
    A_data = {}
    A_list = []
    B_dict = {
        "id": "B",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/平均恢复时间/系统失效恢复时间",
        "des": "发生失效的次数",
        "val": None
    }
    B_data = {}
    B_list = []
    try:
        for i in rTreeList:
            if i.text == '平均恢复时间':
                if '系统失效恢复时间' in i.son_text:
                    idx = i.son_text.index('系统失效恢复时间')
                    id = i.son_id[idx]
                    child = rTreeList[id]
                    for sonid in child.son_id:
                        cchild = rTreeList[sonid]
                        table = cchild.table_info
                        if table:
                            for row in range(1, len(table)):
                                contentA = table[row][3]
                                # pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                # ave_timeA = pattern.findall(contentA)[0]
                                contentslpit = contentA.split("-")
                                content1, content2 = contentslpit[0], contentslpit[1]
                                content1s = content1.split('/')
                                content2s = content2.split(':')
                                Atime = datetime.datetime(int(content1s[0]), int(content1s[1]), int(content1s[2]),
                                                          int(content2s[0]), int(content2s[1]))

                                contentB = table[row][4]

                                # pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                # ave_timeA = pattern.findall(contentA)[0]
                                contentslpit = contentB.split("-")
                                content1, content2 = contentslpit[0], contentslpit[1]
                                content1s = content1.split('/')
                                content2s = content2.split(':')
                                Btime = datetime.datetime(int(content1s[0]), int(content1s[1]), int(content1s[2]),
                                                          int(content2s[0]), int(content2s[1]))
                                # print(str(Btime))
                                A_list.append(
                                    (Btime-Atime).seconds/3600 + (Btime-Atime).days * 24
                                )

                                # A_list.append(float(ave_timeA))
                            # print(ave_time)

        Xlist = []
        sub = []
        num = 0
        for i in A_list:
            num += 1
            A_dict = {
                "id": "A"+str(num),
                "type": "可靠性测试方案",
                "source": reliablityfile + "/平均恢复时间/系统失效恢复时间",
                "des": "由于第i次失效而重新启动,并恢复宕机的软件/系统所花费的总时间",
                "val": i
            }
            sub.append(A_dict)

        # B_dict = {
        #     "id": "B",
        #     "type": "可靠性测试方案",
        #     "source": reliablityfile + "/组件的冗余度/系统组件",
        #     "des": "系统组件数量",
        #     "val": len(B_list)
        # }

        res_dict = {
            "val": round(sum(A_list)/len(A_list), 3),
            "sub": sub
        }
    except:
        A_dict = {
            "id": "A",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/平均恢复时间/系统失效恢复时间",
            "des": "由于第i次失效而重新启动,并恢复宕机的软件/系统所花费的总时间",
            "val": i
        }

        B_dict = {
            "id": "B",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/组件的冗余度/系统组件",
            "des": "系统组件数量",
            "val": None
        }
        res_dict = {
            "val": None,
            "sub": [A_dict]
        }
    return res_dict


def rre2s(rTreeList):
    A_dict = {
        "id": "A",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/数据备份完整性/实际定期备份的数据项",
        "des": "实际定期备份数据项的数量",
        "val": None
    }
    A_data = {}
    Anum = 0
    B_dict = {
        "id": "B",
        "type": "可靠性测试方案",
        "source": reliablityfile + "/数据备份完整性/需要定期备份的数据项",
        "des": "需要备份的数据项的数量",
        "val": None
    }
    B_data = {}
    Bnum = 0
    try:
        for i in rTreeList:
            if i.text == '数据备份完整性':
                if '实际定期备份的数据项' in i.son_text:
                    idx = i.son_text.index('实际定期备份的数据项')
                    id = i.son_id[idx]
                    child = rTreeList[id]
                    for sonid in child.son_id:
                        cchild = rTreeList[sonid]
                        table = cchild.table_info
                        if table:
                            for row in range(1, len(table)):
                                content = table[row][5]
                                # pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                                # ave_time = pattern.findall(content)[0]
                                # contentslpit = content.split("-")
                                # content1,content2 = contentslpit[0],contentslpit[1]
                                # content1s = content1.split('/')
                                # content2s = content2.split(':')
                                # Atime = datetime.datetime(int(content1s[0]),int(content1s[1]),int(content1s[2]),
                                #                           int(content2s[0]),int(content2s[1]))

                                if '是' in content:
                                    Anum += 1
                            # print(ave_time)
                if '需要定期备份的数据项' in i.son_text:
                    idx = i.son_text.index('需要定期备份的数据项')
                    id = i.son_id[idx]
                    child = rTreeList[id]
                    for sonid in child.son_id:
                        cchild = rTreeList[sonid]
                        table = cchild.table_info
                        if table:
                            Bnum = len(table) - 1
                            # for row in range(1,len(table)):
                            #     # content = table[row][4]
                            #     # pattern = re.compile(r".*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*")
                            #     # ave_time = pattern.findall(content)[0]
                            #     content = table[row][1]
                            #     # contentslpit = content.split("-")
                            #     # content1,content2 = contentslpit[0],contentslpit[1]
                            #     # content1s = content1.split('/')
                            #     # content2s = content2.split(':')
                            #     # Btime = datetime.datetime(int(content1s[0]),int(content1s[1]),int(content1s[2]),
                            #     #                           int(content2s[0]),int(content2s[1]))
                            #
                            #     B_data[table[row][2]] = Btime
                            #     # B_data[table[row][2]] = 1

        A_dict = {
            "id": "A",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/数据备份完整性/实际定期备份的数据项",
            "des": "实际定期备份数据项的数量",
            "val": Anum
        }

        B_dict = {
            "id": "B",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/数据备份完整性/需要定期备份的数据项",
            "des": "需要备份的数据项的数量",
            "val": Bnum
        }

        res_dict = {
            "val": round(Anum/Bnum, 3),
            "sub": [A_dict, B_dict]
        }
    except:
        A_dict = {
            "id": "A",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/数据备份完整性/实际定期备份的数据项",
            "des": "实际定期备份数据项的数量",
            "val": None
        }

        B_dict = {
            "id": "B",
            "type": "可靠性测试方案",
            "source": reliablityfile + "/数据备份完整性/需要定期备份的数据项",
            "des": "需要备份的数据项的数量",
            "val": None
        }
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def sco1g(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "未经授权可访问的保密数据项的数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "需求文档",
        "source": reqfile + "/信息安全性",
        "des": "需要访问控制的保密数据项的数量",
        "val": None
    }
    try:
        B = 0
        gn = req[req['tag'] == '安全']
        gn = gn[gn['des'].str.contains('以下保密数据项设置访问控制')]
        for i in gn.itertuples():
            desc = getattr(i, 'des')
            B += getlistlens(desc, '以下保密数据项设置访问控制')
        # B = int(gn.count(axis=0)["r_id"])
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "执行结果", "实际结果"]], left_on="tc_id", right_on="用例编号")
        A = int(all[(all["实际结果"].str.contains("访问控制")) & (all["执行结果"] != "通过")].tc_id.nunique())
        uid = getuid(all)
        A_dict['source'] += uid
        A_dict['val'] = A
        B_dict['val'] = B
        res_dict = {
            "val": 1 - A_dict["val"] / B_dict["val"] if B else 0,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def sco2g(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "正确加密/解密的数据项数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "需求文档",
        "source": reqfile + "/信息安全性",
        "des": "需要加密/解密的数据项数量",
        "val": None
    }
    try:
        gn = req[req['tag'] == '安全']
        gn = gn[gn['des'].str.contains('以下数据项设置加密/解密')]
        B = 0
        for i in gn.itertuples():
            desc = getattr(i, 'des')
            B += getlistlens(desc, '以下数据项设置加密/解密')
        # B = int(gn.count(axis=0)["r_id"])
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "执行结果", "实际结果"]], left_on="tc_id", right_on="用例编号")
        A = int(all[(all["实际结果"].str.contains("加密/解密")) & (all["执行结果"] == "通过")].tc_id.nunique())
        uid = getuid(all)
        A_dict['source'] += uid
        A_dict['val'] = A
        B_dict['val'] = B
        res_dict = {
            "val": A_dict["val"] / B_dict["val"] if B else 0,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def sco3s(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "使用时遭到破坏或存在不可接受风险的加密算法的数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "需求文档",
        "source": reqfile + "/信息安全性",
        "des": "所使用的加密算法的数量",
        "val": None
    }
    try:
        gn = req[req['tag'] == '安全']
        gn = gn[gn['des'].str.contains('使用的加密/解密算法经过严格审查')]
        B = int(gn.count(axis=0)["r_id"])
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "执行结果", "实际结果"]], left_on="tc_id", right_on="用例编号")
        A = int(all[(all["实际结果"].str.contains("使用的加密/解密算法经过严格审查")) & (all["执行结果"] != "通过")].tc_id.nunique())
        uid = getuid(all)
        A_dict['source'] += uid
        A_dict['val'] = A
        B_dict['val'] = B
        res_dict = {
            "val": 1-A_dict["val"] / B_dict["val"] if B else 0,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def sin1g(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "因未经授权访问而破坏或篡改数据项的数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "需求文档",
        "source": reqfile + "/信息安全性",
        "des": "需要避免数据破坏或复改的数据项数量",
        "val": None
    }
    try:
        gn = req[req['tag'] == '安全']
        gn = gn[gn['des'].str.contains('避免数据破坏或篡改')]
        B = 0
        for i in gn.itertuples():
            desc = getattr(i, 'des')
            B += getlistlens(desc, '避免数据破坏或篡改')
        # B = int(gn.count(axis=0)["r_id"])
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "执行结果", "实际结果"]], left_on="tc_id", right_on="用例编号")
        A = int(all[(all["实际结果"].str.contains("避免数据破坏或篡改")) & (all["执行结果"] != "通过")].tc_id.nunique())
        uid = getuid(all)
        A_dict['source'] += uid
        A_dict['val'] = A
        B_dict['val'] = B
        res_dict = {
            "val": 1-A_dict["val"] / B_dict["val"] if B else 0,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def sin2g(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "实际用于数据抗讹误性方法的数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "需求文档",
        "source": reqfile + "/信息安全性",
        "des": "可用及推荐的用于数据抗讹误性方法的数量",
        "val": None
    }
    try:
        gn = req[req['tag'] == '安全']
        gn = gn[gn['des'].str.contains('采用以下方法用于数据抗讹误')]
        B = 0
        for i in gn.itertuples():
            desc = getattr(i, 'des')
            B += getlistlens(desc, '采用以下方法用于数据抗讹误')
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "执行结果", "实际结果"]], left_on="tc_id", right_on="用例编号")
        A = int(all[(all["实际结果"].str.contains("采用以下方法用于数据抗讹误")) & (all["执行结果"] == "通过")].tc_id.nunique())
        uid = getuid(all)
        A_dict['source'] += uid
        A_dict['val'] = A
        B_dict['val'] = B
        res_dict = {
            "val": A_dict["val"] / B_dict["val"] if B else 0,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def sno1g(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "实际使用数字签名确保抗抵赖性事务的数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "需求文档",
        "source": reqfile + "/信息安全性",
        "des": "使用数字签名要求抗抵赖性事务的数量",
        "val": None
    }
    try:
        gn = req[req['tag'] == '安全']
        gn = gn[gn['des'].str.contains('使用数字签名确保抗抵赖性')]
        B = 0
        for i in gn.itertuples():
            desc = getattr(i, 'des')
            B += getlistlens(desc, '使用数字签名确保抗抵赖性')
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "执行结果", "实际结果"]], left_on="tc_id", right_on="用例编号")
        A = int(all[(all["实际结果"].str.contains("数字签名")) & (all["执行结果"] == "通过")].tc_id.nunique())
        uid = getuid(all)
        A_dict['source'] += uid
        A_dict['val'] = A
        B_dict['val'] = B
        res_dict = {
            "val": A_dict["val"] / B_dict["val"] if B else 0,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def sac1g(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "所有日志中记录的访问次数",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "对系统或数据的访问次数",
        "val": None
    }
    try:
        gn = req[req['tag'] == '安全']
        gn = gn[gn['des'].str.contains('对系统或数据的访问以日志形式记录')]
        # B = int(gn.count(axis=0)["r_id"])
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "执行结果", "实际结果"]], left_on="tc_id", right_on="用例编号")
        # A = int(all[(all["实际结果"].str.contains("数字签名")) & (all["执行结果"] == "通过")].tc_id.nunique())
        B = 0
        A = 0
        uid = getuid(all)
        A_dict['source'] += uid
        B_dict['source'] += uid
        for i in all.itertuples():
            resdes = getattr(i, "实际结果")
            if '系统或数据的访问次数' in resdes:
                idx = resdes.index('系统或数据的访问次数')
                pattern = re.compile(r'.*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*')
                B += int(pattern.findall(resdes[idx:])[0])
            if '日志记录的访问次数' in resdes:
                idx = resdes.index('日志记录的访问次数')
                pattern = re.compile(r'.*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*')
                A += int(pattern.findall(resdes[idx:])[0])

        A_dict['val'] = A
        B_dict['val'] = B
        res_dict = {
            "val": A_dict["val"] / B_dict["val"] if B else 0,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def sac2s(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "要求系统日志存储在稳定存储器中的时间",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "需求文档",
        "source": reqfile + "/信息安全性",
        "des": "要求系统日志存储在稳定存储器中的时间",
        "val": None
    }
    try:
        gn = req[req['tag'] == '安全']
        gn = gn[gn['des'].str.contains('日志存储在稳定存储器中时间')]
        Blist = []
        for i in gn.itertuples():
            desc = getattr(i, 'des')
            tmp = gettime(desc)
            Blist.append(tmp)
        if Blist:
            B = average(Blist)
        else:
            B = 0
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "执行结果", "实际结果"]], left_on="tc_id", right_on="用例编号")
        # A = int(all[(all["实际结果"].str.contains("数字签名")) & (all["执行结果"] == "通过")].tc_id.nunique())
        # A_dict['val'] = A
        uid = getuid(all)
        A_dict['source'] += uid
        # B_dict['source'] += uid
        Alist = []
        for j in all.itertuples():
            desc = getattr(j, '实际结果')
            if '日志实际存储时间' in desc:
                tmp = gettime(desc)
                Alist.append(tmp)
        if Alist:
            A = average(Alist)
        else:
            A = 0
        B_dict['val'] = B
        A_dict['val'] = A
        res_dict = {
            "val": round(A_dict["val"] / B_dict["val"], 2) if B else 0,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def sau1g(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "提供鉴别机制的数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "需求文档",
        "source": reqfile + "/信息安全性",
        "des": "规定的鉴别机制数量",
        "val": None
    }
    try:
        gn = req[req['tag'] == '安全']
        gn = gn[gn['des'].str.contains('提供以下鉴别机制')]
        B = 0
        for i in gn.itertuples():
            desc = getattr(i, 'des')
            B += getlistlens(desc, '提供以下鉴别机制')
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "执行结果", "实际结果"]], left_on="tc_id", right_on="用例编号")
        A = int(all[(all["实际结果"].str.contains("鉴别机制")) & (all["执行结果"] == "通过")].tc_id.nunique())
        # A_dict['val'] = A
        uid = getuid(all)
        A_dict['source'] += uid
        # B_dict['source'] += uid

        B_dict['val'] = B
        A_dict['val'] = A
        res_dict = {
            "val": round(A_dict["val"] / B_dict["val"], 2) if B else 0,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def sau2s(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "已实现的鉴别规则的数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "需求文档",
        "source": reqfile + "/信息安全性",
        "des": "规定的鉴别规则的数量",
        "val": None
    }
    try:
        gn = req[req['tag'] == '安全']
        gn = gn[gn['des'].str.contains('包含以下鉴别规则')]
        B = 0
        for i in gn.itertuples():
            desc = getattr(i, 'des')
            B += getlistlens(desc, '包含以下鉴别规则')
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "执行结果", "实际结果"]], left_on="tc_id", right_on="用例编号")
        A = int(all[(all["实际结果"].str.contains("鉴别规则")) & (all["执行结果"] == "通过")].tc_id.nunique())
        # A_dict['val'] = A
        uid = getuid(all)
        A_dict['source'] += uid
        # B_dict['source'] += uid

        B_dict['val'] = B
        A_dict['val'] = A
        res_dict = {
            "val": round(A_dict["val"] / B_dict["val"], 2) if B else 0,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def scl1g(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "在评价中证实的已正确实现与信息安全性依从相关的项数",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "与信息安全性的依从性相关的项数",
        "val": None
    }
    try:
        xl = req[req["tag"] == "安全"]
        xl = xl[xl['standard'] is not None & xl['standard'] != '无']
        joined = xl.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '测试步骤', '测试类型', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")

        A_dict['val'] = all[all["执行结果"] == "通过"].shape[0]
        B_dict["val"] = all.shape[0]
        uid = getuid(all)
        A_dict['source'] += uid
        B_dict['source'] += uid
        res_dict = {
            "val": round(A_dict["val"] / B_dict["val"], 2) if B_dict["val"] else 0,
            "sub": [A_dict, B_dict]
        }

    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def pcp1g(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "在评价中证实的已正确实现与可移植性依从相关的项数",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "与可移植性的依从性相关的项数",
        "val": None
    }
    try:
        xl = req[req["tag"] == "可移植"]
        xl = xl[xl['standard'] is not None & xl['standard'] != '无']
        joined = xl.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '测试步骤', '测试类型', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")

        A_dict['val'] = all[all["执行结果"] == "通过"].shape[0]
        B_dict["val"] = all.shape[0]
        uid = getuid(all)
        A_dict['source'] += uid
        B_dict['source'] += uid
        res_dict = {
            "val": round(A_dict["val"] / B_dict["val"], 2) if B_dict["val"] else 0,
            "sub": [A_dict, B_dict]
        }

    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def pad1g(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "测试期间未完成或结果没有达到要求的功能数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "需求文档",
        "source": reqfile + "/可移植性",
        "des": "不同硬件环境中需要测试的功能数量",
        "val": None
    }
    try:
        gn = req[req['tag'] == '可移植']
        gn = gn[gn['des'].str.contains('以下硬件环境中功能达到要求')]
        B = 0
        # for i in gn.itertuples():
        #     desc = getattr(i, 'des')
        #     B += getlistlens(desc, '以下硬件环境中功能达到要求')
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "执行结果", "预期结果"]], left_on="tc_id", right_on="用例编号")
        B = all.shape[0]
        A = int(all[(all["预期结果"].str.contains("硬件环境")) & (all["执行结果"] != "通过")].tc_id.nunique())
        uid = getuid(all)
        A_dict['source'] += uid
        # B_dict['source'] += uid

        B_dict['val'] = B
        A_dict['val'] = A
        res_dict = {
            "val": 1 - round(A_dict["val"] / B_dict["val"], 2) if B else 0,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def pad2g(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "测试期间未完成或结果没有达到要求的功能数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "需求文档",
        "source": reqfile + "/可移植性",
        "des": "不同系统软件环境下需要测试的功能数量",
        "val": None
    }
    try:
        gn = req[req['tag'] == '可移植']
        gn = gn[gn['des'].str.contains('以下软件环境中功能达到要求')]
        B = 0
        # for i in gn.itertuples():
        #     desc = getattr(i, 'des')
        #     B += getlistlens(desc, '以下软件环境中功能达到要求')
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "执行结果", "预期结果"]], left_on="tc_id", right_on="用例编号")
        B = all.shape[0]
        A = int(all[(all["预期结果"].str.contains("软件环境")) & (all["执行结果"] != "通过")].tc_id.nunique())
        uid = getuid(all)
        A_dict['source'] += uid
        # B_dict['source'] += uid

        B_dict['val'] = B
        A_dict['val'] = A
        res_dict = {
            "val": round(1-A_dict["val"] / B_dict["val"], 2) if B else 0,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def pad3s(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "在带有用户环境的运营测试中,测试期间没有完成或结果没有达到要求的功能数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "需求文档",
        "source": reqfile + "/可移植性",
        "des": "在不同运营环境中测试的功能数量",
        "val": None
    }
    try:
        gn = req[req['tag'] == '可移植']
        gn = gn[gn['des'].str.contains('以下运营环境中功能达到要求')]
        B = 0
        # for i in gn.itertuples():
        #     desc = getattr(i, 'des')
        #     B += getlistlens(desc, '以下软件环境中功能达到要求')
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "执行结果", "预期结果"]], left_on="tc_id", right_on="用例编号")
        B = all.shape[0]
        A = int(all[(all["预期结果"].str.contains("运营环境")) & (all["执行结果"] != "通过")].tc_id.nunique())
        uid = getuid(all)
        A_dict['source'] += uid
        # B_dict['source'] += uid

        B_dict['val'] = B
        A_dict['val'] = A
        res_dict = {
            "val": round(1-A_dict["val"] / B_dict["val"], 2) if B else 0,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def pin1g(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "第i次安装所消耗的总工作时间",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "第i次安装的预期时间",
        "val": None
    }
    try:
        gn = req[req['tag'] == '可移植']
        gn = gn[gn['des'].str.contains('安装所需时间')]
        B = 0
        # for i in gn.itertuples():
        #     desc = getattr(i, 'des')
        #     B += getlistlens(desc, '以下软件环境中功能达到要求')
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "实际结果", "预期结果"]], left_on="tc_id", right_on="用例编号")
        sub = []
        xlist = []
        index = 0

        for row in all.itertuples():
            index += 1
            i = getattr(row, "实际结果")
            Atime = gettime(i)
            j = getattr(row, "预期结果")
            Btime = gettime(j)
            A_dict = {
                "id": "A" + str(index),
                "type": "测试文档",
                "source": testfile + '/用例编号:' + getattr(row, "用例编号"),
                "des": "第i次安装所消耗的总工作时间",
                "val": Atime
            }
            B_dict = {
                "id": "B"+str(index),
                "type": "测试文档",
                "source": testfile + '/用例编号:' + getattr(row, "用例编号"),
                "des": "第i次安装的预期时间",
                "val": Btime
            }
            # uid = getuid(all)
            # A_dict['source'] += uid
            # B_dict['source'] += uid
            x = Atime/Btime if Btime else 0
            xlist.append(x)
            sub.append(A_dict)
            sub.append(B_dict)

        # B = all.shape[0]
        # A = int(all[(all["预期结果"].str.contains("运营环境")) & (all["执行结果"] != "通过")].tc_id.nunique())

        # B_dict['val'] = B
        # A_dict['val'] = A
        res_dict = {
            "val": round(average(xlist), 2) if xlist else 0,
            "sub": sub
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def pin2g(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "用户成功自定义安装规程的数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "需求文档",
        "source": reqfile + "/可移植性",
        "des": "为使用方便,用户尝试自定义安装规程的数量",
        "val": None
    }
    try:
        gn = req[req['tag'] == '可移植']
        gn = gn[gn['des'].str.contains('自定义安装规程的数量')]
        B = 0
        A = 0

        # for i in gn.itertuples():
        #     desc = getattr(i, 'des')
        #     B += getlistlens(desc, '以下软件环境中功能达到要求')
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "实际结果", "预期结果"]], left_on="tc_id", right_on="用例编号")
        uid = getuid(all)
        A_dict['source'] += uid
        # B_dict['source'] += uid
        for i in all.itertuples():
            resdes = getattr(i, 'des')
            idx = resdes.index('自定义安装规程的数量')
            pattern = re.compile(r'.*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*')
            B += int(pattern.findall(resdes[idx:])[0])
            result = getattr(i, '实际结果')
            idx = result.index('自定义安装规程')
            pattern = re.compile(r'.*?([1-9]\d*\.\d+|0\.\d+|0|[1-9]\d*).*')
            A += int(pattern.findall(result[idx:])[0])

        # A = int(all[(all["预期结果"].str.contains("运营环境")) & (all["执行结果"] != "通过")].tc_id.nunique())

        B_dict['val'] = B
        A_dict['val'] = A
        res_dict = {
            "val": round(A_dict["val"] / B_dict["val"], 2) if B else 0,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def pre1g(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "替换原软件产品后，本软件产品在没有任何额外的学习或变通的情况下,能够执行的用户功能数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "需求文档",
        "source": reqfile + "/可移植性",
        "des": "替换原软件产品后，本软件产品中用户功能的数量",
        "val": None
    }
    try:
        gn = req[req['tag'] == '可移植']
        gn = gn[gn['des'].str.contains(r"以下用户功能的执行与(\w+)一致", regex=True)]
        B = 0
        A = 0
        for i in gn.itertuples():
            desc = str(getattr(i, 'des'))
            B += getlistlens(desc, '以下用户功能的执行与')
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "实际结果", "执行结果"]], left_on="tc_id", right_on="用例编号")
        A = int(all[(all["执行结果"] == "通过")].tc_id.nunique())
        uid = getuid(all)
        A_dict['source'] += uid
        # B_dict['source'] += uid

        # A = int(all[(all["预期结果"].str.contains("运营环境")) & (all["执行结果"] != "通过")].tc_id.nunique())

        B_dict['val'] = B
        A_dict['val'] = A
        res_dict = {
            "val": round(A_dict["val"] / B_dict["val"], 2) if B else 0,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def pre2s(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "优于或等于被替换产品的新产品质量测度数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "需求文档",
        "source": reqfile + "/可移植性/",
        "des": "被替换软件产品中的质量测度数量",
        "val": None
    }
    try:
        gn = req[req['tag'] == '可移植']
        gn = gn[gn['des'].str.contains("以下质量测度优于或等于")]
        B = 0
        A = 0
        for i in gn.itertuples():
            desc = str(getattr(i, 'des'))
            B += getlistlens(desc, '以下质量测度优于或等于')
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "预期结果", "执行结果"]], left_on="tc_id", right_on="用例编号")
        A = int(all[all['预期结果'].str.contains('质量测度优于或等于') & (all["执行结果"] == "通过")].tc_id.nunique())
        uid = getuid(all)
        A_dict['source'] += uid
        # B_dict['source'] += uid

        B_dict['val'] = B
        A_dict['val'] = A
        res_dict = {
            "val": round(A_dict["val"] / B_dict["val"], 2) if B else 0,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def pre3s(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "结果与被替换软件产品相似的产品功能数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "需求文档",
        "source": reqfile + "/可移植性",
        "des": "被替换软件产品中需要使用的功能数量",
        "val": None
    }
    try:
        gn = req[req['tag'] == '可移植']
        gn = gn[gn['des'].str.contains(r"保留(\w+)的以下产品功能", regex=True)]
        B = 0
        A = 0
        for i in gn.itertuples():
            desc = str(getattr(i, 'des'))
            B += getlistlens(desc, '产品功能')
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "预期结果", "执行结果"]], left_on="tc_id", right_on="用例编号")
        A = int(all[all['预期结果'].str.contains('保留') & (all["执行结果"] == "通过")].tc_id.nunique())
        uid = getuid(all)
        A_dict['source'] += uid
        # B_dict['source'] += uid

        B_dict['val'] = B
        A_dict['val'] = A
        res_dict = {
            "val": round(A_dict["val"] / B_dict["val"], 2) if B else 0,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def pre4s(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "能像被替换软件产品一样继续使用的数据数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "需求文档",
        "source": reqfile + "/可移植性",
        "des": "被替换软件产品中需要继续使用的数据数量",
        "val": None
    }
    try:
        gn = req[req['tag'] == '可移植']
        gn = gn[gn['des'].str.contains("使用以下数据")]
        B = 0
        A = 0
        for i in gn.itertuples():
            desc = str(getattr(i, 'des'))
            B += getlistlens(desc, '使用以下数据')
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "预期结果", "执行结果"]], left_on="tc_id", right_on="用例编号")
        A = int(all[all['预期结果'].str.contains('数据') & (all["执行结果"] == "通过")].tc_id.nunique())
        uid = getuid(all)
        A_dict['source'] += uid
        # B_dict['source'] += uid

        B_dict['val'] = B
        A_dict['val'] = A
        res_dict = {
            "val": round(A_dict["val"] / B_dict["val"], 2) if B else 0,
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


def ule2s(dTreeList, req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "/",
        "des": "运行过程中自动填充默认值的输人字段数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "设计文档",
        "source": designfile + "/模块设计/输入项",
        "des": "具有默认值的输人字段的数量",
        "val": None
    }
    Bnum = 0
    Anum = 0
    try:
        gn = req[req['tag'] == '易用']
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "测试步骤", "执行结果"]], left_on="tc_id", right_on="用例编号")
        Aid = ''
        for i in all.itertuples():
            if '自动填充' in getattr(i, '测试步骤') and getattr(i, '执行结果') == '通过':
                Aid = Aid + getattr(i, '用例编号') + ';'
                Anum += 1
        for node in dTreeList:
            if node.text == '系统模块设计':
                for id in node.son_id:
                    child = dTreeList[id]
                    if '模块设计' in child.son_text:
                        idx = child.son_text.index('模块设计')
                        cchild = dTreeList[child.son_id[idx]]
                        if '输入项' in cchild.son_text:
                            cidx = cchild.son_text.index('输入项')
                            ccchild = dTreeList[cchild.son_id[cidx]]
                            for ccid in ccchild.son_id:
                                cccchild = dTreeList[ccid]
                                table = cccchild.table_info
                                if table:
                                    for values in table.values():
                                        # print(values)
                                        if values[0] == '输入的方式' and '自动填充' in values[1]:
                                            Bnum += 1

                                for cccid in cccchild.son_id:
                                    table = dTreeList[cccid].table_info
                                    if table:
                                        for values in table.values():
                                            # print(values)
                                            if values[0] == '输入的方式' and '自动填充' in values[1]:
                                                Bnum += 1

                break

        A_dict = {
            "id": "A",
            "type": "测试文档",
            "source": testfile + '/用例编号:',
            "des": "运行过程中自动填充默认值的输人字段数量",
            "val": Anum
        }
        B_dict = {
            "id": "B",
            "type": "设计文档",
            "source": designfile + "/模块设计/输入项",
            "des": "具有默认值的输人字段的数量",
            "val": Bnum
        }
        uid = getuid(all)
        A_dict['source'] += Aid
        # B_dict['source'] += uid
        res_dict = {

            "val": round(Anum / Bnum, 3),

            "sub": [A_dict, B_dict]

        }

    except:

        res_dict = {

            "val": None,

            "sub": [A_dict, B_dict]

        }

    return res_dict


def uop4s(dTreeList, req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "/",
        "des": "可以定制的用户界面元素数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "设计文档",
        "source": designfile + "/界面处理流程及界面约束描述/用户定制",
        "des": "期望能够受益于定制的用户界面元素数量",
        "val": None
    }
    Bnum = 0
    Anum = 0
    try:
        gn = req[req['tag'] == '易用']
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "实际结果", "执行结果"]], left_on="tc_id", right_on="用例编号")
        Aid = ''
        for i in all.itertuples():
            if '用户定制' in getattr(i, '实际结果') and getattr(i, '执行结果') == '通过':
                Aid = Aid + getattr(i, '用例编号') + ';'
                Anum += 1
        for node in dTreeList:
            if node.text == '界面处理流程及界面约束描述':
                for id in node.son_id:
                    child = dTreeList[id]
                    for cid in child.son_id:
                        cchild = dTreeList[cid]
                        table = cchild.table_info
                        if table:
                            if '用户定制' in table[0]:
                                idx = table[0].index('用户定制')
                                for row in range(1, len(table)):
                                    content = table[row][idx]
                                    if '是' in content:
                                        Bnum += 1
                        for ccid in cchild.son_id:
                            ccchild = dTreeList[ccid]
                            table = ccchild.table_info
                            if table:
                                if '用户定制' in table[0]:
                                    idx = table[0].index('用户定制')
                                    for row in range(1, len(table)):
                                        content = table[row][idx]
                                        if '是' in content:
                                            Bnum += 1

                break
        A_dict = {
            "id": "A",
            "type": "测试文档",
            "source": testfile + '/用例编号' + ":",
            "des": "可以定制的用户界面元素数量",
            "val": Anum
        }
        B_dict = {
            "id": "B",
            "type": "设计文档",
            "source": designfile + "/界面处理流程及界面约束描述/用户定制",
            "des": "期望能够受益于定制的用户界面元素数量",
            "val": Bnum
        }
        uid = getuid(all)
        print(uid)
        A_dict['source'] += Aid
        # B_dict['source'] += uid
        res_dict = {

            "val": round(Anum / Bnum, 3),

            "sub": [A_dict, B_dict]

        }

    except:

        res_dict = {

            "val": None,

            "sub": [A_dict, B_dict]

        }

    return res_dict


def uin1s(dTreeList, req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "在外观舒适性上令人愉悦的显示界面数量",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "设计文档",
        "source": designfile + "/用户界面与模块关系表",
        "des": "显示界面数量",
        "val": None
    }
    Bnum = 0
    Anum = 0
    try:
        gn = req[req['tag'] == '易用']
        joined = gn.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", "实际结果", "执行结果"]], left_on="tc_id", right_on="用例编号")
        Aid = ''
        for i in all.itertuples():
            if '外观舒适' in getattr(i, '实际结果') and getattr(i, '执行结果') == '通过':
                Aid = Aid + getattr(i, '用例编号') + ';'
                Anum += 1
        for node in dTreeList:
            if node.text == '用户界面与模块关系表':
                for id in node.son_id:
                    child = dTreeList[id]
                    table = child.table_info
                    # print(table)
                    if table:
                        Bnum += len(table) - 1
                break
        A_dict = {
            "id": "A",
            "type": "测试文档",
            "source": testfile + '/用例编号:' + "",
            "des": "在外观舒适性上令人愉悦的显示界面数量",
            "val": Anum
        }
        B_dict = {
            "id": "B",
            "type": "设计文档",
            "source": designfile + "/用户界面与模块关系表",
            "des": "显示界面数量",
            "val": Bnum
        }
        uid = getuid(all)
        A_dict['source'] += Aid
        # B_dict['source'] += uid
        res_dict = {
            "val": round(Anum/Bnum, 3),
            "sub": [A_dict, B_dict]
        }
    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }

    return res_dict


def rcl1g(req, trace, testcases):
    A_dict = {
        "id": "A",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "在评价中证实的已正确实现与可靠性依从相关的项数",
        "val": None
    }
    B_dict = {
        "id": "B",
        "type": "测试文档",
        "source": testfile + '/用例编号' + "",
        "des": "与可靠性的依从性相关的项数",
        "val": None
    }
    try:
        xl = req[req["tag"] == "可靠"]
        xl = xl[xl['standard'] is not None & xl['standard'] != '无']
        joined = xl.merge(trace[["r_id", "tc_id"]], on="r_id", how="inner")
        all = joined.merge(testcases[["用例编号", '测试步骤', '测试类型', "执行结果"]], left_on="tc_id",
                           right_on="用例编号")

        A_dict['val'] = all[all["执行结果"] == "通过"].shape[0]
        B_dict["val"] = all.shape[0]
        uid = getuid(all)
        A_dict['source'] += uid
        B_dict['source'] += uid
        res_dict = {
            "val": round(A_dict["val"] / B_dict["val"], 2) if B_dict["val"] else 0,
            "sub": [A_dict, B_dict]
        }

    except:
        res_dict = {
            "val": None,
            "sub": [A_dict, B_dict]
        }
    return res_dict


class MaintainTestFileExtract():
    def __init__(self, UserManualFile: str, MaintainTestFile: str,
                 demandMtrix: str, getTestManual: str, usermanual_rTreeList, maintaintest_rTreeList, yanshouceshi_file):
        self.UserManualFile = UserManualFile
        self.MaintainTestFile = MaintainTestFile
        self.manualDocx = usermanual_rTreeList
        self.testDocx = maintaintest_rTreeList
        self.yanshouceshi_file = yanshouceshi_file

        self.trace = get_metric(demandMtrix)
        self.testcases = get_testcase(getTestManual)

        self.manual_root = self.manualDocx[0]
        for n in self.manualDocx:
            if n.text == '维护性：“四规”，在规定条件下，规定的时间内，使用规定的工具或方法修复规定功能的能力;':
                self.manual_root = n
                break

        self.test_root = self.testDocx[0]
        for n in self.testDocx:
            if n.text == '测试方法':
                self.test_root = n
                break
        # self.manual_root.node_info_print()
        # self.test_root.node_info_print()

        self.mmo1g()
        self.mmo2s()
        self.mre1g()
        self.mre2s()

        self.man1g()
        self.man2s()
        self.man3s()
        self.mmd1g()
        self.mmd2g()
        self.mmd3s()
        self.mte1g()
        self.mte2s()
        self.mte3s()
        self.mcl1g()
        self.mcl1g()

    def check(self, node):
        flag = False
        for i in node:
            if i == self.test_root:
                flag = True
                break
        if flag:
            return True
        return False

    def mmo1g(self):
        A, B = 0, 0
        user_demand_id = ''
        B_dict = {
            "id": "B",
            "type": "软件需求规格说明书",
            "source": self.UserManualFile + "/系统需求说明/非功能需求/维护性",
            "des": "软件需求规格说明书中规定的独立组件的数量",
            "val": 0
        }
        A_dict = {
            "id": "A",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/模块化测度测试方法/组件间的耦合度计算",
            "des": "软件模块中独立组件的数量",
            "val": 0
        }
        node = self.manual_root
        flag = False

        for i in self.manual_root.son_id:
            # print('son_id = ', i)
            # print(self.manualDocx[int(i)].table_info)
            # print(self.manualDocx[int(i)].table_info[0][1])
            if self.manualDocx[int(i)].table_info[2][1].__contains__("独立的组件数量"):
                user_demand_id = self.manualDocx[int(i)].table_info[0][1]
                node = self.manualDocx[i]
                flag = True
                break

        # 如果需求文档中没有该项，则直接返回
        if not flag:
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

        B_text = node.table_info[2][1]
        B = re.search(r'独立的组件数量至少有(\d+)个', B_text).group(1)

        node_3_1 = self.test_root
        for i in self.test_root.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '模块化测度测试方法':
                node_3_1 = self.testDocx[int(i)]
                flag = True
                break
        node_3_1_1 = self.test_root
        for i in node_3_1.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '组件间的耦合度':
                node_3_1_1 = self.testDocx[int(i)]
                break

        node_3_1_1_1 = self.test_root
        for i in node_3_1_1.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '系统独立组件':
                node_3_1_1_1 = self.testDocx[int(i)]
                break
        # print(node_3_1_1_1.son_id[0])
        if self.check([node_3_1, node_3_1_1, node_3_1_1_1]):
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

        table_node = self.testDocx[int(node_3_1_1_1.son_id[0])]

        # print('table_node[2] = {0}'.format(table_node.text))
        table_tmp = eval(table_node.text)
        A = len(table_tmp) - 1
        A_dict['val'] = A
        B_dict['val'] = B
        # print('A = {0}, B = {1}'.format(A, B))
        if A and B:
            res_dict = {
                "val": int(A)/int(B),
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

    def mmo2s(self):
        A, B = 0, 0
        user_demand_id = ''
        B_dict = {
            "id": "B",
            "type": "软件需求规格说明书",
            "source": self.UserManualFile + "/系统需求说明/非功能需求/维护性",
            "des": "软件需求规格说明书中规定的圈复杂度阈值",
            "val": 0
        }
        A_dict = {
            "id": "A",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/模块化测度测试方法/圈复杂度",
            "des": "软件模块的圈复杂度",
            "val": 0
        }

        node = self.manual_root
        flag = False
        for i in self.manual_root.son_id:
            # print('son_id = ', i)
            # print(self.manualDocx[int(i)].table_info)
            # print(self.manualDocx[int(i)].table_info[0][1])
            if self.manualDocx[int(i)].table_info[2][1].__contains__('模块的圈复杂度得分'):
                user_demand_id = self.manualDocx[int(i)].table_info[0][1]
                node = self.manualDocx[i]
                flag = True
                break
        # print('user_demand_id = ', user_demand_id)
        if flag:
            # print('node.id = ',node.id)
            # print('node.text = ',node.text)
            # print('node.table_info = ',node.table_info)
            B_text = node.table_info[2][1]
            B = re.search(r'圈复杂度得分低于(\d+)，', B_text).group(1)
            # print(B)

        node_3_1 = self.test_root
        for i in self.test_root.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '模块化测度测试方法':
                node_3_1 = self.testDocx[int(i)]
                break
        node_3_1_2 = self.test_root
        for i in node_3_1.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '圈复杂度':
                node_3_1_2 = self.testDocx[int(i)]
                break

        node_3_1_2_1 = self.test_root
        for i in node_3_1_2.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '圈复杂度阈值':
                node_3_1_2_1 = self.testDocx[int(i)]
                break

        if self.check([node_3_1, node_3_1_2, node_3_1_2_1]):
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict
        # print(node_3_1_2_1.son_id[0])

        table_node = self.testDocx[int(node_3_1_2_1.son_id[0])]

        # print('table_node[2] = {0}'.format(table_node.text))
        table_tmp = eval(table_node.text)
        # print('table_tmp=',table_tmp)
        for i in range(1, len(table_tmp)):
            if table_tmp[i][5] > B:
                A += 1

        # print('A = {0}, B = {1}'.format(A, B))
        A_dict['val'] = A
        B_dict['val'] = B
        if A and B:
            res_dict = {
                "val": int(A)/int(B),
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

    def mre1g(self):
        A, B = 0, 0
        user_demand_id = ''
        B_dict = {
            "id": "B",
            "type": "软件需求规格说明书",
            "source": self.UserManualFile + "/系统需求说明/非功能需求/维护性",
            "des": "软件需求规格说明书中规定的函数被调用次数阈值",
            "val": 0
        }
        A_dict = {
            "id": "A",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/可重用性测度测试方法/资产的可重用性",
            "des": "软件的函数被调用次数",
            "val": 0
        }

        node = self.manual_root
        flag = False
        for i in self.manual_root.son_id:
            if self.manualDocx[int(i)].table_info[2][1].__contains__('函数调用次数'):
                user_demand_id = self.manualDocx[int(i)].table_info[0][1]
                node = self.manualDocx[i]
                flag = True
                break

        if flag:
            B_text = node.table_info[2][1]
            B = re.search(r'函数调用次数大于(\d+)，', B_text).group(1)
            # print("B = ", B)

        node_3_2 = self.test_root
        for i in self.test_root.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '可重用性测度测试方法':
                node_3_2 = self.testDocx[int(i)]
                break
        node_3_2_1 = self.test_root
        for i in node_3_2.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '资产的可重用性':
                node_3_2_1 = self.testDocx[int(i)]
                break

        node_3_2_1_2 = self.test_root
        for i in node_3_2_1.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '函数调用图中调用次数':
                node_3_2_1_2 = self.testDocx[int(i)]
                break

        # print(node_3_2_1_2.son_id[0])
        if self.check([node_3_2, node_3_2_1, node_3_2_1_2]):
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict
        table_node = self.testDocx[int(node_3_2_1_2.son_id[0])]

        # print('table_node[2] = {0}'.format(table_node.text))
        table_tmp = eval(table_node.text)
        # print('table_tmp=',table_tmp)
        for i in range(1, len(table_tmp)):
            if table_tmp[i][2] > B:
                A += 1

        # print('A = {0}, B = {1}'.format(A, B))
        A_dict['val'] = A
        B_dict['val'] = B
        if A and B:
            res_dict = {
                "val": int(A)/int(B),
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

    # OK
    def mre2s(self):
        user_demand_id = ''
        node = self.manual_root
        flag = False
        for i in self.manual_root.son_id:
            if self.manualDocx[int(i)].table_info[2][1].__contains__('符合编码规则的模块'):
                user_demand_id = self.manualDocx[int(i)].table_info[0][1]
                node = self.manualDocx[i]
                flag = True
                break

        software_demand_id = user_demand_id
        filtered_rows = self.trace[self.trace['软件需求编号'] == user_demand_id]
        test_case_id = filtered_rows.iloc[0]['tc_id']
        filtered_rows = self.testcases[self.testcases['用例编号'] == test_case_id]

        A, B = 0, 0
        if flag:
            B_text = node.table_info[2][1]
            B = re.search(r'符合编码规则的模块不少于(\d+)，', B_text).group(1)
            # print("B = ", B)

        # print(filtered_rows)
        for i in filtered_rows.index:
            print(i)
            if filtered_rows['执行结果'][i] == '通过':
                A += 1
        # print('software_demand_id = {0} \n test_case_id = {1}'.format(software_demand_id,test_case_id))
        # print('A=',A)
        B_dict = {
            "id": "B",
            "type": "软件需求规格说明书",
            "source": self.UserManualFile + "/系统需求说明/非功能需求/维护性",
            "des": "软件需求规格说明书中规定的需要遵循编码规则的模块最少数",
            "val": B
        }
        A_dict = {
            "id": "A",
            "type": "验收测试用例",
            "source": self.yanshouceshi_file + "/用例编号/" + str(user_demand_id),
            "des": "软件符合编码规则的模块数",
            "val": A
        }

        if A and B:
            res_dict = {
                "val": int(A)/int(B),
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

    def man1g(self):
        A, B = 0, 0
        B_dict = {
            "id": "B",
            "type": "维护性测试方案",
            "source": self.UserManualFile + "/系统需求说明/非功能需求/维护性",
            "des": "审计所需日志条数",
            "val": 0
        }
        A_dict = {
            "id": "A",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/易分析性测度测试方法/系统日志完整性",
            "des": "实际运行时间",
            "val": 0
        }

        node_3_3 = self.test_root
        for i in self.test_root.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '易分析性测度测试方法':
                node_3_3 = self.testDocx[int(i)]
                break
        node_3_3_1 = self.test_root
        for i in node_3_3.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '系统日志完整性':
                node_3_3_1 = self.testDocx[int(i)]
                break

        node_3_3_1_1 = self.test_root
        for i in node_3_3_1.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '审计追踪所需日志条数':
                node_3_3_1_1 = self.testDocx[int(i)]
                break
        node_3_3_1_2 = self.test_root
        for i in node_3_3_1.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '系统记录日志条数':
                node_3_3_1_2 = self.testDocx[int(i)]
                break

        if self.check([node_3_3, node_3_3_1, node_3_3_1_1, node_3_3_1_2]):
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

        table_node_3_3_1_1 = self.testDocx[int(node_3_3_1_1.son_id[0])].text
        table_node_3_3_1_2 = self.testDocx[int(node_3_3_1_2.son_id[0])].text
        table_node_3_3_1_1 = eval(table_node_3_3_1_1)
        table_node_3_3_1_2 = eval(table_node_3_3_1_2)
        count = 0
        res = 0

        for i in range(1, len(table_node_3_3_1_1)):
            for j in range(1, len(table_node_3_3_1_2)):
                if table_node_3_3_1_1[i][2] == table_node_3_3_1_2[j][2]:
                    A += int(table_node_3_3_1_2[i][4])
                    B += int(table_node_3_3_1_1[i][4])
                    res += int(table_node_3_3_1_2[i][4]) / int(table_node_3_3_1_1[i][4])
                    count += 1
                    break

        A_dict['val'] = A
        B_dict['val'] = B
        # print('A = {0}, B = {1}'.format(A, B))
        # print('res = {}'.format(res / count))
        if A and B:
            res_dict = {
                "val": res / count,
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

    def man2s(self):
        A, B = 0, 0
        B_dict = {
            "id": "B",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/易分析性测度测试方法/诊断功能有效性/已实现的诊断功能",
            "des": "3.3.2.2 表格中的诊断功能名称集合大小",
            "val": 0
        }
        A_dict = {
            "id": "A",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/易分析性测度测试方法/诊断功能有效性/对原因分析有效的诊断功能",
            "des": "3.3.2.1 表格中原因分析是否有效为“是”的诊断功能名称集合大小",
            "val": 0
        }

        node_3_3 = self.test_root
        for i in self.test_root.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '易分析性测度测试方法':
                node_3_3 = self.testDocx[int(i)]
                break
        node_3_3_2 = self.test_root
        for i in node_3_3.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '诊断功能有效性':
                node_3_3_2 = self.testDocx[int(i)]
                break

        node_3_3_2_1 = self.test_root
        for i in node_3_3_2.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '对原因分析有效的诊断功能':
                node_3_3_2_1 = self.testDocx[int(i)]
                break
        node_3_3_2_2 = self.test_root
        for i in node_3_3_2.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '已实现的诊断功能':
                node_3_3_2_2 = self.testDocx[int(i)]
                break

        if self.check([node_3_3, node_3_3_2, node_3_3_2_2, node_3_3_2_1]):
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

        table_node_3_3_2_1 = self.testDocx[int(node_3_3_2_1.son_id[0])].text
        table_node_3_3_2_2 = self.testDocx[int(node_3_3_2_2.son_id[0])].text
        table_node_3_3_2_1 = eval(table_node_3_3_2_1)
        table_node_3_3_2_2 = eval(table_node_3_3_2_2)

        for i in range(1, len(table_node_3_3_2_1)):
            if table_node_3_3_2_1[i][4] == '是':
                A += 1
        B = len(table_node_3_3_2_2) - 1

        A_dict['val'] = A
        B_dict['val'] = B

        # print('A = {0}, B = {1}'.format(A, B))
        if A and B:
            res_dict = {
                "val": A/B,
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

    def man3s(self):
        A, B = 0, 0
        B_dict = {
            "id": "B",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/易分析性测度测试方法/诊断功能充分性/已实现的诊断功能",
            "des": "软件需求规格说明书中规定的需要遵循编码规则的模块最少数",
            "val": 0
        }
        A_dict = {
            "id": "A",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/易分析性测度测试方法/诊断功能充分性/需要实现的诊断功能",
            "des": "软件符合编码规则的模块数",
            "val": 0
        }

        node_3_3 = self.test_root
        for i in self.test_root.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '易分析性测度测试方法':
                node_3_3 = self.testDocx[int(i)]
                break
        node_3_3_3 = self.test_root
        for i in node_3_3.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '诊断功能充分性':
                node_3_3_3 = self.testDocx[int(i)]
                break

        node_3_3_3_1 = self.test_root
        for i in node_3_3_3.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '已实现的诊断功能':
                node_3_3_3_1 = self.testDocx[int(i)]
                break
        node_3_3_3_2 = self.test_root
        for i in node_3_3_3.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '需要实现的诊断功能':
                node_3_3_3_2 = self.testDocx[int(i)]
                break

        if self.check([node_3_3, node_3_3_3, node_3_3_3_1, node_3_3_3_2]):
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

        table_node_3_3_3_1 = self.testDocx[int(node_3_3_3_1.son_id[0])].text
        table_node_3_3_3_2 = self.testDocx[int(node_3_3_3_2.son_id[0])].text
        table_node_3_3_3_1 = eval(table_node_3_3_3_1)
        table_node_3_3_3_2 = eval(table_node_3_3_3_2)

        for i in range(1, len(table_node_3_3_3_1)):
            if table_node_3_3_3_1[i][4] == '是':
                A += 1
        B = len(table_node_3_3_3_2) - 1

        A_dict['val'] = A
        B_dict['val'] = B

        # print('A = {0}, B = {1}'.format(A, B))
        if A and B:
            res_dict = {
                "val": A/B,
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

    def mmd1g(self):
        A, B = 0, 0
        node_3_4 = self.test_root
        for i in self.test_root.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '易修改性测度':
                node_3_4 = self.testDocx[int(i)]
                break
        node_3_4_1 = self.test_root
        for i in node_3_4.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '修改的效率':
                node_3_4_1 = self.testDocx[int(i)]
                break

        node_3_4_1_1 = self.test_root
        for i in node_3_4_1.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '修改的实际耗时':
                node_3_4_1_1 = self.testDocx[int(i)]
                break
        node_3_4_1_2 = self.test_root
        for i in node_3_4_1.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '修改的预计耗时':
                node_3_4_1_2 = self.testDocx[int(i)]
                break

        if self.check([node_3_4, node_3_4_1, node_3_4_1_1, node_3_4_1_2]):
            B_dict = {
                "id": "B",
                "type": "维护性测试方案",
                "source": self.MaintainTestFile + "/3.4.1.1",
                "des": "预期修改的实际时间",
                "val": 0
            }
            A_dict = {
                "id": "A",
                "type": "维护性测试方案",
                "source": self.MaintainTestFile + "/3.4.1.2",
                "des": "特定的时间周期内测量指定类型修改的实际时间",
                "val": 0
            }
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

        table_node_3_4_1_1 = self.testDocx[int(node_3_4_1_1.son_id[0])].text
        table_node_3_4_1_2 = self.testDocx[int(node_3_4_1_2.son_id[0])].text

        table_node_3_4_1_1 = eval(table_node_3_4_1_1)
        table_node_3_4_1_2 = eval(table_node_3_4_1_2)
        res = 0
        count = len(table_node_3_4_1_1) - 1
        sublist = []
        try:
            for i in range(1, len(table_node_3_4_1_1)):
                res += int(table_node_3_4_1_2[i][5]) / int(table_node_3_4_1_1[i][5])
                B_dict = {
                    "id": "B",
                    "type": "维护性测试方案",
                    "source": self.MaintainTestFile + "/测试方法/易修改性测度/修改的效率/修改的实际耗时",
                    "des": "预期修改的实际时间",
                    "val": int(table_node_3_4_1_1[i][5])
                }
                A_dict = {
                    "id": "A",
                    "type": "维护性测试方案",
                    "source": self.MaintainTestFile + "/测试方法/易修改性测度/修改的效率/修改的预计耗时",
                    "des": "特定的时间周期内测量指定类型修改的实际时间",
                    "val": int(table_node_3_4_1_2[i][5])
                }
                sublist.append(A_dict)
                sublist.append(B_dict)
            res_dict = {
                "val": int(res) / count,
                "sub": sublist
            }
            return res_dict
        except:
            B_dict = {
                "id": "B",
                "type": "维护性测试方案",
                "source": self.MaintainTestFile + "/测试方法/易修改性测度/修改的效率/修改的实际耗时",
                "des": "预期修改的实际时间",
                "val": 0
            }
            A_dict = {
                "id": "A",
                "type": "维护性测试方案",
                "source": self.MaintainTestFile + "/测试方法/易修改性测度/修改的效率/修改的预计耗时",
                "des": "特定的时间周期内测量指定类型修改的实际时间",
                "val": 0
            }

            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

    def mmd2g(self):
        A, B = 0, 0
        node_3_4 = self.test_root
        for i in self.test_root.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '易修改性测度':
                node_3_4 = self.testDocx[int(i)]
                break
        node_3_4_2 = self.test_root
        for i in node_3_4.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '修改的正确性':
                node_3_4_2 = self.testDocx[int(i)]
                break

        node_3_4_2_1 = self.test_root
        for i in node_3_4_2.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '导致事故发生的修改':
                node_3_4_2_1 = self.testDocx[int(i)]
                break
        node_3_4_2_2 = self.test_root
        for i in node_3_4_2.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '实施的修改':
                node_3_4_2_2 = self.testDocx[int(i)]
                break

        if self.check([node_3_4, node_3_4_2, node_3_4_2_1, node_3_4_2_2]):
            B_dict = {
                "id": "B",
                "type": "维护性测试方案",
                "source": self.MaintainTestFile + "/测试方法/易修改性测度/修改的正确性/实施的修改",
                "des": "预期修改的实际时间",
                "val": 0
            }
            A_dict = {
                "id": "A",
                "type": "维护性测试方案",
                "source": self.MaintainTestFile + "/测试方法/易修改性测度/修改的正确性/导致事故发生的修改",
                "des": "特定的时间周期内测量指定类型修改的实际时间",
                "val": 0
            }
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

        table_node_3_4_2_1 = self.testDocx[int(node_3_4_2_1.son_id[0])].text
        table_node_3_4_2_2 = self.testDocx[int(node_3_4_2_2.son_id[0])].text

        table_node_3_4_2_1 = eval(table_node_3_4_2_1)
        table_node_3_4_2_2 = eval(table_node_3_4_2_2)
        B = len(table_node_3_4_2_2) - 1
        # print('B = ',B)

        B_dict = {
            "id": "B",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/易修改性测度/修改的正确性/实施的修改",
            "des": "预期修改的实际时间",
            "val": 0
        }
        A_dict = {
            "id": "A",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/易修改性测度/修改的正确性/导致事故发生的修改",
            "des": "特定的时间周期内测量指定类型修改的实际时间",
            "val": 0
        }
        # print('1111')
        try:
            # print('2222')
            for i in range(len(table_node_3_4_2_1)):
                if table_node_3_4_2_1[i][5] == '是':
                    A += 1
            A_dict['val'] = A
            B_dict['val'] = B

            print('A=', A)
            res_dict = {
                "val": A/B,
                "sub": [A_dict, B_dict]
            }
            return res_dict
        except:
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

    def mmd3s(self):
        A, B = 0, 0
        node_3_4 = self.test_root
        for i in self.test_root.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '易修改性测度':
                node_3_4 = self.testDocx[int(i)]
                break
        node_3_4_2 = self.test_root
        for i in node_3_4.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '修改的能力':
                node_3_4_2 = self.testDocx[int(i)]
                break

        node_3_4_2_1 = self.test_root
        for i in node_3_4_2.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '实际修改的项目':
                node_3_4_2_1 = self.testDocx[int(i)]
                break
        node_3_4_2_2 = self.test_root
        for i in node_3_4_2.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '要求修改的项目':
                node_3_4_2_2 = self.testDocx[int(i)]
                break

        if self.check([node_3_4, node_3_4_2, node_3_4_2_1, node_3_4_2_2]):
            B_dict = {
                "id": "B",
                "type": "维护性测试方案",
                "source": self.MaintainTestFile + "/测试方法/易修改性测度/修改的能力/要求修改的项目",
                "des": "预期修改的实际时间",
                "val": 0
            }
            A_dict = {
                "id": "A",
                "type": "维护性测试方案",
                "source": self.MaintainTestFile + "/测试方法/易修改性测度/修改的能力/实际修改的项目",
                "des": "特定的时间周期内测量指定类型修改的实际时间",
                "val": 0
            }
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

        table_node_3_4_2_1 = self.testDocx[int(node_3_4_2_1.son_id[0])].text
        table_node_3_4_2_2 = self.testDocx[int(node_3_4_2_2.son_id[0])].text

        table_node_3_4_2_1 = eval(table_node_3_4_2_1)
        table_node_3_4_2_2 = eval(table_node_3_4_2_2)
        A = len(table_node_3_4_2_1) - 1
        B = len(table_node_3_4_2_2) - 1

        B_dict = {
            "id": "B",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/易修改性测度/修改的能力/要求修改的项目",
            "des": "预期修改的实际时间",
            "val": B
        }
        A_dict = {
            "id": "A",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/易修改性测度/修改的能力/实际修改的项目",
            "des": "特定的时间周期内测量指定类型修改的实际时间",
            "val": A
        }

        res_dict = {
            "val": A/B,
            "sub": [A_dict, B_dict]
        }
        return res_dict

    def mte1g(self):
        A, B = 0, 0
        node_3_5 = self.test_root
        for i in self.test_root.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '易测试性测度':
                node_3_5 = self.testDocx[int(i)]
                break
        node_3_5_1 = self.test_root
        for i in node_3_5.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '测试功能的完整性':
                node_3_5_1 = self.testDocx[int(i)]
                break

        node_3_5_1_1 = self.test_root
        for i in node_3_5_1.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '已实现的测试功能':
                node_3_5_1_1 = self.testDocx[int(i)]
                break
        node_3_5_1_2 = self.test_root
        for i in node_3_5_1.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '需要的测试功能':
                node_3_5_1_2 = self.testDocx[int(i)]
                break

        if self.check([node_3_5, node_3_5_1, node_3_5_1_1, node_3_5_1_2]):
            B_dict = {
                "id": "B",
                "type": "维护性测试方案",
                "source": self.MaintainTestFile + "/测试方法/易测试性测度/测试功能的完整性/需要的测试功能",
                "des": "预期修改的实际时间",
                "val": 0
            }
            A_dict = {
                "id": "A",
                "type": "维护性测试方案",
                "source": self.MaintainTestFile + "/测试方法/易测试性测度/测试功能的完整性/已实现的测试功能",
                "des": "特定的时间周期内测量指定类型修改的实际时间",
                "val": 0
            }
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

        table_node_3_5_1_1 = self.testDocx[int(node_3_5_1_1.son_id[0])].text
        table_node_3_5_1_2 = self.testDocx[int(node_3_5_1_2.son_id[0])].text

        table_node_3_5_1_1 = eval(table_node_3_5_1_1)
        table_node_3_5_1_2 = eval(table_node_3_5_1_2)

        for i in range(len(table_node_3_5_1_1)):
            if table_node_3_5_1_1[i][5] == '是':
                A += 1
        B = len(table_node_3_5_1_2) - 1

        B_dict = {
            "id": "B",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/易测试性测度/测试功能的完整性/需要的测试功能",
            "des": "预期修改的实际时间",
            "val": B
        }
        A_dict = {
            "id": "A",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/易测试性测度/测试功能的完整性/已实现的测试功能",
            "des": "特定的时间周期内测量指定类型修改的实际时间",
            "val": A
        }

        res_dict = {
            "val": A/B,
            "sub": [A_dict, B_dict]
        }
        return res_dict

    def mte2s(self):
        A, B = 0, 0
        node_3_5 = self.test_root
        for i in self.test_root.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '易测试性测度':
                node_3_5 = self.testDocx[int(i)]
                break
        node_3_5_2 = self.test_root
        for i in node_3_5.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '测试独立性':
                node_3_5_2 = self.testDocx[int(i)]
                break

        node_3_5_2_1 = self.test_root
        for i in node_3_5_2.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '能被桩模拟的测试':
                node_3_5_2_1 = self.testDocx[int(i)]
                break
        node_3_5_2_2 = self.test_root
        for i in node_3_5_2.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '依赖其他系统的测试':
                node_3_5_2_2 = self.testDocx[int(i)]
                break

        if self.check([node_3_5, node_3_5_2, node_3_5_2_1, node_3_5_2_2]):
            B_dict = {
                "id": "B",
                "type": "维护性测试方案",
                "source": self.MaintainTestFile + "/测试方法/易测试性测度/测试独立性/依赖其他系统的测试",
                "des": "预期修改的实际时间",
                "val": 0
            }
            A_dict = {
                "id": "A",
                "type": "维护性测试方案",
                "source": self.MaintainTestFile + "/测试方法/易测试性测度/测试独立性/能被桩模拟的测试",
                "des": "特定的时间周期内测量指定类型修改的实际时间",
                "val": 0
            }
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

        table_node_3_5_2_1 = self.testDocx[int(node_3_5_2_1.son_id[0])].text
        table_node_3_5_2_2 = self.testDocx[int(node_3_5_2_2.son_id[0])].text

        table_node_3_5_2_1 = eval(table_node_3_5_2_1)
        table_node_3_5_2_2 = eval(table_node_3_5_2_2)

        for i in range(len(table_node_3_5_2_1)):
            if table_node_3_5_2_1[i][5] == '是':
                A += 1
        B = len(table_node_3_5_2_2) - 1

        B_dict = {
            "id": "B",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/易测试性测度/测试独立性/依赖其他系统的测试",
            "des": "预期修改的实际时间",
            "val": B
        }
        A_dict = {
            "id": "A",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/易测试性测度/测试独立性/能被桩模拟的测试",
            "des": "特定的时间周期内测量指定类型修改的实际时间",
            "val": A
        }

        res_dict = {
            "val": A/B,
            "sub": [A_dict, B_dict]
        }
        return res_dict

    def mte3s(self):
        A, B = 0, 0
        node_3_5 = self.test_root
        for i in self.test_root.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '易测试性测度':
                node_3_5 = self.testDocx[int(i)]
                break
        node_3_5_3 = self.test_root
        for i in node_3_5.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '测试的重启动性':
                node_3_5_3 = self.testDocx[int(i)]
                break

        node_3_5_3_1 = self.test_root
        for i in node_3_5_3.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '执行中能暂停并重启的测试用例':
                node_3_5_3_1 = self.testDocx[int(i)]
                break
        node_3_5_3_2 = self.test_root
        for i in node_3_5_3.son_id:
            # 先找到 组件间的耦合度
            if self.testDocx[int(i)].text == '执行中能暂停的测试用例':
                node_3_5_3_2 = self.testDocx[int(i)]
                break

        if self.check([node_3_5, node_3_5_3, node_3_5_3_1, node_3_5_3_2]):
            B_dict = {
                "id": "B",
                "type": "维护性测试方案",
                "source": self.MaintainTestFile + "/测试方法/易测试性测度/测试的重启性/执行中能暂停的测试用例",
                "des": "预期修改的实际时间",
                "val": 0
            }
            A_dict = {
                "id": "A",
                "type": "维护性测试方案",
                "source": self.MaintainTestFile + "/测试方法/易测试性测度/测试的重启性/执行中能暂停并重启的测试用例",
                "des": "特定的时间周期内测量指定类型修改的实际时间",
                "val": 0
            }
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict

        table_node_3_5_3_1 = self.testDocx[int(node_3_5_3_1.son_id[0])].text
        table_node_3_5_3_2 = self.testDocx[int(node_3_5_3_2.son_id[0])].text

        table_node_3_5_3_1 = eval(table_node_3_5_3_1)
        table_node_3_5_3_2 = eval(table_node_3_5_3_2)

        for i in range(len(table_node_3_5_3_1)):
            if table_node_3_5_3_1[i][4] == '是':
                A += 1
        B = len(table_node_3_5_3_2) - 1
        print('A = ', A)
        print('B= ', B)
        B_dict = {
            "id": "B",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/易测试性测度/测试的重启性/执行中能暂停的测试用例",
            "des": "预期修改的实际时间",
            "val": B
        }
        A_dict = {
            "id": "A",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/易测试性测度/测试的重启性/执行中能暂停并重启的测试用例",
            "des": "特定的时间周期内测量指定类型修改的实际时间",
            "val": A
        }

        res_dict = {
            "val": A/B,
            "sub": [A_dict, B_dict]
        }
        return res_dict

    def mcl1g(self):
        A, B = 0, 0
        user_demand_id = ''
        node = self.manual_root
        flag = False

        for i in self.manual_root.son_id:
            # print('son_id = ', i)
            # print(self.manualDocx[int(i)].table_info)
            # print(self.manualDocx[int(i)].table_info[0][1])
            if self.manualDocx[int(i)].table_info[2][1].__contains__("维护通过的用例数"):
                user_demand_id = self.manualDocx[int(i)].table_info[0][1]
                node = self.manualDocx[i]
                flag = True
                break

        filtered_rows = self.trace[self.trace['软件需求编号'] == user_demand_id]
        test_case_id = filtered_rows.iloc[0]['tc_id']
        filtered_rows = self.testcases[self.testcases['用例编号'] == test_case_id]
        # print(filtered_rows)
        A, B = 0, 0
        for i in filtered_rows.index:
            B += 1
            if filtered_rows['执行结果'][i] == '通过':
                A += 1

        B_dict = {
            "id": "B",
            "type": "软件需求规格说明书",
            "source": self.UserManualFile + "/系统需求说明/非功能需求/维护性",
            "des": "软件需求规格说明书中规定的维护后通过的用例数阈值",
            "val": B
        }
        A_dict = {
            "id": "A",
            "type": "维护性测试方案",
            "source": self.MaintainTestFile + "/测试方法/维护的依从性计算方法/维护的依从性",
            "des": "维护后通过的用例数",
            "val": A
        }

        if A and B:
            res_dict = {
                "val": int(A)/int(B),
                "sub": [A_dict, B_dict]
            }
            return res_dict
        else:
            res_dict = {
                "val": 0,
                "sub": [A_dict, B_dict]
            }
            return res_dict


def getTreelist(Filepath):
    (Folderpath, Filename) = os.path.split(Filepath)  # 分离文件夹路径和文件名.后缀
    # 为了后续copy文件改为zip并解压，删除已经存在的同名文件
    (souce_name, souce_suffix) = os.path.splitext(Filename)  # 分离文件名和后缀
    if souce_suffix != '.docx':
        exit('请选择docx文档')
    extract_path = Folderpath + '/' + souce_name + '_copy'
    zip_path = Folderpath + '/' + souce_name + '_copy.zip'
    Delete_Copy(extract_path, zip_path)  # 删除该路径下的文件
    print('docx文件路径:', Filepath)

    extract_path, zip_path, Folderpath, souce_name = Get_xmlmkdir(Filepath)  # 将docx的克隆文件解压成含xml文件夹
    Tag_order = get_order(extract_path)
    TreeList = Parsing_docx(Tag_order, Filepath, extract_path)
    return TreeList


def get_all_metrix(fileslist):
    global reliablityfile
    reliablityfile = fileslist[4].split('/')[-1]
    rTreeList = getTreelist(fileslist[4])

    global designfile
    designfile = fileslist[3].split('/')[-1]
    dTreeList = getTreelist(fileslist[3])

    global testfile
    testfile = fileslist[2].split('/')[-1]

    global reqfile
    reqfile = fileslist[0].split('/')[-1]
    TreeList = getTreelist(fileslist[0])
    req = get_requirements(TreeList)
    interface = get_interface(TreeList)
    global tracefile
    tracefile = fileslist[1].split('/')[-1]
    trace = get_metric(fileslist[1])
    testcases = get_testcase(fileslist[2])

    fcp1gs = fcp1g(req, trace)
    fcr1gs = fcr1g(req, trace, testcases)
    fap1gs = fap1g(req, trace, testcases)
    fap2gs = fap2g(fap1g(req, trace, testcases)["list"])
    fcl1gs = fcl1g(req, trace, testcases)
    ptb1gs = ptb1g(req, trace, testcases)
    ptb2gs = ptb2g(req, ptb1g(req, trace, testcases)["val"])
    ptb3gs = ptb3g(req, trace, testcases)
    ptb4gs = ptb4g(req, ptb3g(req, trace, testcases)["val"])
    ptb5gs = ptb5g(req, trace, testcases)
    pru1gs = pru1g(req, trace, testcases)
    pru2gs = pru2g(req, trace, testcases)
    pru3gs = pru3g(req, trace, testcases)
    pru4ss = pru4s(req, trace, testcases)
    pca1gs = pca1g(req, trace, testcases)
    pca2gs = pca2g(req, trace, testcases)
    pca3ss = pca3s(req, trace, testcases)
    pcl1gs = pcl1g(req, trace, testcases)
    cco1gs = cco1g(req, trace, testcases)
    cin1gs = cin1g(interface, trace, testcases)
    cin2gs = cin2g(interface, trace, testcases)
    cin3ss = cin3s(interface, trace, testcases)
    ccl1gs = ccl1g(interface, req, trace, testcases)
    uap1gs = uap1g(req, trace, testcases)
    uap2gs = uap2g(req, trace, testcases)
    uap3gs = uap3g(req, trace, testcases)
    ule1gs = ule1g(req, trace, testcases)
    ule3ss = ule3s(req, trace, testcases)
    ule4ss = ule4s(req, trace, testcases)
    uop1gs = uop1g(req, trace, testcases)
    uop2gs = uop2g(req, trace, testcases)
    uop3ss = uop3s(req, trace, testcases)
    uop5ss = uop5s(req, trace, testcases)
    uop6ss = uop6s(req, trace, testcases)
    uop7ss = uop7s(req, trace, testcases)
    uop8ss = uop8s(req, trace, testcases)
    uop9ss = uop9s(req, trace, testcases)
    uep1gs = uep1g(req, trace, testcases)
    uep2ss = uep2s(req, trace, testcases)
    uep3ss = uep3s(req, trace, testcases)
    uac1gs = uac1g(req, trace, testcases)
    uac2ss = uac2s(req, trace, testcases)
    ucl1gs = ucl1g(req, trace, testcases)
    rma1gs = rma1g(rTreeList)
    rma2gs = rma2g(rTreeList)
    rma3gs = rma3g(rTreeList)
    rma4ss = rma4s(rTreeList)
    rav1gs = rav1g(rTreeList)
    rav2gs = rav2g(rTreeList)
    rft1gs = rft1g(rTreeList)
    rft2ss = rft2s(rTreeList)
    rft3ss = rft3s(rTreeList)
    rre1gs = rre1g(rTreeList)
    rre2ss = rre2s(rTreeList)
    ule2ss = ule2s(dTreeList, req, trace, testcases)
    uop4ss = uop4s(dTreeList, req, trace, testcases)
    uin1ss = uin1s(dTreeList, req, trace, testcases)
    rcl1gs = rcl1g(req, trace, testcases)

    sco1gs = sco1g(req, trace, testcases)
    sco2gs = sco2g(req, trace, testcases)
    sco3ss = sco3s(req, trace, testcases)
    sin1gs = sin1g(req, trace, testcases)
    sin2gs = sin2g(req, trace, testcases)
    sno1gs = sno1g(req, trace, testcases)
    sac1gs = sac1g(req, trace, testcases)
    sac2ss = sac2s(req, trace, testcases)
    sau1gs = sau1g(req, trace, testcases)
    sau2ss = sau2s(req, trace, testcases)
    scl1gs = scl1g(req, trace, testcases)
    pcp1gs = pcp1g(req, trace, testcases)
    pad1gs = pad1g(req, trace, testcases)
    pad2gs = pad2g(req, trace, testcases)
    pad3ss = pad3s(req, trace, testcases)
    pin1gs = pin1g(req, trace, testcases)
    pin2gs = pin2g(req, trace, testcases)
    pre1gs = pre1g(req, trace, testcases)
    pre2ss = pre2s(req, trace, testcases)
    pre3ss = pre3s(req, trace, testcases)
    pre4ss = pre4s(req, trace, testcases)
    ftmp1 = getTreelist(fileslist[0])
    ftmp2 = getTreelist(fileslist[5])

    maintainTestFileEtract = MaintainTestFileExtract(fileslist[0], fileslist[5], fileslist[1], fileslist[2], ftmp1, ftmp2, fileslist[3])
    mmo1gs = maintainTestFileEtract.mmo1g()
    mmo2ss = maintainTestFileEtract.mmo2s()
    mre1gs = maintainTestFileEtract.mre1g()
    mre2ss = maintainTestFileEtract.mre2s()

    man1gs = maintainTestFileEtract.man1g()
    man2ss = maintainTestFileEtract.man2s()
    man3ss = maintainTestFileEtract.man3s()
    mmd1gs = maintainTestFileEtract.mmd1g()
    mmd2gs = maintainTestFileEtract.mmd2g()
    mmd3ss = maintainTestFileEtract.mmd3s()
    mte1gs = maintainTestFileEtract.mte1g()
    mte2ss = maintainTestFileEtract.mte2s()
    mte3ss = maintainTestFileEtract.mte3s()
    mcl1gs = maintainTestFileEtract.mcl1g()
    mcl1gs = maintainTestFileEtract.mcl1g()

    metrix_25000 = {'功能性': {'功能完备性': {'功能覆盖率': fcp1gs}, '功能正确性': {'功能正确性': fcr1gs},
                            '功能适合性': {'使用目标的功能适合性': fap1gs, '系统的功能适合性': fap2gs},
                            '功能性的依从性': {'功能性的依从性': fcl1gs}},
                    '性能效率': {'时间特性': {'平均响应时间': ptb1gs, '响应时间的充分性': ptb2gs, '平均周转时间': ptb3gs, '周转时间充分性': ptb4gs,
                                      '平均吞吐量': ptb5gs},
                             '资源利用性': {'处理器平均占用率': pru1gs, '内存平均占用率': pru2gs, 'I/O设备平均占用率': pru3gs, '带宽占用率': pru4ss},
                             '容量': {'事务处理容量': pca1gs, '用户访问量': pca2gs, '用户访问增长的充分性': pca3ss},
                             '性能效率的依从性': {'性能效率的依从性': pcl1gs}},
                    '兼容性': {'共存性': {'与其他产品的共存性': cco1gs},
                            '互操作性': {'数据格式可交换性': cin1gs, '数据交换协议充分性': cin2gs, '外部接口充分性': cin3ss},
                            '兼容性的依从性': {'兼容性的依从性': ccl1gs},
                            },
                    '易用性': {'可辨识性': {'描述的完整性': uap1gs, '演示覆盖率': uap2gs, '入口点的自描述性': uap3gs},
                            '易学性': {'用户指导完整性': ule1gs, '输入字段的默认值': ule2ss, '差错信息的易理解性': ule3ss, '用户界面的自解释性': ule4ss},
                            '易操作性': {'操作一致性': uop1gs, '消息的明确性': uop2gs, '功能的易定制性': uop3ss, '用户界面的易定制性': uop4ss,
                                     '监视能力': uop5ss, '撤销操作能力': uop6ss, '信息分类的易理解性': uop7ss, '外观一致性': uop8ss,
                                     '输入设备的支持性': uop9ss},
                            '用户差错防御性': {'抵御误操作': uep1gs, '用户输入差错纠正率': uep2ss, '用户差错易恢复性': uep3ss},
                            '用户界面舒适性': {'用户界面外观舒适性': uin1ss}, '易访问性': {'特殊群体的易访问性': uac1gs, '支持的语种充分性': uac2ss},
                            '易用性的依从性': {'易用性的依从性': ucl1gs}},
                    '可靠性': {'成熟性': {'故障修复率': rma1gs, '平均失效间隔时间(MTBF)': rma2gs, '周期失效率': rma3gs, '测试覆盖率': rma4ss},
                            '可用性': {'系统可用性': rav1gs, '平均宕机时间': rav2gs},
                            '容错性': {'避免失效率': rft1gs, '组件的冗余度': rft2ss, '平均故障通告时间': rft3ss},
                            '易恢复性': {'平均恢复时间': rre1gs, '数据备份完整性': rre2ss},
                            '可靠性的依从性': {'可靠性的依从性': rcl1gs}},
                    '信息安全性': {'保密性': {'访问控制性': sco1gs, '数据加密正确性': sco2gs, '加密算法的强度': sco3ss},
                              '完整性': {'数据完整性': sin1gs, '内部数据抗讹误性': sin2gs, '缓冲区溢出防止率': 1},
                              '抗抵赖性': {'数字签名使用率': sno1gs}, '可核查性': {'用户审计跟踪的完整性': sac1gs, '系统日志保留满足度': sac2ss},
                              '真实性': {'鉴别机制的充分性': sau1gs, '鉴别规则的符合性': sau2ss},
                              '信息安全性的依从性': {'信息安全性的依从性': scl1gs}},
                    '维护性': {'模块化': {'组件间的耦合度': 1, '圈复杂度的充分性': mmo2ss},
                            '可重用性': {'资产的可重用性': mre1gs, '编码规则符合性': mre2ss},
                            '易分析性': {'系统日志完整性': man1gs, '诊断功能有效性': man2ss, '诊断功能充分性': man3ss},
                            '易修改性': {'修改的效率': mmd1gs, '修改的正确性': mmd2gs, '修改的能力': mmd3ss},
                            '易测试性': {'测试功能的完整性': mte1gs, '测试独立性': mte2ss, '测试的重启动性': mte3ss},
                            '维护性的依从性': {'维护性的依从性': mcl1gs}
                            },
                    '可移植性': {'适应性': {'硬件环境的适应性': pad1gs, '系统软件环境的适应性': pad2gs, '运营环境的适应性': pad3ss},
                             '易安装性': {'安装的时间效率': pin1gs, '安装的灵活性': pin2gs},
                             '易替换性': {'使用相似性': pre1gs, '产品质量等价性': pre2ss, '功能的包容性': pre3ss,
                                      '数据复用/导入能力': pre4ss},
                             '可移植性的依从性': {'可移植性的依从性': pcp1gs}}
                    }
    return metrix_25000


if __name__ == '__main__':
    filepath = [
        "D:/Code/test_project/TMMi文档/软件需求规格说明书-update.docx",
        "D:/Code/test_project/TMMi文档/需求跟踪矩阵.xlsx",
        "D:/Code/test_project/TMMi文档/验收测试用例.xlsx",
        "D:/Code/test_project/TMMi文档/软件详细设计说明书.docx",
        "D:/Code/test_project/TMMi文档/可靠性测试方案.docx",
        "D:/Code/test_project/TMMi文档/维护性测试方案.docx"
    ]
    test = get_all_metrix(filepath)
    print(test)
    with open('./metrix.json', 'w', encoding="utf-8") as f:
        json.dump(get_all_metrix(filepath), f, indent=4, ensure_ascii=False)

    # filepath = [
    #     "D:/Workspace/TMMi文档模板/软件需求规格说明书-update.docx",
    #     "D:/Workspace/TMMi文档模板/需求跟踪矩阵.xlsx",
    #     "D:/Workspace/TMMi文档模板/验收测试用例.xlsx",
    #     "D:/Workspace/TMMi文档模板/软件详细设计说明书.docx",
    #     "D:/Workspace/TMMi文档模板/可靠性测试方案.docx",
    #     "D:/Workspace/TMMi文档模板/维护性测试方案.docx"
    # ]
    # maintainTestFileEtract = MaintainTestFileExtract(filepath[0],
    #                                                  filepath[5],
    #                                                  filepath[1],
    #                                                  filepath[2],
    #                                                  getTreelist(filepath[0]),
    #                                                  getTreelist(filepath[5]),
    #                                                  filepath[3]
    #                                                  )
    # print("mmo1g = {0}".format(maintainTestFileEtract.mmo1g()))
    # print("mmo2s = {0}".format(maintainTestFileEtract.mmo2s()))
    # print("mre1g = {0}".format(maintainTestFileEtract.mre1g()))
    # print("mre2s = {0}".format(maintainTestFileEtract.mre2s()))
    # print("man1g = {0}".format(maintainTestFileEtract.man1g()))
    # print("man2s = {0}".format(maintainTestFileEtract.man2s()))
    # print("man3s = {0}".format(maintainTestFileEtract.man3s()))
    # print("mmd1g = {0}".format(maintainTestFileEtract.mmd1g()))
    # print("mmd2g = {0}".format(maintainTestFileEtract.mmd2g()))
    # print("mmd3s = {0}".format(maintainTestFileEtract.mmd3s()))
    # print("mte1g = {0}".format(maintainTestFileEtract.mte1g()))
    # print("mte2s = {0}".format(maintainTestFileEtract.mte2s()))
    # print("mte3s = {0}".format(maintainTestFileEtract.mte3s()))
    # print("mcl1g = {0}".format(maintainTestFileEtract.mcl1g()))
    # print("mcl1g = {0}".format(maintainTestFileEtract.mcl1g()))
